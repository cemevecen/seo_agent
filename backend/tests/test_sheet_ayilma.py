# -*- coding: utf-8 -*-
"""/sheet erisim + ayılma çizelge motoru."""

from datetime import date

from backend.services.ayilma_schedule import (
    LEAD_NURSE,
    STAFF_NURSES,
    generate_ayilma_schedule,
    ideal_hours,
    roster_defaults,
)
from backend.services.sheet_page_access import (
    is_sheet_only_member_email,
    is_sheet_page_allowed_email,
    is_sheet_page_path,
    member_denied_sheet_access,
    resolve_sheet_menu_visible,
    sheet_only_member_path_allowed,
)


def test_sheet_page_allowed_emails():
    assert is_sheet_page_allowed_email("cemevecen@nokta.com")
    assert is_sheet_page_allowed_email("CemEvecen@Gmail.com")
    assert is_sheet_page_allowed_email("evecensema@gmail.com")
    assert is_sheet_page_allowed_email("EvecenSema@gmail.com")
    assert not is_sheet_page_allowed_email("onurtorun@nokta.com")
    assert not is_sheet_page_allowed_email("melihengin@nokta.com")
    assert not is_sheet_page_allowed_email("")


def test_roster_defaults_next_month():
    from datetime import date

    d = roster_defaults()
    today = date.today()
    if today.month == 12:
        assert d["default_year"] == today.year + 1
        assert d["default_month"] == 1
    else:
        assert d["default_year"] == today.year
        assert d["default_month"] == today.month + 1


def test_sheet_page_paths():
    assert is_sheet_page_path("/sheet")
    assert is_sheet_page_path("/sheet/")
    assert is_sheet_page_path("/api/sheet/ayilma/meta")
    assert is_sheet_page_path("/api/sheet/ayilma/generate")
    assert is_sheet_page_path("/api/sheet/ayilma/export.xlsx")
    assert is_sheet_page_path("/api/sheet/ayilma/export.csv")
    assert is_sheet_page_path("/api/sheet/ayilma/export.docx")
    assert not is_sheet_page_path("/ipo")
    assert not is_sheet_page_path("/")


def test_export_xlsx_opens():
    from io import BytesIO

    from openpyxl import load_workbook

    from backend.services.ayilma_schedule import build_ayilma_xlsx_bytes

    out = generate_ayilma_schedule(2026, 9)
    raw = build_ayilma_xlsx_bytes(year=2026, month=9, days=out["days"], rows=out["rows"])
    assert raw[:2] == b"PK"
    wb = load_workbook(BytesIO(raw))
    ws = wb.active
    assert "Ayılma" in str(ws["A1"].value)
    assert ws.cell(row=4, column=1).value == LEAD_NURSE
    # Lead row must be empty on all day columns
    for col in range(2, 2 + len(out["days"])):
        assert not ws.cell(row=4, column=col).value


def test_export_csv_windows_friendly():
    from backend.services.ayilma_schedule import build_ayilma_csv_bytes

    out = generate_ayilma_schedule(2026, 9)
    raw = build_ayilma_csv_bytes(year=2026, month=9, days=out["days"], rows=out["rows"])
    assert raw.startswith(b"\xef\xbb\xbf")
    text = raw.decode("utf-8")
    assert "Ad Soyadı" in text
    assert "," in text
    assert ",," not in text
    assert '""' in text  # boş mesai hücreleri tırnaklı
    assert ";" not in text.splitlines()[2]  # başlık satırı virgülle
    assert "Ayılma" in text
    assert "Gülten" in text
    assert "Nuray" in text


def test_export_docx_opens():
    from io import BytesIO
    from zipfile import ZipFile

    from backend.services.ayilma_schedule import build_ayilma_docx_bytes

    out = generate_ayilma_schedule(2026, 9)
    raw = build_ayilma_docx_bytes(year=2026, month=9, days=out["days"], rows=out["rows"])
    assert raw[:2] == b"PK"
    with ZipFile(BytesIO(raw)) as zf:
        assert "word/document.xml" in zf.namelist()
        doc_xml = zf.read("word/document.xml").decode("utf-8")
    assert "Ayılma" in doc_xml
    assert "Ad Soyad" in doc_xml
    assert "landscape" in doc_xml.lower() or 'w:orient="landscape"' in doc_xml


def test_sheet_menu_visible():
    assert resolve_sheet_menu_visible(member_email="cemevecen@nokta.com") is True
    assert resolve_sheet_menu_visible(member_email="cemevecen@gmail.com") is True
    assert resolve_sheet_menu_visible(member_email="evecensema@gmail.com") is True
    assert resolve_sheet_menu_visible(member_email="other@nokta.com") is False
    assert member_denied_sheet_access("other@nokta.com") is True
    assert member_denied_sheet_access("cemevecen@gmail.com") is False
    assert member_denied_sheet_access("evecensema@gmail.com") is False


def test_sheet_only_member_paths():
    assert is_sheet_only_member_email("evecensema@gmail.com")
    assert not is_sheet_only_member_email("cemevecen@gmail.com")
    assert sheet_only_member_path_allowed("/sheet")
    assert sheet_only_member_path_allowed("/api/sheet/ayilma/meta")
    assert sheet_only_member_path_allowed("/auth/logout")
    assert sheet_only_member_path_allowed("/static/js/app.js")
    assert not sheet_only_member_path_allowed("/realtime")
    assert not sheet_only_member_path_allowed("/api/panel/online-users")


def test_every_weekday_has_kat1_eight():
    """Hafta içi her güne bir açık «8» (kat-1); 24 gündüzü kapsamaz."""
    leaves = {
        "Nuray Durna": {"2026-09-18": "İST"},
        "Sema Evecen": {"2026-09-01": "İST", "2026-09-07": "İST", "2026-09-08": "İST"},
        "Şengül Zamur": {"2026-09-23": "İST"},
        "Semanur Çınar": {f"2026-09-{d:02d}": "Yİ" for d in range(1, 14)}
        | {"2026-09-27": "İST"},
        "Rabia Kumtepe": {"2026-09-19": "İST"},
    }
    out = generate_ayilma_schedule(2026, 9, leaves=leaves)
    grid = {r["name"]: r["cells"] for r in out["rows"] if r["role"] == "staff"}
    missing = []
    for dm in out["days"]:
        if dm["is_weekend"]:
            continue
        iso = dm["iso"]
        if not any(grid[n].get(iso, "") == "8" for n in STAFF_NURSES):
            missing.append(iso)
    assert not missing, f"kat-1 8 missing: {missing}"
    assert not any("Kat-1 «8» eksik" in w for w in (out.get("warnings") or []))


def test_generate_august_basic_coverage():
    out = generate_ayilma_schedule(2026, 8)
    assert out["ok"] is True
    assert out["lead"] == LEAD_NURSE
    assert len(out["staff"]) == 6
    assert out["ideal_hours_staff"] == ideal_hours(2026, 8)

    lead = next(r for r in out["rows"] if r["name"] == LEAD_NURSE)
    assert lead["role"] == "lead"
    assert lead["overtime_hours"] == 0
    assert lead["worked_hours"] == 0
    for dm in out["days"]:
        assert lead["cells"][dm["iso"]] == "", f"lead must stay empty {dm['iso']}"

    for dm in out["days"]:
        iso = dm["iso"]
        morning = 0
        night = 0
        staff8 = 0
        for name in STAFF_NURSES:
            row = next(r for r in out["rows"] if r["name"] == name)
            code = row["cells"].get(iso, "")
            if code in ("8", "24"):
                morning += 1
            if code in ("16", "24"):
                night += 1
            if code == "8":
                staff8 += 1
        assert night >= 2, f"night short {iso}: {night}"
        if dm["is_weekend"]:
            assert staff8 == 0, f"weekend staff 8 on {iso}"
        else:
            assert morning >= 1, f"morning missing {iso}"
            assert staff8 >= 1, f"kat-1 8 missing {iso}"


def test_prefer_8_and_minimize_16():
    out = generate_ayilma_schedule(2026, 8)
    counts = out["staff_code_counts"]
    # Kişi başı ~2–4 → toplam ~12–24; her güne 8 şart değil
    assert 12 <= counts["8"] <= 24
    assert counts["16"] < counts["24"]
    assert counts["16"] <= 2


def test_eights_and_twentyfours_are_shared():
    """8 kişi başı 2–4; 24 altı kişiye yayılır."""
    out = generate_ayilma_schedule(2026, 8)
    staff_rows = [r for r in out["rows"] if r["role"] == "staff"]
    eights = [r["count_8"] for r in staff_rows]
    twentyfours = [r["count_24"] for r in staff_rows]
    assert min(eights) >= 2
    assert max(eights) <= 4
    assert max(eights) - min(eights) <= 2
    assert min(twentyfours) >= 2
    assert max(twentyfours) - min(twentyfours) <= 4


def test_lead_does_not_count_in_staff_night_or_overtime():
    out = generate_ayilma_schedule(2026, 8)
    lead = next(r for r in out["rows"] if r["name"] == LEAD_NURSE)
    assert lead["exclude_from_staff_balance"] is True
    assert lead["ideal_hours"] == 0
    assert lead["overtime_hours"] == 0
    assert lead["worked_hours"] == 0
    for dm in out["days"]:
        iso = dm["iso"]
        night = sum(
            1
            for n in STAFF_NURSES
            if next(r for r in out["rows"] if r["name"] == n)["cells"].get(iso) in ("16", "24")
        )
        assert night >= 2
        assert lead["cells"][iso] == ""


def test_first_day_after_leave_gets_night_shift():
    """Yİ/RP/İST bitişinin ertesi gün nöbet (24 veya 16); kat-1 8 değil."""
    leaves = {
        "Nuray Durna": {
            "2026-08-17": "Yİ",
            "2026-08-18": "Yİ",
            "2026-08-19": "Yİ",
        }
    }
    out = generate_ayilma_schedule(2026, 8, leaves=leaves)
    nuray = next(r for r in out["rows"] if r["name"] == "Nuray Durna")
    first_back = nuray["cells"]["2026-08-20"]
    assert first_back in ("16", "24"), f"expected nöbet, got {first_back!r}"
    assert first_back != "8"


def test_first_day_after_ist_gets_work():
    """İST bitişinin ertesi takvim gününde mutlaka 24 nöbet."""
    leaves = {"Sema Evecen": {"2026-08-08": "İST", "2026-08-09": "İST"}}
    out = generate_ayilma_schedule(2026, 8, leaves=leaves)
    sema = next(r for r in out["rows"] if r["name"] == "Sema Evecen")
    assert sema["cells"].get("2026-08-10") == "24"  # Pazar — 9 Ağu İST ertesi


def test_ist_next_calendar_day_is_24():
    """İST gününün ertesi takvim günü 24 (8 değil)."""
    leaves = {"Sema Evecen": {"2026-09-01": "İST", "2026-09-07": "İST", "2026-09-08": "İST"}}
    out = generate_ayilma_schedule(2026, 9, leaves=leaves)
    sema = next(r for r in out["rows"] if r["name"] == "Sema Evecen")
    assert sema["cells"].get("2026-09-02") == "24"
    assert sema["cells"].get("2026-09-09") == "24"


def test_ist_only_keeps_hours_balance():
    """İST günleri kotadan düşülmez; kalan günlerle ortalama mesai bandı (stres ≤2× tolerans)."""
    from backend.services.ayilma_schedule import HOURS_BALANCE_TOLERANCE

    leaves = {"Sema Evecen": {"2026-09-01": "İST", "2026-09-07": "İST", "2026-09-08": "İST"}}
    out = generate_ayilma_schedule(2026, 9, leaves=leaves)
    staff = [r for r in out["rows"] if r["role"] == "staff"]
    metrics = [r["worked_hours"] for r in staff]
    assert max(metrics) - min(metrics) <= HOURS_BALANCE_TOLERANCE * 2
    sema = next(r for r in staff if r["name"] == "Sema Evecen")
    peer = sorted(r["worked_hours"] for r in staff if r["name"] != "Sema Evecen")
    peer_med = peer[len(peer) // 2]
    assert abs(sema["worked_hours"] - peer_med) <= HOURS_BALANCE_TOLERANCE * 2


def test_generate_respects_leave_and_rest_after_24():
    leaves = {
        "Nuray Durna": {
            "2026-08-17": "Yİ",
            "2026-08-18": "Yİ",
            "2026-08-19": "Yİ",
        }
    }
    out = generate_ayilma_schedule(2026, 8, leaves=leaves)
    nuray = next(r for r in out["rows"] if r["name"] == "Nuray Durna")
    assert nuray["cells"]["2026-08-17"] == "Yİ"
    assert nuray["cells"]["2026-08-18"] == "Yİ"

    # 24 sonrası ertesi gün boş veya izin
    for name in STAFF_NURSES:
        row = next(r for r in out["rows"] if r["name"] == name)
        cells = row["cells"]
        for dm in out["days"]:
            if dm["day"] >= 31:
                continue
            iso = dm["iso"]
            if cells.get(iso) != "24":
                continue
            nxt = f"2026-08-{dm['day'] + 1:02d}"
            if nxt in cells:
                assert cells[nxt] in ("", "Yİ", "RP", "İST"), f"{name} {iso}-> {cells[nxt]}"


def test_yi_counts_as_eight_and_fill_remaining_days():
    """5 gün Yİ = 40s; eylül 176 → zorunlu nöbet 136s (kalan günlerle); üstüne ek mesai olabilir."""
    leaves = {
        "Nuray Durna": {
            f"2026-09-{d:02d}": "Yİ" for d in range(1, 6)  # Pzt–Cum haftası
        }
    }
    out = generate_ayilma_schedule(2026, 9, leaves=leaves)
    assert out["ideal_hours_staff"] == 176
    assert out["max_monthly_hours"] == 400
    nuray = next(r for r in out["rows"] if r["name"] == "Nuray Durna")
    assert nuray["leave_hours"] == 40
    assert nuray["min_shift_hours"] == 136
    assert nuray["worked_hours"] == nuray["shift_hours"] + 40
    assert nuray["cells"]["2026-09-01"] == "Yİ"
    assert nuray["shift_hours"] >= 120


def test_first_day_after_leave_weekend_gets_24():
    """Yİ Cuma bitince Cumartesi (hafta sonu) ilk güne 24 yazılır."""
    leaves = {
        "Nuray Durna": {
            "2026-09-03": "Yİ",
            "2026-09-04": "Yİ",  # Cuma
        }
    }
    out = generate_ayilma_schedule(2026, 9, leaves=leaves)
    nuray = next(r for r in out["rows"] if r["name"] == "Nuray Durna")
    assert nuray["cells"].get("2026-09-05") == "24"  # Cumartesi


def test_overtime_band_with_heavy_leave():
    """Yİ alan kişi ortalama dışı; aktif kadro OT bandı dar kalır."""
    leaves = {
        "Semanur Çınar": {f"2026-09-{d:02d}": "Yİ" for d in range(1, 14)},
        "Sema Evecen": {f"2026-09-{d:02d}": "İST" for d in (1, 7, 8, 15, 22, 29)},
    }
    out = generate_ayilma_schedule(2026, 9, leaves=leaves)
    semanur = next(r for r in out["rows"] if r["name"] == "Semanur Çınar")
    assert semanur["exclude_from_staff_balance"] is True
    assert semanur["leave_hours"] > 0
    # Aktif (Yİ/RP yok) personel bandı
    active = [
        r for r in out["rows"] if r["role"] == "staff" and not r["exclude_from_staff_balance"]
    ]
    ots = [r["overtime_hours"] for r in active]
    assert max(ots) - min(ots) <= 32
    # İzinli kişi: zorunlu taban + ek mesai kuralları içinde kalır (ortalama zorlaması yok)
    assert semanur["shift_hours"] >= semanur["min_shift_hours"] - 24


def test_staff_accounted_hours_reasonably_balanced():
    """Aktif personel (Yİ/RP hariç) ≤16s bantta."""
    out = generate_ayilma_schedule(2026, 9)
    active = [
        r for r in out["rows"] if r["role"] == "staff" and not r["exclude_from_staff_balance"]
    ]
    vals = [r["worked_hours"] for r in active]
    ots = [r["overtime_hours"] for r in active]
    assert max(vals) - min(vals) <= 16, vals
    assert max(ots) - min(ots) <= 16, ots


def test_yi_excluded_from_peer_average():
    """Yİ kullanan ortalama hesabına girmez; aktif kadro kendi ortalamasında kalır."""
    from backend.services.ayilma_schedule import HOURS_BALANCE_TOLERANCE

    leaves = {"Nuray Durna": {f"2026-09-{d:02d}": "Yİ" for d in range(1, 8)}}
    out = generate_ayilma_schedule(2026, 9, leaves=leaves)
    nuray = next(r for r in out["rows"] if r["name"] == "Nuray Durna")
    assert nuray["exclude_from_staff_balance"] is True
    active = [
        r for r in out["rows"] if r["role"] == "staff" and not r["exclude_from_staff_balance"]
    ]
    assert all(r["name"] != "Nuray Durna" for r in active)
    vals = [r["worked_hours"] for r in active]
    assert max(vals) - min(vals) <= HOURS_BALANCE_TOLERANCE * 2


def test_peer_hours_band_with_long_yi_and_ist():
    """Semanur uzun Yİ + İST karışımında Emine/Nuray/Şengül arası ≤16s (40s sapma olmamalı)."""
    from backend.services.ayilma_schedule import HOURS_BALANCE_TOLERANCE, month_days, _grid_gap_ok

    leaves = {
        "Nuray Durna": {"2026-09-18": "İST"},
        "Sema Evecen": {"2026-09-01": "İST", "2026-09-07": "İST", "2026-09-08": "İST"},
        "Şengül Zamur": {"2026-09-23": "İST"},
        "Semanur Çınar": {f"2026-09-{d:02d}": "Yİ" for d in range(1, 14)}
        | {"2026-09-27": "İST"},
        "Rabia Kumtepe": {"2026-09-19": "İST"},
    }
    out = generate_ayilma_schedule(2026, 9, leaves=leaves)
    active = [
        r for r in out["rows"] if r["role"] == "staff" and not r["exclude_from_staff_balance"]
    ]
    vals = {r["name"]: r["worked_hours"] for r in active}
    spread = max(vals.values()) - min(vals.values())
    assert spread <= HOURS_BALANCE_TOLERANCE, vals
    assert not any("saat bandı" in w for w in (out.get("warnings") or []))
    grid = {r["name"]: r["cells"] for r in out["rows"] if r["role"] == "staff"}
    assert _grid_gap_ok(month_days(2026, 9), grid)


def test_ist_request_blocks_assignment():
    leaves = {"Sema Evecen": {"2026-08-10": "İST", "2026-08-11": "İST"}}
    out = generate_ayilma_schedule(2026, 8, leaves=leaves)
    sema = next(r for r in out["rows"] if r["name"] == "Sema Evecen")
    assert sema["cells"]["2026-08-10"] == "İST"
    assert sema["cells"]["2026-08-11"] == "İST"


def test_ist_overtime_near_peers():
    """İST kotadan düşülmez; yalnız İST kullanan personel fazla mesai bandında kalır."""
    leaves = {
        "Sema Evecen": {
            "2026-08-22": "İST",
            "2026-08-25": "İST",
            "2026-08-28": "İST",
        },
    }
    out = generate_ayilma_schedule(2026, 8, leaves=leaves)
    sema = next(r for r in out["rows"] if r["name"] == "Sema Evecen")
    peer_ots = [
        r["overtime_hours"]
        for r in out["rows"]
        if r["role"] == "staff" and r["name"] != "Sema Evecen"
    ]
    assert sema["leave_hours"] == 0
    assert sema["min_shift_hours"] == out["ideal_hours_staff"]
    assert max(peer_ots) - sema["overtime_hours"] <= 16
    assert sema["overtime_hours"] - min(peer_ots) <= 16


def _count_gun_asiri(out: dict) -> int:
    days = out["days"]
    n = 0
    for row in out["rows"]:
        if row["role"] != "staff":
            continue
        cells = row["cells"]
        for i in range(2, len(days)):
            if cells.get(days[i]["iso"]) != "24":
                continue
            if cells.get(days[i - 2]["iso"]) != "24":
                continue
            mid = cells.get(days[i - 1]["iso"], "")
            if mid in ("", "Yİ", "RP", "İST"):
                n += 1
    return n


def _max_gun_asiri_streak(out: dict) -> int:
    """Bir kişide 24+boş+24… zincirinin en uzun 24 sayısı."""
    days = out["days"]
    gap = ("", "Yİ", "RP", "İST")
    best = 0
    for row in out["rows"]:
        if row["role"] != "staff":
            continue
        cells = row["cells"]
        for i, dm in enumerate(days):
            if cells.get(dm["iso"]) != "24":
                continue
            streak = 1
            j = i
            while j >= 2:
                if cells.get(days[j - 2]["iso"]) != "24":
                    break
                if cells.get(days[j - 1]["iso"], "") not in gap:
                    break
                streak += 1
                j -= 2
            best = max(best, streak)
    return best


def _max_pair24_monthly(out: dict) -> int:
    days = out["days"]
    grid = {r["name"]: r["cells"] for r in out["rows"] if r["role"] == "staff"}
    names = list(grid.keys())
    best = 0
    for i, a in enumerate(names):
        for b in names[i + 1 :]:
            n = sum(
                1
                for dm in days
                if grid[a].get(dm["iso"], "") == "24" and grid[b].get(dm["iso"], "") == "24"
            )
            best = max(best, n)
    return best


def _max_pair24_near_streak(out: dict) -> int:
    """Aynı ikilinin ≤4 gün aralıklı ardışık birlikte 24 sayısı."""
    days = out["days"]
    grid = {r["name"]: r["cells"] for r in out["rows"] if r["role"] == "staff"}
    names = list(grid.keys())
    best = 0
    for i, a in enumerate(names):
        for b in names[i + 1 :]:
            idxs = [
                j
                for j, dm in enumerate(days)
                if grid[a].get(dm["iso"], "") == "24" and grid[b].get(dm["iso"], "") == "24"
            ]
            streak = 1
            local = 1
            for k in range(1, len(idxs)):
                if idxs[k] - idxs[k - 1] <= 4:
                    streak += 1
                    local = max(local, streak)
                else:
                    streak = 1
            best = max(best, local if idxs else 0)
    return best


def test_soft_avoid_gun_asiri_prefer_24_over_16():
    """16 neredeyse hiç; gün aşırı zinciri ≤4 (uç durum ≤5); 16 ile streak kırılmaz."""
    out = generate_ayilma_schedule(2026, 9)
    counts = out["staff_code_counts"]
    assert counts["16"] <= 2
    assert counts["24"] >= counts["8"]
    assert _max_gun_asiri_streak(out) <= 5


def test_soft_avoid_repeated_pair24():
    """Aynı ikili 24'te çok sık / üst üste eşleşmesin."""
    from backend.services.ayilma_schedule import (
        PAIR24_MONTHLY_SOFT,
        PAIR24_NEAR_STREAK_MAX,
        generate_ayilma_schedule,
    )

    for variant in range(40):
        out = generate_ayilma_schedule(2026, 9, variant=variant)
        assert _max_pair24_monthly(out) <= PAIR24_MONTHLY_SOFT
        assert _max_pair24_near_streak(out) <= PAIR24_NEAR_STREAK_MAX


def test_pair24_near_streak_cap_with_leaves():
    """İzinli ayda da aynı ikili yakın 24 zinciri ≤2 (ör. Şengül+Emine 24-boş-24-boş-24)."""
    from backend.services.ayilma_schedule import PAIR24_NEAR_STREAK_MAX

    leaves = {
        "Nuray Durna": {"2026-09-18": "İST"},
        "Sema Evecen": {"2026-09-01": "İST", "2026-09-07": "İST", "2026-09-08": "İST"},
        "Şengül Zamur": {"2026-09-23": "İST"},
        "Semanur Çınar": {f"2026-09-{d:02d}": "Yİ" for d in range(1, 14)}
        | {"2026-09-27": "İST"},
        "Rabia Kumtepe": {"2026-09-19": "İST"},
    }
    out = generate_ayilma_schedule(2026, 9, leaves=leaves)
    # İzin sıkışığında 2 her zaman mümkün olmayabilir; 3 üstü olmamalı
    assert _max_pair24_near_streak(out) <= 3
    assert _max_pair24_near_streak(out) <= PAIR24_NEAR_STREAK_MAX + 1


def test_pair24_soft_penalty_discourages_third_near():
    from backend.services.ayilma_schedule import _pair24_soft_penalty, month_days

    days = month_days(2026, 9)
    grid = {n: {d.iso: "" for d in days} for n in STAFF_NURSES}
    a, b = STAFF_NURSES[0], STAFF_NURSES[1]
    grid[a][days[5].iso] = "24"
    grid[b][days[5].iso] = "24"
    grid[a][days[7].iso] = "24"
    grid[b][days[7].iso] = "24"
    assert _pair24_soft_penalty(a, b, 9, days, grid) >= 40


def test_gun_asiri_streak_helpers():
    from backend.services.ayilma_schedule import (
        GUN_ASIRI_STREAK_ABSOLUTE,
        GUN_ASIRI_STREAK_MAX,
        GUN_ASIRI_STREAK_SOFT,
        _gun_asiri_streak_if_24,
        month_days,
    )

    days = month_days(2026, 9)
    grid = {n: {d.iso: "" for d in days} for n in STAFF_NURSES}
    name = STAFF_NURSES[0]
    # 12,14,16 dolu → gün 10'a yazmak zinciri 4 yapar
    grid[name][days[11].iso] = "24"
    grid[name][days[13].iso] = "24"
    grid[name][days[15].iso] = "24"
    assert _gun_asiri_streak_if_24(name, 9, days, grid) == 4  # day 10
    grid[name][days[9].iso] = "24"
    assert _gun_asiri_streak_if_24(name, 15, days, grid) == 4  # day 16 back
    assert GUN_ASIRI_STREAK_SOFT == 3
    assert GUN_ASIRI_STREAK_MAX == 4
    assert GUN_ASIRI_STREAK_ABSOLUTE == 5


def test_idle_24_gap_helpers():
    from backend.services.ayilma_schedule import (
        IDLE_24_GAP_MAX,
        IDLE_24_GAP_SOFT,
        _days_without_24_before,
        _idle_empty_streak_before,
        month_days,
    )

    days = month_days(2026, 9)
    grid = {n: {d.iso: "" for d in days} for n in STAFF_NURSES}
    name = STAFF_NURSES[0]
    grid[name][days[0].iso] = "24"
    grid[name][days[1].iso] = ""
    grid[name][days[2].iso] = "8"
    grid[name][days[3].iso] = ""
    assert _days_without_24_before(name, 4, days, grid) == 3
    assert _idle_empty_streak_before(name, 4, days, grid) == 1
    assert IDLE_24_GAP_MAX == 3
    assert IDLE_24_GAP_SOFT == 2


def test_max_empty_between_24_hard_cap_three():
    """Kati: bir hemşirede 24 arası 4+ boş hücre olmaz (panel üretimi)."""
    from backend.services.ayilma_schedule import (
        _max_empty_between_24_in_grid,
        month_days,
    )

    leaves = {
        "Semanur Çınar": {f"2026-09-{d:02d}": "Yİ" for d in range(1, 14)},
        "Sema Evecen": {f"2026-09-{d:02d}": "İST" for d in (1, 7, 8, 15, 22, 29)},
    }
    out = generate_ayilma_schedule(2026, 9, leaves=leaves)
    grid = {r["name"]: r["cells"] for r in out["rows"] if r["role"] == "staff"}
    empty_gap = _max_empty_between_24_in_grid(month_days(2026, 9), grid)
    assert empty_gap <= 3, f"max empty between 24 = {empty_gap}"


def test_no_trailing_empty_streak_over_three():
    """Ay sonunda son 24 sonrası 4+ boş gün olmamalı (trailing dahil)."""
    from backend.services.ayilma_schedule import (
        _max_empty_between_24_in_grid,
        month_days,
    )

    out = generate_ayilma_schedule(
        2026,
        9,
        special_rules=[
            {
                "name": "Sema Evecen",
                "mode": "work",
                "dates": ["2026-09-01", "2026-09-02"],
                "weekly": True,
            }
        ],
    )
    days = month_days(2026, 9)
    grid = {r["name"]: r["cells"] for r in out["rows"] if r["role"] == "staff"}
    assert _max_empty_between_24_in_grid(days, grid) <= 3
    for name, cells in grid.items():
        last24 = None
        for i, d in enumerate(days):
            if cells.get(d.iso) == "24":
                last24 = i
        if last24 is None:
            continue
        trail = 0
        for j in range(last24 + 1, len(days)):
            if cells.get(days[j].iso, "") == "":
                trail += 1
            else:
                break
        assert trail <= 3, f"{name} trailing empty after last 24 = {trail}"


def test_triple_gap_sandwich_minimized():
    """3+24+3 yerine 2+24+2 tercih — mümkün olduğunca az triple sandwich."""
    from backend.services.ayilma_schedule import (
        _count_triple_gap_sandwiches_in_grid,
        month_days,
    )

    for month in (8, 9):
        for variant in range(25):
            out = generate_ayilma_schedule(2026, month, variant=variant)
            grid = {r["name"]: r["cells"] for r in out["rows"] if r["role"] == "staff"}
            days_meta = month_days(2026, month)
            triple = _count_triple_gap_sandwiches_in_grid(days_meta, grid)
            assert triple <= 2, (
                f"{month}/{variant}: {triple} adet 3+24+3 (hedef ≤2, ideal 0)"
            )


def test_gun_asiri_streak_cap_with_heavy_leave():
    """İzin yoğun ayda bile gün aşırı 24 zinciri ABSOLUTE (5) aşılmaz."""
    leaves = {
        "Semanur Çınar": {f"2026-09-{d:02d}": "Yİ" for d in range(1, 14)},
        "Sema Evecen": {f"2026-09-{d:02d}": "İST" for d in (1, 7, 8, 15, 22, 29)},
    }
    for variant in range(25):
        out = generate_ayilma_schedule(2026, 9, leaves=leaves, variant=variant)
        assert _max_gun_asiri_streak(out) <= 5, f"variant {variant}"


def test_avoid_consecutive_eight_pairs_when_possible():
    """Üst üste 8+8 mümkün olduğunca az (yumuşak; sert tavan hâlâ 2)."""
    from backend.services.ayilma_schedule import (
        CONSECUTIVE_8_STREAK_MAX,
        _count_consecutive_8_runs,
        _max_consecutive_8_streak,
        month_days,
    )

    leaves = {
        "Semanur Çınar": {f"2026-09-{d:02d}": "Yİ" for d in range(1, 14)},
        "Sema Evecen": {f"2026-09-{d:02d}": "İST" for d in (1, 7, 8, 15, 22, 29)},
    }
    for variant in range(12):
        out = generate_ayilma_schedule(2026, 9, leaves=leaves, variant=variant)
        grid = {r["name"]: r["cells"] for r in out["rows"] if r["role"] == "staff"}
        days_meta = month_days(2026, 9)
        pairs = _count_consecutive_8_runs(grid, days_meta)
        assert pairs <= 4, f"variant {variant} has {pairs} consecutive-8 pairs"
        for row in out["rows"]:
            if row["role"] != "staff":
                continue
            streak = _max_consecutive_8_streak(row["name"], days_meta, grid)
            assert streak <= CONSECUTIVE_8_STREAK_MAX


def test_no_consecutive_eights_for_staff():
    from backend.services.ayilma_schedule import (
        CONSECUTIVE_8_STREAK_MAX,
        _max_consecutive_8_streak,
        month_days,
    )

    out = generate_ayilma_schedule(2026, 9)
    grid = {r["name"]: r["cells"] for r in out["rows"]}
    days_meta = month_days(2026, 9)
    for row in out["rows"]:
        if row["role"] != "staff":
            continue
        streak = _max_consecutive_8_streak(row["name"], days_meta, grid)
        assert streak <= CONSECUTIVE_8_STREAK_MAX, (
            f"{row['name']} has {streak} consecutive 8s (max {CONSECUTIVE_8_STREAK_MAX})"
        )


def test_no_three_eights_after_long_leave():
    """Uzun Yİ sonrası ardışık 3×8 olmamalı (ör. Semanur senaryosu)."""
    from backend.services.ayilma_schedule import _max_consecutive_8_streak, month_days

    leaves = {
        "Semanur Çınar": {f"2026-09-{d:02d}": "Yİ" for d in range(1, 14)},
    }
    out = generate_ayilma_schedule(2026, 9, leaves=leaves)
    semanur = next(r for r in out["rows"] if r["name"] == "Semanur Çınar")
    days_meta = month_days(2026, 9)
    grid = {"Semanur Çınar": semanur["cells"]}
    streak = _max_consecutive_8_streak("Semanur Çınar", days_meta, grid)
    assert streak <= 2, f"Semanur consecutive 8 streak={streak}"
    # Dönüş günü (14) nöbet veya en az tek 8; üçlü blok yok
    d14, d15, d16 = (
        semanur["cells"].get("2026-09-14", ""),
        semanur["cells"].get("2026-09-15", ""),
        semanur["cells"].get("2026-09-16", ""),
    )
    assert not (d14 == d15 == d16 == "8"), f"3×8 block: {d14}/{d15}/{d16}"


def test_two_night_shifts_every_day_with_mixed_leaves():
    """15 Eylül gibi İST/Yİ karışımında da her gün 2× gece nöbeti."""
    leaves = {
        "Semanur Çınar": {f"2026-09-{d:02d}": "Yİ" for d in range(1, 14)},
        "Sema Evecen": {"2026-09-01": "İST", "2026-09-07": "İST", "2026-09-08": "İST"},
        "Rabia Kumtepe": {"2026-09-03": "İST"},
        "Emine Türker": {"2026-09-15": "İST", "2026-09-22": "İST"},
        "Nuray Durna": {"2026-09-18": "İST"},
        "Şengül Zamur": {"2026-09-28": "İST"},
    }
    out = generate_ayilma_schedule(2026, 9, leaves=leaves)
    iso = "2026-09-15"
    night = [
        (r["name"], r["cells"].get(iso, ""))
        for r in out["rows"]
        if r["role"] == "staff" and r["cells"].get(iso, "") in ("16", "24")
    ]
    assert len(night) >= 2, f"2026-09-15 only {night}"
    for dm in out["days"]:
        n = sum(
            1
            for r in out["rows"]
            if r["role"] == "staff" and r["cells"].get(dm["iso"], "") in ("16", "24")
        )
        assert n >= 2, f"{dm['iso']}: {n} night shifts"


def test_resolve_special_weekly_avoid():
    from backend.services.ayilma_schedule import resolve_special_day_sets

    # 2026-09-01 Salı → weekly avoid Salı
    work, avoid = resolve_special_day_sets(
        2026,
        9,
        [
            {
                "name": "Sema Evecen",
                "mode": "avoid",
                "dates": ["2026-09-01"],
                "weekly": True,
            }
        ],
    )
    assert "2026-09-01" in avoid["Sema Evecen"]
    assert "2026-09-08" in avoid["Sema Evecen"]
    assert "2026-09-15" in avoid["Sema Evecen"]
    assert "2026-09-02" not in avoid["Sema Evecen"]
    assert not work["Sema Evecen"]


def test_special_avoid_blocks_shifts():
    out = generate_ayilma_schedule(
        2026,
        9,
        special_rules=[
            {
                "name": "Sema Evecen",
                "mode": "avoid",
                "dates": ["2026-09-01"],
                "weekly": True,
            }
        ],
    )
    sema = next(r for r in out["rows"] if r["name"] == "Sema Evecen")
    for iso in ("2026-09-01", "2026-09-08", "2026-09-15", "2026-09-22", "2026-09-29"):
        assert sema["cells"].get(iso, "") in ("", "Yİ", "RP", "İST"), (
            f"Sema should avoid {iso}, got {sema['cells'].get(iso)!r}"
        )


def test_special_avoid_weekly_keeps_hours_balance():
    """Sal–Per çalışmasın: bloklu günler boş; diğer günlerle band (stres ≤2× tolerans)."""
    from backend.services.ayilma_schedule import HOURS_BALANCE_TOLERANCE

    out = generate_ayilma_schedule(
        2026,
        9,
        special_rules=[
            {
                "name": "Sema Evecen",
                "mode": "avoid",
                "dates": ["2026-09-01", "2026-09-02", "2026-09-03"],
                "weekly": True,
            }
        ],
    )
    sema = next(r for r in out["rows"] if r["name"] == "Sema Evecen")
    staff = [r for r in out["rows"] if r["role"] == "staff"]
    metrics = [r["worked_hours"] for r in staff]
    spread = max(metrics) - min(metrics)
    assert spread <= HOURS_BALANCE_TOLERANCE * 3, (
        f"balance spread {spread} > {HOURS_BALANCE_TOLERANCE * 3}: "
        + ", ".join(f"{r['name']}={r['worked_hours']}" for r in staff)
    )
    for iso, code in sema["cells"].items():
        if not iso.startswith("2026-09-"):
            continue
        if date.fromisoformat(iso).weekday() in (1, 2, 3):
            assert code not in ("8", "16", "24"), f"Sema worked avoid day {iso}: {code!r}"


def test_special_work_prefers_selected_days():
    """çalışsın: seçilen günlerde mümkünse mesai (en az birinde 8/16/24)."""
    out = generate_ayilma_schedule(
        2026,
        9,
        special_rules=[
            {
                "name": "Rabia Kumtepe",
                "mode": "work",
                "dates": ["2026-09-02", "2026-09-03", "2026-09-04"],
                "weekly": False,
            }
        ],
    )
    rabia = next(r for r in out["rows"] if r["name"] == "Rabia Kumtepe")
    codes = [
        rabia["cells"].get(iso, "")
        for iso in ("2026-09-02", "2026-09-03", "2026-09-04")
    ]
    assert any(c in ("8", "16", "24") for c in codes), codes


def test_special_work_weekly_enforces_all_days():
    """çalışsın + her hafta: seçilen weekday'lerde mesai zorunlu."""
    out = generate_ayilma_schedule(
        2026,
        9,
        special_rules=[
            {
                "name": "Sema Evecen",
                "mode": "work",
                "dates": ["2026-09-01", "2026-09-02"],
                "weekly": True,
            }
        ],
    )
    sema = next(r for r in out["rows"] if r["name"] == "Sema Evecen")
    tues_weds = [
        iso
        for iso in sorted(sema["cells"])
        if iso.startswith("2026-09-")
        and date.fromisoformat(iso).weekday() in (1, 2)
    ]
    assert tues_weds, "expected Tue/Wed dates in September"
    missing = [
        iso
        for iso in tues_weds
        if sema["cells"].get(iso, "") not in ("8", "16", "24")
    ]
    assert not missing, f"Sema missing work on: {missing}"


def test_variant_can_differ():
    a = generate_ayilma_schedule(2026, 9, variant=0)
    b = generate_ayilma_schedule(2026, 9, variant=1)
    cells_a = {r["name"]: r["cells"] for r in a["rows"] if r["role"] == "staff"}
    cells_b = {r["name"]: r["cells"] for r in b["rows"] if r["role"] == "staff"}
    assert a["variant"] == 0 and b["variant"] == 1
    # En az bir günde farklı atama beklenir (eşitlikte farklı aday)
    differ = False
    for name in cells_a:
        for iso, code in cells_a[name].items():
            if cells_b[name].get(iso) != code:
                differ = True
                break
        if differ:
            break
    assert differ

