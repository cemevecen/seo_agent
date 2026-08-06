import io
import json
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook

from backend.database import SessionLocal, init_db
from backend.models import AdReportRow
from backend.services import ad_analytics_store as store

FIXTURE = Path(__file__).resolve().parent / "fixtures" / "ad_sample.csv"


def test_xlsx_header_after_title_rows():
    wb = Workbook()
    ws = wb.active
    ws.append(["Rapor özeti"])
    ws.append(["Dönem", "2025"])
    ws.append([])
    ws.append(
        [
            "Ad Unit",
            "Date",
            "Income Type",
            "Net Revenue",
        ]
    )
    ws.append(["web_unit_1", 45658, "Open Auction", 10.5])
    buf = io.BytesIO()
    wb.save(buf)
    rows = store.parse_xlsx_bytes(buf.getvalue(), filename="dovizcom1_Report_2025.xlsx")
    assert len(rows) == 1
    assert rows[0]["ad_unit"] == "web_unit_1"


def test_xlsx_turkish_headers():
    wb = Workbook()
    ws = wb.active
    ws.append(["Reklam birimi", "Tarih", "Gelir tipi", "Net gelir"])
    ws.append(["web_test", 45658, "Mediation", 3.0])
    buf = io.BytesIO()
    wb.save(buf)
    rows = store.parse_xlsx_bytes(buf.getvalue(), filename="dovizcom2_Report_2026.xlsx")
    assert len(rows) == 1


def test_incremental_append_upserts_same_day():
    init_db()
    d = date(2026, 6, 10)
    serial = _excel_serial(d)
    base = (
        "Ad Unit,Month,Date,Income Type,Ad Request,Matched Request,Impression,Click,"
        "Ad Request Ecpm,Ad Impression Ecpm,CTR,Coverage,Viewability,Net Revenue\n"
        f"web_unit_1,1,{serial},Open Auction,1,1,1,0,0,0,0,0,0,10\n"
    )
    with SessionLocal() as db:
        store.reset_all(db)
        store.import_append_to_stream(
            db,
            base.encode("utf-8"),
            stream_key="doviz:desktop",
            original_filename="gunluk.csv",
        )
        store.import_append_to_stream(
            db,
            base.replace(",10\n", ",15\n").encode("utf-8"),
            stream_key="doviz:desktop",
            original_filename="gunluk2.csv",
        )
        n = db.query(AdReportRow).filter(AdReportRow.report_date == d).count()
        rev = db.query(AdReportRow).filter(AdReportRow.report_date == d).one().net_revenue
    assert n == 1
    assert rev == 15.0


def test_channel_and_surface_from_filename_and_ad_unit():
    rows = store.parse_csv_text(
        "Ad Unit,Month,Date,Income Type,Ad Request,Matched Request,Impression,Click,"
        "Ad Request Ecpm,Ad Impression Ecpm,CTR,Coverage,Viewability,Net Revenue\n"
        "m_doviz_kripto_320x50,45658,45658,Open Auction,1,1,1,0,0,0,0,0,0,1\n"
        "web_doviz_kripto_970x90,45658,45658,Mediation,1,1,1,0,0,0,0,0,0,2\n",
        filename="dovizcom2_Report_2026.xlsx",
    )
    assert len(rows) == 2
    assert rows[0]["channel"] == "dovizcom"
    assert rows[0]["surface"] == "mweb"
    assert rows[1]["surface"] == "web"


def test_revenue_week_anomaly_requires_14_days():
    days = [{"date": f"2026-01-{i:02d}", "net_revenue": 10.0} for i in range(1, 11)]
    out = store._revenue_week_anomaly(days)
    assert out["ok"] is False


def test_revenue_week_anomaly_delta():
    days = [{"date": f"2026-01-{i:02d}", "net_revenue": 10.0 if i <= 7 else (5.0 if i <= 14 else 1.0)} for i in range(1, 21)]
    out = store._revenue_week_anomaly(days)
    assert out["ok"] is True
    assert out["last7_revenue"] == 7.0
    assert out["prev7_revenue"] == 70.0


def test_resolve_compare_range_previous_period():
    start, end = store.resolve_compare_range("2026-01-10", "2026-01-16", "previous_period")
    assert start == "2026-01-03"
    assert end == "2026-01-09"


def test_resolve_compare_range_previous_year():
    start, end = store.resolve_compare_range("2025-03-01", "2025-03-07", "previous_year")
    assert start == "2024-03-01"
    assert end == "2024-03-07"


def test_resolve_compare_range_custom():
    start, end = store.resolve_compare_range(
        "2026-01-01",
        "2026-01-31",
        "custom",
        "2025-06-01",
        "2025-06-30",
    )
    assert start == "2025-06-01"
    assert end == "2025-06-30"


def test_align_by_date_series_calendar_not_index():
    """Karşı dönem daha az satır içerse indeks hizalaması sıfıra düşürür; takvim eşlemesi korur."""
    primary = [
        {"date": "2025-06-01", "net_revenue": 10},
        {"date": "2025-06-02", "net_revenue": 20},
        {"date": "2025-06-03", "net_revenue": 30},
        {"date": "2025-06-04", "net_revenue": 40},
    ]
    compare = [
        {"date": "2024-06-01", "net_revenue": 1},
        {"date": "2024-06-02", "net_revenue": 2},
    ]
    aligned = store.align_by_date_series(
        primary,
        compare,
        "net_revenue",
        mode="previous_year",
        primary_start="2025-06-01",
        compare_start="2024-06-01",
    )
    assert len(aligned) == 4
    assert aligned[0]["compare"] == 1.0
    assert aligned[1]["compare"] == 2.0
    assert aligned[2]["compare"] is None
    assert aligned[3]["compare"] is None
    assert aligned[2]["compare_date"] == "2024-06-03"
    assert aligned[3]["compare_date"] == "2024-06-04"


def test_compute_kpi_deltas():
    deltas = store.compute_kpi_deltas(
        {"net_revenue": 150.0, "impression": 1000},
        {"net_revenue": 100.0, "impression": 800},
    )
    assert deltas["net_revenue"]["pct"] == 50.0
    assert deltas["impression"]["abs"] == 200

    zero_base = store._kpi_delta(50.0, 0.0)
    assert zero_base["pct"] is None
    assert zero_base["abs"] == 50.0


def _excel_serial(d: date) -> int:
    return (d - date(1899, 12, 30)).days


def test_compare_ad_units_param_uses_primary_top_units():
    primary = [{"ad_unit": "unit_a"}, {"ad_unit": "unit_b"}]
    assert store._compare_ad_units_param(primary, None) == "unit_a,unit_b"
    assert store._compare_ad_units_param(primary, "unit_x") == "unit_x"


def test_merge_breakdown_str_keys_and_delta_abs():
    primary = [{"ad_unit": "unit_a", "net_revenue": 100.0, "impression": 10}]
    compare = [{"ad_unit": "unit_a", "net_revenue": 50.0, "impression": 5}]
    merged = store._merge_breakdown(primary, compare, "ad_unit")
    assert merged[0]["net_revenue_compare"] == 50.0
    assert merged[0]["net_revenue_delta_pct"] == 100.0
    assert merged[0]["net_revenue_delta_abs"] == 50.0


def test_query_summary_with_compare():
    init_db()
    text = (
        "Ad Unit,Month,Date,Income Type,Ad Request,Matched Request,Impression,Click,"
        "Ad Request Ecpm,Ad Impression Ecpm,CTR,Coverage,Viewability,Net Revenue\n"
    )
    d1 = date(2026, 1, 5)
    d2 = date(2026, 1, 12)
    d0 = date(2025, 12, 29)
    text += f"unit_a,1,{_excel_serial(d1)},Open Auction,10,10,10,1,0,0,0,0,0,50\n"
    text += f"unit_a,1,{_excel_serial(d2)},Open Auction,10,10,10,1,0,0,0,0,0,30\n"
    text += f"unit_a,1,{_excel_serial(d0)},Open Auction,10,10,10,1,0,0,0,0,0,20\n"
    rows = store.parse_csv_text(text, filename="dovizcom1_Report_2026.xlsx")
    with SessionLocal() as db:
        store.reset_all(db)
        store.import_rows(db, rows)
        summ = store.query_summary(
            db,
            start=d1.isoformat(),
            end=d2.isoformat(),
            compare_mode="previous_period",
        )
        assert "compare" in summ
        assert summ["compare"]["deltas"]["net_revenue"]["compare"] == 20.0
        assert summ["compare"]["deltas"]["net_revenue"]["current"] == 80.0
        db.execute(__import__("sqlalchemy").delete(AdReportRow))
        db.commit()


def test_query_summary_rows_in_range_zero_outside_data():
    init_db()
    d = date(2026, 5, 15)
    text = (
        "Ad Unit,Month,Date,Income Type,Ad Request,Matched Request,Impression,Click,"
        "Ad Request Ecpm,Ad Impression Ecpm,CTR,Coverage,Viewability,Net Revenue\n"
        f"unit_a,1,{_excel_serial(d)},Open Auction,10,10,10,1,0,0,0,0,0,50\n"
    )
    rows = store.parse_csv_text(text, filename="dovizweb1.csv")
    with SessionLocal() as db:
        store.reset_all(db)
        store.import_rows(db, rows)
        inside = store.query_summary(db, start="2026-05-01", end="2026-05-31")
        assert inside["rows_in_range"] == 1
        outside = store.query_summary(db, start="2026-06-01", end="2026-06-16")
        assert outside["rows_in_range"] == 0
        assert outside["kpis"]["ad_request"] == 0
        db.execute(__import__("sqlalchemy").delete(AdReportRow))
        db.commit()


def test_empower_metrics_from_extra_header():
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "Ad Unit",
            "Date",
            "Income Type",
            "Impression",
            "Net Revenue",
            "Empower Pageviews",
        ]
    )
    d = date(2026, 2, 1)
    serial = (d - date(1899, 12, 30)).days
    ws.append(["web_x", serial, "Open Auction", 100, 50.0, 1200])
    buf = io.BytesIO()
    wb.save(buf)
    rows = store.parse_xlsx_bytes(buf.getvalue(), filename="dovizcom1_Report_2026.xlsx")
    assert len(rows) == 1
    extras = json.loads(rows[0]["extra_metrics"])
    assert extras.get("empower_pageview") == 1200.0


def test_kpi_available_omits_missing_empower():
    init_db()
    text = (
        "Ad Unit,Month,Date,Income Type,Ad Request,Matched Request,Impression,Click,"
        "Ad Request Ecpm,Ad Impression Ecpm,CTR,Coverage,Viewability,Net Revenue\n"
        "unit_a,1,45658,Open Auction,10,8,10,1,0,0,0,0,0,50\n"
    )
    rows = store.parse_csv_text(text, filename="dovizcom1_Report_2026.xlsx")
    with SessionLocal() as db:
        store.reset_all(db)
        store.import_rows(db, rows)
        summ = store.query_summary(db)
        assert "empower_pageview" not in summ["kpi_available"]
        assert "net_revenue" in summ["kpi_available"]
        assert "impression" in summ["kpi_available"]
        db.execute(__import__("sqlalchemy").delete(AdReportRow))
        db.commit()


def test_aggregate_ctr_sub_percent_precision():
    init_db()
    text = (
        "Ad Unit,Month,Date,Income Type,Ad Request,Matched Request,Impression,Click,"
        "Ad Request Ecpm,Ad Impression Ecpm,CTR,Coverage,Viewability,Net Revenue\n"
        "unit_a,1,45658,Open Auction,1000,800,193290271,858,0,0,0,25.2,0,50\n"
    )
    rows = store.parse_csv_text(text, filename="dovizcom1_Report_2026.xlsx")
    with SessionLocal() as db:
        store.reset_all(db)
        store.import_rows(db, rows)
        summ = store.query_summary(db)
        assert summ["kpis"]["click"] == 858
        expected_ctr = 858 / 193_290_271 * 100.0
        assert summ["kpis"]["ctr_pct"] == round(expected_ctr, 6)
        assert summ["kpis"]["ctr_pct"] > 0
        assert summ["kpis"]["coverage_pct"] == 25.2
        db.execute(__import__("sqlalchemy").delete(AdReportRow))
        db.commit()


def test_viewability_coverage_percent_scale():
    init_db()
    text = (
        "Ad Unit,Month,Date,Income Type,Ad Request,Matched Request,Impression,Click,"
        "Ad Request Ecpm,Ad Impression Ecpm,CTR,Coverage,Viewability,Net Revenue\n"
        "unit_a,1,45658,Open Auction,100,80,100,1,0,0,0,80,90,50\n"
    )
    rows = store.parse_csv_text(text, filename="dovizcom1_Report_2026.xlsx")
    with SessionLocal() as db:
        store.reset_all(db)
        store.import_rows(db, rows)
        summ = store.query_summary(db)
        assert summ["kpis"]["coverage_pct"] == 80.0
        assert summ["kpis"]["viewability_pct"] == 90.0
        db.execute(__import__("sqlalchemy").delete(AdReportRow))
        db.commit()


def test_parse_csv_and_import():
    init_db()
    text = FIXTURE.read_text(encoding="utf-8")
    rows = store.parse_csv_text(text, filename="doviz_android_test.csv")
    assert len(rows) >= 3
    assert rows[0]["ad_unit"] == "test_sticky_unit"
    assert isinstance(rows[0]["report_date"], date)
    with SessionLocal() as db:
        store.reset_all(db)
        out = store.import_rows(db, rows)
        assert out["inserted"] >= 3
        assert rows[0].get("project") == "doviz"
        assert rows[0].get("branch") == "android"
        summ = store.query_summary(db)
        assert summ["kpis"]["net_revenue"] > 20
        assert len(summ["by_income_type"]) >= 2
        db.execute(__import__("sqlalchemy").delete(AdReportRow))
        db.commit()


def test_build_upload_batch_summary():
    per_file = [
        {"filename": "a.xlsx", "parsed": 100, "inserted": 100, "stream_key": "doviz:web"},
        {"filename": "b.xlsx", "error": "bozuk"},
        {"filename": "c.xlsx", "parsed": 0, "parse_error": "başlık yok"},
        {"filename": "d.xlsx", "parsed": 50, "warning": "dal?", "stream_key": None},
    ]
    s = store.build_upload_batch_summary(per_file)
    assert s["file_count"] == 4
    assert s["ok_count"] == 2
    assert s["failed_count"] == 1
    assert s["empty_count"] == 1
    assert s["has_errors"] is True
    assert s["has_warnings"] is True
    assert s["integrated_rows"] == 150


def test_build_heatmap_calendar():
    days = [
        {"date": "2026-06-01", "net_revenue": 10.0},
        {"date": "2026-06-02", "net_revenue": 20.0},
    ]
    hm = store._build_heatmap_calendar(days)
    assert len(hm) == 2
    assert hm[0]["dow_label"] in store._DOW_LABELS


def test_facets_returns_bounds_and_row_count():
    init_db()
    with SessionLocal() as db:
        out = store.facets(db)
    assert "min_date" in out
    assert "max_date" in out
    assert isinstance(out["total_rows"], int)
    assert isinstance(out["streams"], list)


def test_reset_all_clears_db_and_facets():
    init_db()
    d = date(2026, 6, 10)
    csv_text = (
        "Ad Unit,Month,Date,Income Type,Ad Request,Matched Request,Impression,Click,"
        "Ad Request Ecpm,Ad Impression Ecpm,CTR,Coverage,Viewability,Net Revenue\n"
        f"web_unit_1,1,{_excel_serial(d)},Open Auction,1,1,1,0,0,0,0,0,0,10\n"
    )
    with SessionLocal() as db:
        store.reset_all(db)
        store.import_append_to_stream(
            db,
            csv_text.encode("utf-8"),
            stream_key="doviz:desktop",
            original_filename="dovizweb1.csv",
        )
        warm = store.facets(db)
        assert warm["total_rows"] > 0
        assert len(warm["imports"]) > 0
        store.reset_all(db)
        fresh = store.facets(db)
        assert fresh["total_rows"] == 0
        assert fresh["imports"] == []
        assert fresh["source_files"] == []
        assert store.count_rows(db) == 0


def test_delete_source_file_removes_rows_and_catalog():
    init_db()
    d = date(2026, 6, 10)
    serial = _excel_serial(d)
    csv_a = (
        "Ad Unit,Month,Date,Income Type,Ad Request,Matched Request,Impression,Click,"
        "Ad Request Ecpm,Ad Impression Ecpm,CTR,Coverage,Viewability,Net Revenue\n"
        f"web_unit_a,1,{serial},Open Auction,1,1,1,0,0,0,0,0,0,10\n"
    )
    csv_b = csv_a.replace("web_unit_a", "web_unit_b").replace(",10\n", ",20\n")
    with SessionLocal() as db:
        store.reset_all(db)
        store.import_upload_file(db, csv_a.encode("utf-8"), filename="dovizweb1.csv", commit=True)
        store.import_upload_file(db, csv_b.encode("utf-8"), filename="dovizweb2.csv", commit=True)
        assert store.count_rows(db) == 2
        out = store.delete_source_file(db, "dovizweb1.csv")
        assert out["deleted_rows"] == 1
        assert store.count_rows(db) == 1
        remaining = db.query(AdReportRow).one()
        assert remaining.source_file == "dovizweb2.csv"
        assert remaining.net_revenue == 20.0


def test_delete_overwritten_file_restores_previous_period():
    init_db()
    d = date(2026, 6, 10)
    serial = _excel_serial(d)
    csv = (
        "Ad Unit,Month,Date,Income Type,Ad Request,Matched Request,Impression,Click,"
        "Ad Request Ecpm,Ad Impression Ecpm,CTR,Coverage,Viewability,Net Revenue\n"
        f"web_unit_x,1,{serial},Open Auction,1,1,1,0,0,0,0,0,0,10\n"
    )
    csv2 = csv.replace(",10\n", ",99\n")
    with SessionLocal() as db:
        store.reset_all(db)
        store.import_upload_file(db, csv.encode("utf-8"), filename="dovizweb1.csv", commit=True)
        store.import_upload_file(db, csv2.encode("utf-8"), filename="dovizweb2.csv", commit=True)
        row = db.query(AdReportRow).one()
        assert row.source_file == "dovizweb2.csv"
        assert row.net_revenue == 99.0
        out = store.delete_source_file(db, "dovizweb2.csv")
        assert out["deleted_rows"] == 1
        assert out["restored_rows"] == 1
        restored = db.query(AdReportRow).one()
        assert restored.source_file == "dovizweb1.csv"
        assert restored.net_revenue == 10.0


def test_delete_source_files_bulk_restores_once():
    init_db()
    d = date(2026, 6, 10)
    serial = _excel_serial(d)
    hdr = (
        "Ad Unit,Month,Date,Income Type,Ad Request,Matched Request,Impression,Click,"
        "Ad Request Ecpm,Ad Impression Ecpm,CTR,Coverage,Viewability,Net Revenue\n"
    )
    with SessionLocal() as db:
        store.reset_all(db)
        store.import_upload_file(
            db,
            (hdr + f"u,1,{serial},Open Auction,1,1,1,0,0,0,0,0,0,10\n").encode(),
            filename="dovizweb1.csv",
            commit=True,
        )
        store.import_upload_file(
            db,
            (hdr + f"u,1,{serial},Open Auction,1,1,1,0,0,0,0,0,0,99\n").encode(),
            filename="dovizweb2.csv",
            commit=True,
        )
        store.import_upload_file(
            db,
            (hdr + f"v,1,{serial},Open Auction,1,1,1,0,0,0,0,0,0,5\n").encode(),
            filename="dovizmweb1.csv",
            commit=True,
        )
        out = store.delete_source_files_bulk(db, ["dovizweb2.csv", "dovizmweb1.csv"])
        assert out["deleted_files"] == 2
        assert out["restored_rows"] == 1
        rows = db.query(AdReportRow).all()
        assert len(rows) == 1
        assert rows[0].source_file == "dovizweb1.csv"
        assert rows[0].net_revenue == 10.0


def test_delete_and_reupload_same_file():
    init_db()
    d = date(2026, 6, 11)
    serial = _excel_serial(d)
    csv = (
        "Ad Unit,Month,Date,Income Type,Ad Request,Matched Request,Impression,Click,"
        "Ad Request Ecpm,Ad Impression Ecpm,CTR,Coverage,Viewability,Net Revenue\n"
        f"web_unit_y,1,{serial},Open Auction,1,1,1,0,0,0,0,0,0,42\n"
    )
    with SessionLocal() as db:
        store.reset_all(db)
        store.import_upload_file(db, csv.encode("utf-8"), filename="dovizweb1.csv", commit=True)
        store.delete_source_file(db, "dovizweb1.csv")
        assert store.count_rows(db) == 0
        store.import_upload_file(db, csv.encode("utf-8"), filename="dovizweb1.csv", commit=True)
        row = db.query(AdReportRow).one()
        assert row.net_revenue == 42.0


def test_delete_file3_restores_file2_then_reupload_file3():
    """dovizandroid3 sil → android2 geri gelir; yeni android3 yüklenince duplike olmaz."""
    init_db()
    d = date(2026, 6, 10)
    serial = _excel_serial(d)
    csv = (
        "Ad Unit,Month,Date,Income Type,Ad Request,Matched Request,Impression,Click,"
        "Ad Request Ecpm,Ad Impression Ecpm,CTR,Coverage,Viewability,Net Revenue\n"
        f"app_unit,1,{serial},Open Auction,1,1,1,0,0,0,0,0,0,10\n"
    )
    csv2 = csv.replace(",10\n", ",50\n")
    csv3 = csv.replace(",10\n", ",80\n")
    with SessionLocal() as db:
        store.reset_all(db)
        store.import_upload_file(db, csv.encode("utf-8"), filename="dovizandroid1.csv", commit=True)
        store.import_upload_file(db, csv2.encode("utf-8"), filename="dovizandroid2.csv", commit=True)
        store.import_upload_file(db, csv3.encode("utf-8"), filename="dovizandroid3.csv", commit=True)
        row = db.query(AdReportRow).one()
        assert row.net_revenue == 80.0
        store.delete_source_file(db, "dovizandroid3.csv")
        row = db.query(AdReportRow).one()
        assert row.source_file == "dovizandroid2.csv"
        assert row.net_revenue == 50.0
        store.import_upload_file(db, csv3.replace(",80\n", ",95\n").encode("utf-8"), filename="dovizandroid3.csv", commit=True)
        row = db.query(AdReportRow).one()
        assert row.source_file == "dovizandroid3.csv"
        assert row.net_revenue == 95.0
        assert store.count_rows(db) == 1


def test_bulk_import_empty_file_reports_zero_bytes_hint():
    out = store.import_upload_files_bulk([(b"", "dovizweb3.xlsx")])
    summary = out["summary"]
    assert summary["failed_count"] == 1
    assert summary["integrated_rows"] == 0
    reason = summary["failed"][0]["reason"]
    assert "0 bayt" in reason
    assert summary["failed"][0]["filename"] == "dovizweb3.xlsx"


def test_suggested_detail_favorites_stream_keys_and_top_n():
    init_db()
    with SessionLocal() as db:
        out = store.suggested_detail_favorites(db, period_days=30)
    assert out["period_days"] == 30
    assert len(out["streams"]) == len(store.AD_STREAMS)
    assert out["streams"]["doviz:desktop"]["top_n"] == 15
    assert out["streams"]["doviz:android"]["top_n"] == 3
    assert isinstance(out["streams"]["doviz:desktop"]["units"], list)


def test_n_strips_tl_suffix_from_sheet_cells():
    assert store._n("1,44TL") == 1.44
    assert abs(store._n("5.060,54TL") - 5060.54) < 0.01
    assert store._n("0,00TL") == 0.0
    assert store._n("%12,3") == 12.3


def test_sheet_csv_parses_tr_revenue_and_stream_lock():
    from backend.services.ad_sheets_config import (
        AD_SHEET_SOURCES,
        is_sheet_catalog_filename,
        sheet_catalog_filename,
    )

    assert len(AD_SHEET_SOURCES) == 6
    assert sheet_catalog_filename("doviz:android") == "doviz_android_google_sheet.csv"
    assert is_sheet_catalog_filename("doviz_android_google_sheet.csv")
    assert not is_sheet_catalog_filename("dovizandroid1.xlsx")

    csv = (
        "Ad Unit,Month,Date,Income Type,Ad Request,Matched Request,Impression,Click,"
        "Ad Request Ecpm,Ad Impression Ecpm,CTR,Coverage,Viewability,Net Revenue\n"
        "doviz_android_x,Ocak 2024,01.01.2024,Open Auction,10,5,1,0,0,0,0,0,0,\"1,44TL\"\n"
    )
    stream = store._STREAM_BY_KEY["doviz:android"]
    rows = list(
        store.parse_csv_text(
            csv,
            filename="doviz_android_google_sheet.csv",
            stream=stream,
        )
    )
    assert len(rows) == 1
    assert rows[0]["net_revenue"] == 1.44
    assert rows[0]["project"] == "doviz"
    assert rows[0]["branch"] == "android"


def test_dedupe_batch_keeps_last_fingerprint():
    a = {"fingerprint": "fp1", "net_revenue": 1.0}
    b = {"fingerprint": "fp1", "net_revenue": 9.0}
    c = {"fingerprint": "fp2", "net_revenue": 2.0}
    out = store._dedupe_batch_by_fingerprint([a, b, c])
    by_fp = {r["fingerprint"]: r["net_revenue"] for r in out}
    assert by_fp == {"fp1": 9.0, "fp2": 2.0}


def test_sheet_incremental_keeps_rows_from_last_date_inclusive():
    """Son veri 2026-08-05 ise bir sonraki sync ≥ 05.08 satırlarını alır."""
    from backend.services.ad_analytics_store import iter_csv_text

    csv = (
        "Ad Unit,Month,Date,Income Type,Ad Request,Matched Request,Impression,Click,"
        "Ad Request Ecpm,Ad Impression Ecpm,CTR,Coverage,Viewability,Net Revenue\n"
        "u,Ağustos 2026,04.08.2026,Open Auction,1,1,1,0,0,0,0,0,0,10\n"
        "u,Ağustos 2026,05.08.2026,Open Auction,1,1,1,0,0,0,0,0,0,20\n"
        "u,Ağustos 2026,06.08.2026,Open Auction,1,1,1,0,0,0,0,0,0,30\n"
    )
    stream = store._STREAM_BY_KEY["doviz:android"]
    resume = date(2026, 8, 5)
    rows = list(
        iter_csv_text(
            csv,
            filename="doviz_android_google_sheet.csv",
            stream=stream,
            min_date=resume,
        )
    )
    dates = sorted({r["report_date"] for r in rows})
    assert dates == [date(2026, 8, 5), date(2026, 8, 6)]
    assert sum(r["net_revenue"] for r in rows) == 50.0


def test_web_sheet_excludes_sibling_mweb_business_keys():
    """Web sheet Mweb satırlarını da içeriyorsa desktop import’ta düşülür."""
    from backend.services.ad_analytics_store import (
        business_keys_from_csv_text,
        iter_csv_text,
        row_business_key,
    )
    from backend.services.ad_sheets_config import AD_SHEET_SOURCES

    desktop = next(s for s in AD_SHEET_SOURCES if s.stream_key == "sinemalar:desktop")
    assert desktop.exclude_sibling_stream_key == "sinemalar:mweb"
    doviz_desk = next(s for s in AD_SHEET_SOURCES if s.stream_key == "doviz:desktop")
    assert doviz_desk.exclude_sibling_stream_key == "doviz:mweb"

    header = (
        "Ad Unit,Month,Date,Income Type,Ad Request,Matched Request,Impression,Click,"
        "Ad Request Ecpm,Ad Impression Ecpm,CTR,Coverage,Viewability,Net Revenue\n"
    )
    mweb_csv = header + (
        "m_sinemalar_x,Haziran 2026,10.06.2026,Open Auction,1,1,1,0,0,0,0,0,0,100\n"
        "amp_shared,Haziran 2026,10.06.2026,Open Auction,1,1,1,0,0,0,0,0,0,50\n"
    )
    web_csv = header + (
        "m_sinemalar_x,Haziran 2026,10.06.2026,Open Auction,1,1,1,0,0,0,0,0,0,100\n"
        "amp_shared,Haziran 2026,10.06.2026,Open Auction,1,1,1,0,0,0,0,0,0,50\n"
        "web_sinemalar_y,Haziran 2026,10.06.2026,Open Auction,1,1,1,0,0,0,0,0,0,20\n"
    )
    exclude = business_keys_from_csv_text(mweb_csv)
    assert len(exclude) == 2
    stream = store._STREAM_BY_KEY["sinemalar:desktop"]
    kept = list(
        iter_csv_text(
            web_csv,
            filename="sinemalar_desktop_google_sheet.csv",
            stream=stream,
            exclude_keys=exclude,
        )
    )
    assert len(kept) == 1
    assert kept[0]["ad_unit"] == "web_sinemalar_y"
    assert kept[0]["net_revenue"] == 20.0
    assert row_business_key(kept[0]) not in exclude


def test_area_label_prefers_branch_over_surface():
    """Desktop dalındaki m_* surface’li satır WEB alanına yazılır."""
    init_db()
    header = (
        "Ad Unit,Month,Date,Income Type,Ad Request,Matched Request,Impression,Click,"
        "Ad Request Ecpm,Ad Impression Ecpm,CTR,Coverage,Viewability,Net Revenue\n"
    )
    csv = header + "m_sinemalar_x,Haziran 2026,10.06.2026,Open Auction,1,1,1,0,0,0,0,0,0,100\n"
    with SessionLocal() as db:
        store.reset_all(db)
        store.import_upload_file(
            db,
            csv.encode("utf-8"),
            filename="sinemalar_desktop_google_sheet.csv",
            commit=True,
            stream_key="sinemalar:desktop",
        )
        # Import surface=mweb (m_ prefix) ama branch=desktop
        row = db.query(AdReportRow).one()
        assert row.branch == "desktop"
        assert row.surface == "mweb"
        summary = store.query_summary(
            db, start="2026-06-01", end="2026-06-30", project="sinemalar"
        )
        by_area = {r["area"]: r["net_revenue"] for r in summary.get("by_area") or []}
        assert by_area.get("web", 0) == 100.0
        assert by_area.get("mweb", 0) == 0.0
        store.reset_all(db)

def test_iter_csv_scan_progress_every_n_rows():
    from backend.services.ad_analytics_store import iter_csv_text
    from backend.services.ad_sheets_sync import _fmt_row_count

    header = (
        "Ad Unit,Month,Date,Income Type,Ad Request,Matched Request,Impression,Click,"
        "Ad Request Ecpm,Ad Impression Ecpm,CTR,Coverage,Viewability,Net Revenue\n"
    )
    body = "".join(
        f"u,A,01.01.2024,Open Auction,1,1,1,0,0,0,0,0,0,1\n" for _ in range(12)
    )
    scans: list[tuple[int, int]] = []
    list(
        iter_csv_text(
            header + body,
            filename="doviz_android_google_sheet.csv",
            stream=store._STREAM_BY_KEY["doviz:android"],
            on_scan=lambda d, t: scans.append((d, t)),
            scan_every=5,
        )
    )
    assert scans[0] == (0, 12)
    assert (5, 12) in scans
    assert (10, 12) in scans
    assert scans[-1] == (12, 12)
    assert _fmt_row_count(111234) == "111k"
    assert _fmt_row_count(352897) == "352k"


def test_request_cancel_sync_job_flags_running_job():
    from backend.services import ad_sheets_sync as sheets_sync

    sheets_sync._set_job(
        running=False,
        cancel_requested=False,
        cancelled=False,
        phase="idle",
    )
    denied = sheets_sync.request_cancel_sync_job()
    assert denied["accepted"] is False

    sheets_sync._set_job(running=True, cancel_requested=False, phase="import", detail="…")
    accepted = sheets_sync.request_cancel_sync_job()
    assert accepted["accepted"] is True
    assert sheets_sync.is_sync_cancel_requested() is True
    assert accepted["job"]["phase"] == "cancelling"

    sheets_sync._set_job(
        running=False,
        cancel_requested=False,
        cancelled=False,
        phase="idle",
    )


def test_atomic_sync_cancel_rolls_back_stream_changes(monkeypatch):
    """İptalde bu oturumdaki yazımlar geri alınır; önceki satırlar kalır."""
    from backend.services import ad_sheets_sync as sheets_sync

    init_db()
    header = (
        "Ad Unit,Month,Date,Income Type,Ad Request,Matched Request,Impression,Click,"
        "Ad Request Ecpm,Ad Impression Ecpm,CTR,Coverage,Viewability,Net Revenue\n"
    )
    prior = header + "keep_unit,Haziran 2026,01.06.2026,Open Auction,1,1,1,0,0,0,0,0,0,42\n"
    incoming = header + (
        "keep_unit,Haziran 2026,01.06.2026,Open Auction,1,1,1,0,0,0,0,0,0,999\n"
        "new_unit,Haziran 2026,02.06.2026,Open Auction,1,1,1,0,0,0,0,0,0,99\n"
    )

    with SessionLocal() as db:
        store.reset_all(db)
        store.import_upload_file(
            db,
            prior.encode("utf-8"),
            filename="sinemalar_mweb_google_sheet.csv",
            commit=True,
            stream_key="sinemalar:mweb",
        )
        assert db.query(AdReportRow).count() == 1

        monkeypatch.setattr(
            sheets_sync,
            "fetch_public_sheet_csv",
            lambda url, timeout=300: incoming,
        )

        cancel_after_clear = {"v": False}

        def on_progress(payload):
            if payload.get("phase") == "clear":
                cancel_after_clear["v"] = True

        raised = False
        try:
            sheets_sync.sync_from_google_sheets(
                db,
                force=True,
                stream_key="sinemalar:mweb",
                full=True,
                atomic=True,
                on_progress=on_progress,
                cancel_check=lambda: cancel_after_clear["v"],
            )
        except sheets_sync.SyncCancelled:
            raised = True
            db.rollback()

        assert raised
        rows = db.query(AdReportRow).all()
        assert len(rows) == 1
        assert rows[0].ad_unit == "keep_unit"
        assert float(rows[0].net_revenue) == 42.0
        store.reset_all(db)
