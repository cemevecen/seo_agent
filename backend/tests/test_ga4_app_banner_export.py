from backend.services.ga4_app_banner_export import (
    build_app_banner_xlsx,
    build_dump_columns,
    collect_dump_dates,
)
from openpyxl import load_workbook
import io


def _sample_payload() -> dict:
    return {
        "project": "doviz",
        "profile": "android",
        "start": "2025-11-01",
        "end": "2025-11-03",
        "chart_start": "2025-11-01",
        "chart_end": "2025-11-03",
        "total_daily": {
            "dates": ["2025-11-01", "2025-11-02", "2025-11-03"],
            "values": [10, 0, 5],
        },
        "campaigns": [
            {
                "campaign": "test_camp",
                "daily": {
                    "dates": ["2025-11-01", "2025-11-02", "2025-11-03"],
                    "values": [3, 0, 1],
                },
            }
        ],
        "mweb_banner": {
            "ok": True,
            "events": [
                {
                    "event_name": "app_download_banner_click",
                    "daily": {
                        "dates": ["2025-11-01", "2025-11-02", "2025-11-03"],
                        "values": [0, 2, 0],
                    },
                }
            ],
        },
    }


def test_build_dump_columns_matches_ui_order():
    cols = build_dump_columns(_sample_payload())
    keys = [c.key for c in cols]
    assert keys[0] == "download"
    assert keys[1] == "camp:test_camp"
    assert keys[2].startswith("mweb:")


def test_collect_dump_dates_from_total_daily():
    assert collect_dump_dates(_sample_payload()) == ["2025-11-01", "2025-11-02", "2025-11-03"]


def test_build_xlsx_active_only_filters_rows():
    blob = build_app_banner_xlsx(_sample_payload(), active_only=True)
    wb = load_workbook(io.BytesIO(blob))
    ws = wb.active
    # başlık + header + 2 aktivite günü (01 ve 03; 02 sadece mweb)
    data_rows = [r for r in ws.iter_rows(min_row=3, values_only=True) if r and r[0]]
    dates = [str(r[0])[:10] for r in data_rows]
    assert "2025-11-01" in dates
    assert "2025-11-02" in dates
    assert "2025-11-03" in dates


def test_build_xlsx_has_headers():
    blob = build_app_banner_xlsx(_sample_payload(), active_only=False)
    wb = load_workbook(io.BytesIO(blob))
    ws = wb.active
    headers = [c.value for c in ws[2]]
    assert headers[0] == "Tarih"
    assert "download" in headers
    assert any("kampanya" in str(h) for h in headers)
