"""Notification Analytics — paylaşımlı workspace (tüm admin oturumları aynı veriyi görür)."""

from __future__ import annotations

import csv
import io
import json
import logging
import re
import time
import warnings
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from backend.models import NotificationAnalyticsWorkspace
from backend.services.backlink_csv import fetch_public_sheet_csv

LOGGER = logging.getLogger(__name__)
WORKSPACE_ID = 1

# Kaynak Google Sheet (herkese açık görüntüleyici) — dosya yükleme yerine tek kaynak.
NOTIFICATION_ANALYTICS_SHEET_URL = (
    "https://docs.google.com/spreadsheets/d/1NnizUEsaKpaabB0sCvksiHgx7Kfgt_upj1L2IXgUHeM/"
    "edit?gid=0#gid=0"
)
_SHEET_SYNC_TTL_SEC = 300.0
_last_sheet_sync_mono: float = 0.0


def _iso_utc_z(dt: datetime | None) -> str | None:
    """Naive UTC datetime → ISO string with Z (tarayıcıda Europe/Istanbul dönüşümü için)."""
    if dt is None:
        return None
    if dt.tzinfo is not None:
        from datetime import timezone

        dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
    return dt.isoformat() + "Z"


def _n(value: Any) -> float:
    """Oran / CTR — ondalık korunur (3,877 → 3.877)."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value) if value == value else 0.0
    s = str(value).strip()
    if not s:
        return 0.0
    s = re.sub(r"[%\s]", "", s)
    has_dot = "." in s
    has_comma = "," in s
    if has_dot and has_comma:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif has_dot:
        parts = s.split(".")
        if len(parts) > 2 and all(re.fullmatch(r"\d{3}", p) for p in parts[1:]):
            s = "".join(parts)
        elif len(parts) == 2 and re.fullmatch(r"\d{3}", parts[1]):
            if len(parts[0]) > 3:
                s = "".join(parts)
            else:
                s = parts[0] + "." + parts[1]
    elif has_comma:
        parts = s.split(",")
        if len(parts) > 1 and all(re.fullmatch(r"\d{3}", p) for p in parts[1:]):
            s = "".join(parts)
        else:
            s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _n_count(value: Any) -> float:
    """Click / impression — tam sayı; 48.521 → 48521, 1.670 → 1670."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        f = float(value)
        return f if f == f else 0.0
    s = str(value).strip()
    if not s:
        return 0.0
    s = re.sub(r"[%\s]", "", s)
    has_dot = "." in s
    has_comma = "," in s
    if has_dot and has_comma:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    if has_dot:
        parts = s.split(".")
        if len(parts) >= 2 and all(re.fullmatch(r"\d+", p) for p in parts):
            if all(re.fullmatch(r"\d{3}", p) for p in parts[1:]):
                return float("".join(parts))
    if has_comma:
        parts = s.split(",")
        if len(parts) >= 2 and all(re.fullmatch(r"\d{3}", p) for p in parts[1:]):
            return float("".join(parts))
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _normalize_id(raw: Any) -> str:
    return re.sub(r"[\s\u00a0.,·']", "", str(raw or "").strip())


def _normalize_header(h: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (h or "").lower())


def _parse_date_smart(raw: str) -> str | None:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.isoformat()
    if isinstance(raw, date):
        return datetime(raw.year, raw.month, raw.day).isoformat()
    if not raw:
        return None
    s = str(raw).strip()
    try:
        direct = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return direct.isoformat()
    except ValueError:
        pass
    m = re.match(
        r"^(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2,4})(?:\s+(\d{1,2}):(\d{2}))?$",
        s,
    )
    if not m:
        return None
    year = int(m.group(3))
    if year < 100:
        year += 2000
    try:
        dt = datetime(year, int(m.group(2)), int(m.group(1)), int(m.group(4) or 0), int(m.group(5) or 0))
        return dt.isoformat()
    except ValueError:
        return None


def _detect_delimiter(header_line: str) -> str:
    best = ","
    best_count = -1
    for delim in (",", ";", "\t"):
        count = len(header_line.split(delim))
        if count > best_count:
            best = delim
            best_count = count
    return best


_HEADER_SCAN_MAX_ROWS = 25


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day).isoformat()
    if isinstance(value, float):
        if value != value:
            return ""
        if value == int(value) and abs(value) < 1e15:
            return str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _row_has_cells(row: tuple[Any, ...] | list[Any]) -> bool:
    return any(c is not None and str(c).strip() for c in row)


def _column_indices(headers: list[str]) -> dict[str, int] | None:
    def pick(names: list[str], *, occurrence: int = 0) -> int:
        found = 0
        for i, h in enumerate(headers):
            if h in names:
                if found == occurrence:
                    return i
                found += 1
        return -1

    idx = {
        "id": pick(["id", "bildirimid", "notificationid"]),
        "text": pick(
            [
                "text",
                "title",
                "headline",
                "baslik",
                "icerik",
                "metin",
                "bildirimmetni",
                "notificationtext",
                "message",
                "content",
            ]
        ),
        "date": pick(
            [
                "date",
                "datetime",
                "timestamp",
                "tarih",
                "gonderimtarihi",
                "sentat",
                "publishdate",
                "gun",
            ]
        ),
        "ai": pick(["androidappimpression"]),
        "ac": pick(["androidappclick"]),
        "atr": pick(["androidappctr"]),
        "ic": pick(["iosappclick"]),
        "itr": pick(["iosappctr"]),
        "di": pick(["desktopimpression"]),
        "dc": pick(["desktopclick"]),
        "dtr": pick(["desktopctr"]),
        "mi": pick(["mobilewebimpression"]),
        "mc": pick(["mobilewebclick"]),
        "mtr": pick(["mobilewebctr"]),
    }
    # Kaynak tabloda CTR sütunu bazen yanlışlıkla ikinci "android app impression" diye adlandırılıyor.
    if idx["atr"] < 0:
        idx["atr"] = pick(["androidappimpression"], occurrence=1)
    if idx["text"] < 0 or idx["date"] < 0:
        return None
    return idx


def _build_row_from_cells(cols: list[str], idx: dict[str, int]) -> dict | None:
    def col(i: int) -> str:
        return cols[i] if 0 <= i < len(cols) else ""

    iso = _parse_date_smart(col(idx["date"]))
    if not iso:
        return None
    item = {
        "id": _normalize_id(col(idx["id"])) if idx["id"] >= 0 else "",
        "text": col(idx["text"]).strip(),
        "date": iso,
        "platforms": {
            "android": {
                "impression": _n_count(col(idx["ai"])),
                "click": _n_count(col(idx["ac"])),
                "ctr": _n(col(idx["atr"])),
            },
            "ios": {
                "click": _n_count(col(idx["ic"])),
                "ctr": _n(col(idx["itr"])),
            },
            "desktop": {
                "impression": _n_count(col(idx["di"])),
                "click": _n_count(col(idx["dc"])),
                "ctr": _n(col(idx["dtr"])),
            },
            "mobileweb": {
                "impression": _n_count(col(idx["mi"])),
                "click": _n_count(col(idx["mc"])),
                "ctr": _n(col(idx["mtr"])),
            },
        },
    }
    if not item["text"]:
        return None
    return _sanitize_row(item)


def _parse_tabular_rows(header_cells: list[str], data_rows: list[list[str]]) -> list[dict]:
    headers = [_normalize_header(h) for h in header_cells]
    idx = _column_indices(headers)
    if idx is None:
        return []
    out: list[dict] = []
    for raw_row in data_rows:
        cols = list(raw_row)
        if idx["text"] >= len(cols) and idx["date"] >= len(cols) and not any(cols):
            continue
        item = _build_row_from_cells(cols, idx)
        if item:
            out.append(item)
    return out


def parse_csv_text(text: str) -> list[dict]:
    raw = (text or "").strip()
    if not raw:
        return []
    sample = raw[:4096]
    delim = _detect_delimiter(sample.splitlines()[0] if sample.splitlines() else ",")
    try:
        reader = csv.reader(io.StringIO(raw), delimiter=delim)
        matrix = [list(row) for row in reader if any(str(c or "").strip() for c in row)]
    except csv.Error:
        lines = [ln for ln in raw.splitlines() if ln.strip()]
        if len(lines) < 2:
            return []
        matrix = [line.split(delim) for line in lines]
    if len(matrix) < 2:
        return []
    header_cells: list[str] | None = None
    header_idx = -1
    for i, row in enumerate(matrix[:_HEADER_SCAN_MAX_ROWS]):
        headers_norm = [_normalize_header(c) for c in row]
        if _column_indices(headers_norm) is not None:
            header_cells = [str(c or "") for c in row]
            header_idx = i
            break
    if header_cells is None:
        return []
    data_rows = [[str(c or "") for c in row] for row in matrix[header_idx + 1 :]]
    return _parse_tabular_rows(header_cells, data_rows)


def _load_notification_workbook(data: bytes):
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Workbook contains no default style*",
            category=UserWarning,
        )
        return load_workbook(io.BytesIO(data), read_only=True, data_only=True)


def parse_xlsx_bytes(raw: bytes) -> list[dict]:
    if not raw:
        return []
    wb = _load_notification_workbook(raw)
    try:
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if len(all_rows) < 2:
        return []
    header_cells: list[str] | None = None
    header_idx = -1
    for i, row in enumerate(all_rows[:_HEADER_SCAN_MAX_ROWS]):
        if not _row_has_cells(row):
            continue
        cells = [_cell_to_str(c) for c in row]
        headers_norm = [_normalize_header(c) for c in cells]
        if _column_indices(headers_norm) is not None:
            header_cells = cells
            header_idx = i
            break
    if header_cells is None:
        return []
    data_rows: list[list[str]] = []
    for row in all_rows[header_idx + 1 :]:
        if not _row_has_cells(row):
            continue
        data_rows.append([_cell_to_str(c) for c in row])
    return _parse_tabular_rows(header_cells, data_rows)


def parse_upload_bytes(raw: bytes, filename: str = "") -> list[dict]:
    name = (filename or "").lower()
    is_xlsx = name.endswith((".xlsx", ".xlsm", ".xltx"))
    if not is_xlsx and not name.endswith((".csv", ".txt", ".tsv")) and raw[:4] == b"PK\x03\x04":
        is_xlsx = True
    if is_xlsx:
        return parse_xlsx_bytes(raw)
    return parse_csv_text(decode_csv_bytes(raw))


def _highest_id(rows: list[dict]) -> int:
    best = 0
    for row in rows:
        try:
            val = int(_n(row.get("id")))
        except (TypeError, ValueError):
            val = 0
        if val > best:
            best = val
    return best


def _row_key(row: dict) -> str:
    return f"{row.get('id') or ''}|{row.get('text')}|{row.get('date')}"


def _sanitize_row(row: dict) -> dict:
    """iOS yalnızca click tutulur; impression alanı kullanılmaz."""
    platforms = row.get("platforms")
    if not isinstance(platforms, dict):
        return row
    ios = platforms.get("ios")
    if not isinstance(ios, dict) or "impression" not in ios:
        return row
    clean_ios = {k: v for k, v in ios.items() if k != "impression"}
    return {**row, "platforms": {**platforms, "ios": clean_ios}}


def _merge_rows(existing: list[dict], incoming: list[dict]) -> list[dict]:
    merged: dict[str, dict] = {}
    for row in existing + incoming:
        merged[_row_key(row)] = _sanitize_row(row)
    return sorted(merged.values(), key=lambda r: r.get("date") or "")


def _get_workspace(db: Session) -> NotificationAnalyticsWorkspace:
    row = db.get(NotificationAnalyticsWorkspace, WORKSPACE_ID)
    if row is None:
        row = NotificationAnalyticsWorkspace(id=WORKSPACE_ID)
        db.add(row)
        db.flush()
    return row


def _load_rows(row: NotificationAnalyticsWorkspace) -> list[dict]:
    try:
        data = json.loads(row.rows_json or "[]")
        return [_sanitize_row(r) for r in data if isinstance(r, dict)]
    except json.JSONDecodeError:
        return []


def _row_day_key(iso: str | None) -> str:
    return str(iso or "")[:10]


def _rows_date_bounds(rows: list[dict]) -> tuple[str | None, str | None]:
    days = sorted({_row_day_key(r.get("date")) for r in rows} - {""})
    if not days:
        return None, None
    return days[0], days[-1]


def filter_rows_by_date(
    rows: list[dict],
    *,
    start: str | None = None,
    end: str | None = None,
) -> list[dict]:
    """Tarih aralığı (YYYY-MM-DD); boş = filtre yok."""
    s = (start or "").strip()[:10] or None
    e = (end or "").strip()[:10] or None
    if not s and not e:
        return rows
    out: list[dict] = []
    for r in rows:
        d = _row_day_key(r.get("date"))
        if not d:
            continue
        if s and d < s:
            continue
        if e and d > e:
            continue
        out.append(r)
    return out


def workspace_rows_chunk(
    db: Session,
    *,
    offset: int,
    limit: int,
    start: str | None = None,
    end: str | None = None,
) -> dict:
    row = _get_workspace(db)
    all_rows = _load_rows(row)
    rows = filter_rows_by_date(all_rows, start=start, end=end)
    total = len(rows)
    off = max(0, int(offset))
    end_idx = min(total, off + int(limit))
    chunk = rows[off:end_idx]
    return {
        "ok": True,
        "rows": chunk,
        "offset": off,
        "limit": int(limit),
        "total": total,
        "total_unfiltered": len(all_rows),
        "has_more": end_idx < total,
        "filter_start": (start or "")[:10],
        "filter_end": (end or "")[:10],
    }


def _is_admin_notification_source(source: str | None) -> bool:
    s = (source or "").lower()
    return ("admin" in s) or ("bridge" in s)


def workspace_state(db: Session, *, include_rows: bool = True) -> dict:
    from backend.config import settings
    from backend.services.doviz_notification_admin import (
        admin_credentials_configured,
        admin_http_proxy,
        stats_url,
    )

    row = _get_workspace(db)
    rows = _load_rows(row)
    _min_d, _max_d = _rows_date_bounds(rows)
    admin_ready = admin_credentials_configured()
    stored_source = str(getattr(row, "source", None) or "").strip()
    stored_url = str(getattr(row, "source_url", None) or "").strip()
    if _is_admin_notification_source(stored_source):
        display_source = stored_source
        display_url = stored_url or stats_url()
    elif stored_source:
        display_source = stored_source
        display_url = stored_url or NOTIFICATION_ANALYTICS_SHEET_URL
    else:
        display_source = "doviz_admin" if admin_ready else "google_sheet"
        display_url = stats_url() if admin_ready else NOTIFICATION_ANALYTICS_SHEET_URL
    out: dict[str, Any] = {
        "ok": True,
        "last_id": int(row.last_id or 0),
        "start": row.filter_start or "",
        "end": row.filter_end or "",
        "preset": row.preset or "1y",
        "row_count": len(rows),
        "data_min_date": _min_d or "",
        "data_max_date": _max_d or "",
        "updated_at": _iso_utc_z(row.updated_at),
        "last_file_upload_at": _iso_utc_z(row.last_file_upload_at),
        "last_sheet_sync_at": _iso_utc_z(row.last_file_upload_at),
        "source": display_source,
        "source_url": display_url,
        "admin_source_url": stats_url(),
        "sheet_backup_url": NOTIFICATION_ANALYTICS_SHEET_URL,
        "admin_credentials_configured": admin_ready,
        "admin_requires_vpn": True,
        "admin_proxy_configured": bool(admin_http_proxy()),
        "auto_sync_minutes": 15,
        "ingest_configured": bool((settings.notification_ingest_token or "").strip()),
    }
    if include_rows:
        out["rows"] = rows
    return out


def save_workspace(
    db: Session,
    *,
    rows: list[dict] | None = None,
    last_id: int | None = None,
    start: str | None = None,
    end: str | None = None,
    preset: str | None = None,
) -> dict:
    row = _get_workspace(db)
    if rows is not None:
        row.rows_json = json.dumps(rows, ensure_ascii=False)
        if last_id is None:
            last_id = _highest_id(rows)
    if last_id is not None:
        row.last_id = int(last_id)
    if start is not None:
        row.filter_start = str(start or "")[:10]
    if end is not None:
        row.filter_end = str(end or "")[:10]
    if preset is not None:
        row.preset = str(preset or "1y")[:10]
    row.updated_at = datetime.utcnow()
    db.commit()
    return workspace_state(db)


def append_rows(db: Session, incoming: list[dict]) -> dict:
    row = _get_workspace(db)
    existing = _load_rows(row)
    max_id_before = max(int(row.last_id or 0), _highest_id(existing))
    filtered: list[dict] = []
    for item in incoming:
        rid = _n(item.get("id"))
        if rid > max_id_before or not rid:
            filtered.append(item)
    if not filtered:
        return {
            **workspace_state(db),
            "added": 0,
            "message": f"Yeni satır yok (son ID: {max_id_before}).",
        }
    merged = _merge_rows(existing, filtered)
    row.rows_json = json.dumps(merged, ensure_ascii=False)
    row.last_id = max(max_id_before, _highest_id(merged))
    row.updated_at = datetime.utcnow()
    db.commit()
    return {
        **workspace_state(db),
        "added": len(filtered),
        "message": f"{len(filtered)} yeni satır eklendi.",
    }


def upload_csv_text(db: Session, csv_text: str) -> dict:
    """CSV satırlarını id|text|date anahtarıyla birleştirir (mevcut ID'ler de güncellenir)."""
    return upload_parsed_rows(db, parse_csv_text(csv_text))


def upload_file_bytes(db: Session, raw: bytes, filename: str = "") -> dict:
    """CSV veya Excel (.xlsx) — aynı sütun eşlemesi ve birleştirme mantığı."""
    return upload_parsed_rows(db, parse_upload_bytes(raw, filename))


def upload_parsed_rows(db: Session, parsed: list[dict]) -> dict:
    if not parsed:
        return {
            **workspace_state(db, include_rows=False),
            "added": 0,
            "updated": 0,
            "parsed": 0,
            "message": (
                "Dosya parse edilemedi. Başlık satırında metin (text/title/başlık) ve tarih (date/tarih) "
                "sütunları ve en az bir veri satırı gerekli (CSV veya .xlsx)."
            ),
        }
    row = _get_workspace(db)
    existing = _load_rows(row)
    existing_keys = {_row_key(r) for r in existing}
    added = 0
    updated = 0
    seen_incoming: set[str] = set()
    for item in parsed:
        key = _row_key(item)
        if key in seen_incoming:
            continue
        seen_incoming.add(key)
        if key in existing_keys:
            updated += 1
        else:
            added += 1
            existing_keys.add(key)
    merged = _merge_rows(existing, parsed)
    min_day, max_day = _rows_date_bounds(merged)
    fe = (row.filter_end or "").strip()[:10]
    if fe and max_day and max_day > fe:
        row.filter_end = max_day
    row.rows_json = json.dumps(merged, ensure_ascii=False)
    row.last_id = max(int(row.last_id or 0), _highest_id(merged))
    row.last_file_upload_at = datetime.utcnow()
    row.updated_at = datetime.utcnow()
    db.commit()
    return {
        **workspace_state(db, include_rows=False),
        "added": added,
        "updated": updated,
        "parsed": len(parsed),
        "data_min_date": min_day or "",
        "data_max_date": max_day or "",
        "message": f"{len(parsed)} satır işlendi: {added} yeni, {updated} güncellendi.",
    }


def replace_workspace_from_rows(
    db: Session,
    parsed: list[dict],
    *,
    source: str = "",
    source_url: str = "",
) -> dict:
    """Tam yenileme — DB içeriği gelen satırlarla değiştirilir."""
    if not parsed:
        return {
            **workspace_state(db, include_rows=False),
            "added": 0,
            "updated": 0,
            "parsed": 0,
            "replaced": True,
            "message": (
                "Sheet parse edilemedi. Başlık satırında metin (text) ve tarih (date) sütunları gerekli."
            ),
        }
    # Aynı anahtar birden fazla gelirse son satır kalsın.
    by_key: dict[str, dict] = {}
    for item in parsed:
        by_key[_row_key(item)] = item
    merged = list(by_key.values())
    min_day, max_day = _rows_date_bounds(merged)
    row = _get_workspace(db)
    fe = (row.filter_end or "").strip()[:10]
    if max_day and (not fe or max_day > fe):
        row.filter_end = max_day
    fs = (row.filter_start or "").strip()[:10]
    if min_day and (not fs or min_day < fs):
        # özel aralık kullanıcı seçimiyse start'ı zorla geri çekme — yalnızca boşsa doldur
        if not fs:
            row.filter_start = min_day
    row.rows_json = json.dumps(merged, ensure_ascii=False)
    row.last_id = _highest_id(merged)
    if source:
        try:
            row.source = (source or "")[:64]
            row.source_url = (source_url or "")[:512]
        except Exception:
            pass
    row.last_file_upload_at = datetime.utcnow()
    row.updated_at = datetime.utcnow()
    db.commit()
    return {
        **workspace_state(db, include_rows=False),
        "added": len(merged),
        "updated": 0,
        "parsed": len(parsed),
        "replaced": True,
        "data_min_date": min_day or "",
        "data_max_date": max_day or "",
        "message": f"Kaynak senkronize edildi · {len(merged)} kayıt.",
    }


def sync_from_doviz_admin(db: Session, *, force: bool = False) -> dict:
    """Doviz.com admin notifications/stats → workspace (TTL ile throttle)."""
    global _last_sheet_sync_mono
    from backend.config import is_railway_runtime, settings
    from backend.services.doviz_notification_admin import (
        admin_credentials_configured,
        admin_http_proxy,
        fetch_notification_rows_from_admin,
    )

    if not settings.doviz_admin_notification_sync_enabled:
        return {
            **workspace_state(db, include_rows=False),
            "synced": False,
            "skipped": True,
            "ok": False,
            "message": "DOVIZ_ADMIN_NOTIFICATION_SYNC_ENABLED=false",
        }
    if not admin_credentials_configured():
        return {
            **workspace_state(db, include_rows=False),
            "synced": False,
            "skipped": True,
            "ok": False,
            "message": "DOVIZ_ADMIN_EMAIL / DOVIZ_ADMIN_PASSWORD tanımlı değil.",
        }

    if not force:
        return {
            **workspace_state(db, include_rows=False),
            "synced": False,
            "skipped": True,
            "message": "Kayıtlı veri kullanılıyor; canlı tarama yalnızca «Sayfayı güncelle».",
            "source": "doviz_admin",
        }

    if is_railway_runtime() and not (admin_http_proxy() or "").strip():
        LOGGER.info("Notification admin live fetch skipped on Railway (no VPN proxy)")
        return {
            **workspace_state(db, include_rows=False),
            "synced": False,
            "skipped": True,
            "message": "Railway canlı admin taraması yok (VPN proxy yok); Mac köprüsü / «Sayfayı güncelle».",
            "source": "doviz_admin",
        }

    now = time.monotonic()
    if (
        _last_sheet_sync_mono > 0
        and (now - _last_sheet_sync_mono) < _SHEET_SYNC_TTL_SEC
    ):
        return {
            **workspace_state(db, include_rows=False),
            "synced": False,
            "skipped": True,
            "message": "Son senkronizasyon taze; admin yeniden çekilmedi.",
            "source": "doviz_admin",
        }

    try:
        fetched = fetch_notification_rows_from_admin()
    except Exception as exc:  # noqa: BLE001
        LOGGER.warning("Notification admin fetch failed: %s", exc)
        return {
            **workspace_state(db, include_rows=False),
            "synced": False,
            "skipped": False,
            "ok": False,
            "source": "doviz_admin",
            "message": str(exc) or "Doviz admin okunamadı.",
        }

    parsed = fetched.get("rows") or []
    result = replace_workspace_from_rows(
        db,
        parsed,
        source="doviz_admin",
        source_url=fetched.get("source_url") or "https://www.doviz.com/admin/notifications/stats",
    )
    if result.get("parsed"):
        _last_sheet_sync_mono = time.monotonic()
        result["synced"] = True
        result["skipped"] = False
        result["message"] = (
            f"Doviz admin senkronize edildi · {result.get('added') or result.get('parsed')} kayıt."
        )
    else:
        result["synced"] = False
        result["skipped"] = False
        result["message"] = result.get("message") or "Admin tablosundan satır çıkarılamadı."
    result["source"] = "doviz_admin"
    result["source_url"] = fetched.get("source_url") or "https://www.doviz.com/admin/notifications/stats"
    result["fetch_meta"] = {
        "elapsed_sec": fetched.get("elapsed_sec"),
        "html_chars": fetched.get("html_chars"),
        "csv_chars": fetched.get("csv_chars"),
    }
    return result


def sync_notification_analytics(db: Session, *, force: bool = False) -> dict:
    """Otomatik: Doviz admin/notifications/stats. Google Sheet yedek kapalı.

    UI’ya bilgi girilmez — Railway/env credentials veya Mac köprüsü ingest.
    """
    from backend.config import settings
    from backend.services.doviz_notification_admin import admin_credentials_configured

    try_admin = bool(
        settings.doviz_admin_notification_sync_enabled and admin_credentials_configured()
    )

    if try_admin:
        return sync_from_doviz_admin(db, force=force)

    return {
        "ok": False,
        "synced": False,
        "skipped": True,
        "message": "Admin tarama yok — Mac köprüsü ingest bekleniyor.",
        "source": "doviz_admin_bridge",
    }


def ingest_notification_rows(
    db: Session,
    rows: list[dict],
    *,
    source: str = "doviz_admin_bridge",
) -> dict:
    """VPN köprüsü / harici worker’dan gelen satırları yazar (manuel UI yok)."""
    global _last_sheet_sync_mono
    from backend.services.doviz_notification_admin import stats_url

    parsed = [r for r in (rows or []) if isinstance(r, dict)]
    result = replace_workspace_from_rows(
        db,
        parsed,
        source=(source or "doviz_admin_bridge").strip() or "doviz_admin_bridge",
        source_url=stats_url(),
    )
    if result.get("parsed"):
        _last_sheet_sync_mono = time.monotonic()
        result["synced"] = True
        result["skipped"] = False
        result["message"] = (
            f"Admin bridge ingest · {result.get('added') or result.get('parsed')} kayıt."
        )
    else:
        result["synced"] = False
        result["skipped"] = False
        result["ok"] = False
        result["message"] = result.get("message") or "Ingest: satır yok."
    result["source"] = source or "doviz_admin_bridge"
    result["source_url"] = stats_url()
    return result


def sync_from_google_sheet(
    db: Session,
    *,
    force: bool = False,
    prefer_sheet: bool = False,
) -> dict:
    """Google Sheet yedek / fallback yolu.

    prefer_sheet=False iken admin/bridge snapshot varsa sheet ile ezme
    (sheet bugünü gecikmeli getirir → 43 yerine 40 gibi sapma).
    """
    global _last_sheet_sync_mono
    row = _get_workspace(db)
    stored_source = str(getattr(row, "source", None) or "")
    if not prefer_sheet and _is_admin_notification_source(stored_source):
        existing = _load_rows(row)
        if existing:
            LOGGER.info(
                "Notification sheet atlandı — admin/bridge snapshot korunuyor (%s, %s kayıt)",
                stored_source,
                len(existing),
            )
            return {
                **workspace_state(db, include_rows=False),
                "synced": False,
                "skipped": True,
                "sheet_skipped": True,
                "source": stored_source,
                "message": (
                    "Admin/bridge verisi korunuyor; Google Sheet yedek olarak atlandı "
                    "(eksik günler ezilmesin)."
                ),
            }

    now = time.monotonic()
    if (
        not force
        and _last_sheet_sync_mono > 0
        and (now - _last_sheet_sync_mono) < _SHEET_SYNC_TTL_SEC
    ):
        return {
            **workspace_state(db, include_rows=False),
            "synced": False,
            "skipped": True,
            "source": "google_sheet",
            "message": "Son senkron taze; sheet yeniden çekilmedi.",
        }
    try:
        csv_text = fetch_public_sheet_csv(NOTIFICATION_ANALYTICS_SHEET_URL, timeout=60)
    except Exception as exc:  # noqa: BLE001
        LOGGER.warning("Notification sheet fetch failed: %s", exc)
        return {
            **workspace_state(db, include_rows=False),
            "synced": False,
            "skipped": False,
            "ok": False,
            "source": "google_sheet",
            "message": str(exc) or "Google Sheet okunamadı.",
        }
    parsed = parse_csv_text(csv_text)
    # Kaynak etiketi olmasa bile: sheet mevcut veriden daha eskiyse ezme
    if not prefer_sheet:
        existing = _load_rows(row)
        if existing:
            _emin, emax = _rows_date_bounds(existing)
            _smin, smax = _rows_date_bounds(parsed)
            if emax and smax and smax < emax:
                LOGGER.info(
                    "Notification sheet geride (sheet_max=%s db_max=%s) — overwrite yok",
                    smax,
                    emax,
                )
                return {
                    **workspace_state(db, include_rows=False),
                    "synced": False,
                    "skipped": True,
                    "sheet_skipped": True,
                    "message": (
                        f"Sheet yedek geride (son gün {smax} < {emax}); mevcut veri korundu."
                    ),
                }
    result = replace_workspace_from_rows(
        db,
        parsed,
        source="google_sheet",
        source_url=NOTIFICATION_ANALYTICS_SHEET_URL,
    )
    if result.get("parsed"):
        _last_sheet_sync_mono = time.monotonic()
        result["synced"] = True
        result["skipped"] = False
        result["message"] = (
            f"Google Sheet senkronize edildi · {result.get('added') or result.get('parsed')} kayıt."
        )
    else:
        result["synced"] = False
        result["skipped"] = False
    result["source"] = "google_sheet"
    result["source_url"] = NOTIFICATION_ANALYTICS_SHEET_URL
    return result


def decode_csv_bytes(raw: bytes) -> str:
    """UTF-8 / Windows Türkçe CSV kodlamalarını dene."""
    if not raw:
        return ""
    for enc in ("utf-8-sig", "utf-8", "cp1254", "iso-8859-9", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def reset_workspace(db: Session) -> dict:
    row = _get_workspace(db)
    row.rows_json = "[]"
    row.last_id = 0
    row.last_file_upload_at = None
    row.updated_at = datetime.utcnow()
    db.commit()
    return workspace_state(db)
