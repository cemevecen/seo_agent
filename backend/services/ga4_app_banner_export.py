"""GA4 app-banner paneli — günlük döküm tablosunu Excel olarak dışa aktar."""

from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import timedelta
from typing import Any

from backend.services.ga4_app_attribution import _parse_iso_date


def _has_series_signal(series: dict[str, Any] | None) -> bool:
    if not series or not isinstance(series, dict):
        return False
    for v in series.get("values") or []:
        if float(v or 0) > 0:
            return True
    return False


def _value_at(series: dict[str, Any] | None, date_iso: str) -> float | None:
    if not series or not isinstance(series, dict):
        return None
    dates = series.get("dates") or []
    values = series.get("values") or []
    target = str(date_iso)[:10]
    for i, d in enumerate(dates):
        if str(d)[:10] == target:
            if i >= len(values):
                return None
            try:
                return float(values[i])
            except (TypeError, ValueError):
                return None
    return None


def _short_banner_event_name(full: str) -> str:
    s = str(full or "").strip()
    prefix = "app_download_banner_"
    if s.startswith(prefix):
        return s[len(prefix) :]
    if s == "app_download_banner":
        return "banner"
    return s


def _filter_campaigns_with_signal(campaigns: list[dict[str, Any]] | None) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for c in campaigns or []:
        if isinstance(c, dict) and _has_series_signal(c.get("daily")):
            out.append(c)
    return out


def _filter_mweb_events_with_signal(events: list[dict[str, Any]] | None) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for ev in events or []:
        if isinstance(ev, dict) and _has_series_signal(ev.get("daily")):
            out.append(ev)
    return out


@dataclass(frozen=True)
class _DumpColumn:
    key: str
    label: str
    group: str
    series: dict[str, Any]


def build_dump_columns(payload: dict[str, Any]) -> list[_DumpColumn]:
    """UI `buildDailyDumpColumns` ile aynı sütun sırası."""
    cols: list[_DumpColumn] = []
    if _has_series_signal(payload.get("total_daily")):
        cols.append(
            _DumpColumn(
                key="download",
                label="download",
                group="app",
                series=payload["total_daily"],
            )
        )
    for c in _filter_campaigns_with_signal(payload.get("campaigns")):
        name = str(c.get("campaign") or "(not set)")
        cols.append(
            _DumpColumn(
                key=f"camp:{name}",
                label=name,
                group="campaign",
                series=c["daily"],
            )
        )
    mw = payload.get("mweb_banner")
    if isinstance(mw, dict):
        for ev in _filter_mweb_events_with_signal(mw.get("events")):
            label = _short_banner_event_name(str(ev.get("event_name") or ""))
            cols.append(
                _DumpColumn(
                    key=f"mweb:{ev.get('event_name') or label}",
                    label=label,
                    group="mweb",
                    series=ev["daily"],
                )
            )
    asc_c = payload.get("app_store_campaign_downloads")
    prof = str(payload.get("profile") or "").strip().lower()
    if prof == "ios" and isinstance(asc_c, dict) and asc_c.get("ok"):
        combined = asc_c.get("combined_daily")
        if isinstance(combined, dict) and _has_series_signal(combined):
            cols.append(
                _DumpColumn(
                    key="asc_total",
                    label="ASC banner (toplam)",
                    group="asc",
                    series=combined,
                )
            )
        for c in _filter_campaigns_with_signal(asc_c.get("campaigns")):
            camp = str(c.get("campaign") or "")
            cols.append(
                _DumpColumn(
                    key=f"asc:{camp}",
                    label=f"ASC {camp}",
                    group="asc",
                    series=c["daily"],
                )
            )
    return cols


def collect_dump_dates(payload: dict[str, Any]) -> list[str]:
    total = payload.get("total_daily")
    if isinstance(total, dict):
        dates = [str(d)[:10] for d in (total.get("dates") or []) if str(d)[:10]]
        if dates:
            return dates
    start_s = str(payload.get("chart_start") or payload.get("start") or "")[:10]
    end_s = str(payload.get("chart_end") or payload.get("end") or "")[:10]
    if not start_s or not end_s:
        return []
    start_d = _parse_iso_date(start_s)
    end_d = _parse_iso_date(end_s)
    if end_d < start_d:
        return []
    out: list[str] = []
    cur = start_d
    while cur <= end_d:
        out.append(cur.isoformat())
        cur += timedelta(days=1)
    return out


def _header_for_column(col: _DumpColumn) -> str:
    tag = col.group
    if tag == "app":
        return col.label
    if tag == "campaign":
        return f"kampanya · {col.label}"
    if tag == "mweb":
        return f"mweb · {col.label}"
    if tag == "asc":
        return col.label
    return col.label


def build_app_banner_xlsx(payload: dict[str, Any], *, active_only: bool = False) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    cols = build_dump_columns(payload)
    dates = collect_dump_dates(payload)

    wb = Workbook()
    ws = wb.active
    ws.title = "Günlük döküm"

    project = str(payload.get("project") or "doviz")
    profile = str(payload.get("profile") or "")
    start_s = str(payload.get("chart_start") or payload.get("start") or "")[:10]
    end_s = str(payload.get("chart_end") or payload.get("end") or "")[:10]
    ws.append([f"GA4 app banner — {project} · {profile} · {start_s} – {end_s}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(2, len(cols) + 1))
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=12)

    headers = ["Tarih"] + [_header_for_column(c) for c in cols]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_row = 2
    for col_i in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_i)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left" if col_i == 1 else "right", vertical="center")

    row_idx = header_row + 1
    for date_iso in dates:
        nums: list[float] = []
        for col in cols:
            v = _value_at(col.series, date_iso)
            nums.append(0.0 if v is None else float(v))
        if active_only and not any(n > 0 for n in nums):
            continue
        ws.append([date_iso] + nums)
        for col_i in range(2, len(headers) + 1):
            ws.cell(row=row_idx, column=col_i).alignment = Alignment(horizontal="right")
        row_idx += 1

    ws.column_dimensions["A"].width = 12
    for i in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = min(42, max(10, len(headers[i - 1]) + 2))

    ws.freeze_panes = "A3"
    if ws.max_row >= header_row:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{ws.max_row}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
