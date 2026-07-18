"""Empower uygulama export (xlsx) — yükleme, silme, grafik overlay."""

from __future__ import annotations

import hashlib
import io
import logging
import re
import unicodedata
from datetime import date, datetime, timezone
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session

from backend.models import AppEmpowerDailyRow, AppEmpowerImport
from backend.services.app_empower_config import APP_EMPOWER_SERIES, SERIES_BY_KEY

LOGGER = logging.getLogger(__name__)

_FILENAME_RE = re.compile(r"(?i)(android|ios)(?:empower|empove)(\d+)")
_MAX_BYTES = 25 * 1024 * 1024


def _compact_filename_stem(filename: str) -> str:
    """Boşluk, nokta, tire → tek parça (doviz Android Empower 1 → dovizandroidempower1)."""
    base = (filename or "").strip().split("/")[-1]
    stem = base.rsplit(".", 1)[0] if "." in base else base
    compact = re.sub(r"[^a-z0-9]", "", stem.lower())
    # Sık yazım: dovizanroidempower → dovizandroidempower
    compact = re.sub(r"anroidempow", "androidempow", compact)
    return compact


def _norm_header(cell: str) -> str:
    s = unicodedata.normalize("NFKD", (cell or "").strip().lower())
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s


def parse_filename_meta(filename: str) -> tuple[str, int]:
    """dovizandroidempower1.xlsx → ('android', 1)."""
    compact = _compact_filename_stem(filename)
    m = _FILENAME_RE.search(compact)
    if not m:
        raise ValueError(
            "Dosya adı androidempower1 / iosempower2 biçiminde olmalı "
            "(ör. dovizandroidempower1.xlsx; boşluk/nokta kabul edilir)"
        )
    platform = m.group(1).lower()
    doc_index = int(m.group(2))
    if doc_index < 1:
        raise ValueError("Dosya numarası 1 veya daha büyük olmalı")
    return platform, doc_index


def is_empower_filename(filename: str) -> bool:
    try:
        parse_filename_meta(filename)
        return True
    except ValueError:
        return False


def partition_mz_upload_files(
    files: list[tuple[bytes, str]],
) -> tuple[list[tuple[bytes, str]], list[tuple[bytes, str]]]:
    """Reklam MX yüklemesinden Empower xlsx dosyalarını ayırır.

    Returns:
        (ad_files, empower_files) — sıra önemli; çağıranlar bunu ters unpack etmemeli.
    """
    empower: list[tuple[bytes, str]] = []
    ad: list[tuple[bytes, str]] = []
    for raw, name in files:
        if is_empower_filename(name):
            empower.append((raw, name))
        else:
            ad.append((raw, name))
    return ad, empower


def _parse_float(raw: Any) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip()
    if not s or s in ("-", "—"):
        return None
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _parse_date_cell(raw: Any) -> date | None:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, (int, float)):
        n = float(raw)
        if 1 < n < 600000:
            try:
                from openpyxl.utils.datetime import from_excel

                dt = from_excel(n)
                if isinstance(dt, datetime):
                    return dt.date()
                if isinstance(dt, date):
                    return dt
            except Exception:  # noqa: BLE001
                pass
    s = str(raw).strip()
    if not s:
        return None
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return date.fromisoformat(s[:10])
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _header_map(headers: list[str]) -> dict[str, int]:
    norm = [_norm_header(h) for h in headers]
    idx: dict[str, int] = {}

    def pick(*candidates: str) -> int | None:
        for c in candidates:
            c = _norm_header(c)
            if c in norm:
                return norm.index(c)
        return None

    date_i = pick("date", "tarih", "gun", "gün", "day")
    if date_i is None:
        raise ValueError("Date sütunu bulunamadı")
    idx["date"] = date_i
    for key, cands in {
        "sessions": ("sessions",),
        "dau_7d": ("dau (7 days)", "dau 7 days", "dau"),
        "crash_affected_users": ("crash affected users", "crash affected user"),
        "avg_session_duration": ("average session duration", "avg session duration"),
        "engagement_rate": ("engagement rate",),
        "arpdau_usd": ("arpdau ($)", "arpdau", "arpdau usd"),
        "app_version": ("app version",),
    }.items():
        col = pick(*cands)
        if col is not None:
            idx[key] = col
    return idx


def parse_empower_xlsx(raw: bytes) -> list[dict[str, Any]]:
    if not raw:
        raise ValueError("Boş dosya")
    if len(raw) > _MAX_BYTES:
        raise ValueError("Dosya çok büyük (max 25 MB)")
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            raise ValueError("Başlık satırı yok")
        headers = [str(c) if c is not None else "" for c in header_row]
        colmap = _header_map(headers)
        out: list[dict[str, Any]] = []
        for row in rows_iter:
            if not row:
                continue
            d_i = colmap["date"]
            d_raw = row[d_i] if d_i < len(row) else None
            rd = _parse_date_cell(d_raw)
            if not rd:
                continue

            def cell(key: str) -> Any:
                i = colmap.get(key)
                if i is None or i >= len(row):
                    return None
                return row[i]

            out.append(
                {
                    "report_date": rd,
                    "sessions": _parse_float(cell("sessions")) or 0.0,
                    "dau_7d": _parse_float(cell("dau_7d")) or 0.0,
                    "crash_affected_users": _parse_float(cell("crash_affected_users")) or 0.0,
                    "avg_session_duration": _parse_float(cell("avg_session_duration")) or 0.0,
                    "engagement_rate": _parse_float(cell("engagement_rate")) or 0.0,
                    "arpdau_usd": _parse_float(cell("arpdau_usd")) or 0.0,
                    "app_version": str(cell("app_version") or "").strip()[:64],
                }
            )
        if not out:
            raise ValueError("Veri satırı okunamadı")
        return out
    finally:
        wb.close()


def list_imports(db: Session) -> list[dict[str, Any]]:
    rows = (
        db.execute(select(AppEmpowerImport).order_by(AppEmpowerImport.platform, AppEmpowerImport.doc_index))
        .scalars()
        .all()
    )
    return [
        {
            "source_file": r.source_file,
            "platform": r.platform,
            "doc_index": r.doc_index,
            "row_count": r.row_count,
            "has_raw": bool(r.raw_bytes),
            "uploaded_at": r.uploaded_at.isoformat() if r.uploaded_at else None,
        }
        for r in rows
    ]


def import_file(db: Session, raw: bytes, filename: str) -> dict[str, Any]:
    platform, doc_index = parse_filename_meta(filename)
    parsed = parse_empower_xlsx(raw)
    name = (filename or "empower.xlsx").strip().split("/")[-1][:255]
    sha = hashlib.sha256(raw).hexdigest()
    now = datetime.now(timezone.utc).replace(tzinfo=None)

    imp = db.execute(select(AppEmpowerImport).where(AppEmpowerImport.source_file == name)).scalar_one_or_none()
    if imp:
        imp.platform = platform
        imp.doc_index = doc_index
        imp.row_count = len(parsed)
        imp.file_sha256 = sha
        imp.raw_bytes = raw
        imp.uploaded_at = now
    else:
        imp = AppEmpowerImport(
            source_file=name,
            platform=platform,
            doc_index=doc_index,
            row_count=len(parsed),
            file_sha256=sha,
            raw_bytes=raw,
            uploaded_at=now,
        )
        db.add(imp)

    db.execute(
        delete(AppEmpowerDailyRow).where(
            AppEmpowerDailyRow.platform == platform,
            AppEmpowerDailyRow.doc_index == doc_index,
        )
    )
    for p in parsed:
        db.add(
            AppEmpowerDailyRow(
                platform=platform,
                doc_index=doc_index,
                source_file=name,
                report_date=p["report_date"],
                sessions=p["sessions"],
                dau_7d=p["dau_7d"],
                crash_affected_users=p["crash_affected_users"],
                avg_session_duration=p["avg_session_duration"],
                engagement_rate=p["engagement_rate"],
                arpdau_usd=p["arpdau_usd"],
                app_version=p["app_version"],
                uploaded_at=now,
            )
        )
    db.commit()
    dates = [p["report_date"] for p in parsed]
    try:
        from backend.services.ad_analytics_store import invalidate_facets_cache

        invalidate_facets_cache()
    except Exception:  # noqa: BLE001
        pass
    return {
        "ok": True,
        "source_file": name,
        "platform": platform,
        "doc_index": doc_index,
        "rows": len(parsed),
        "min_date": min(dates).isoformat(),
        "max_date": max(dates).isoformat(),
    }


def import_files_bulk(db: Session, files: list[tuple[bytes, str]]) -> dict[str, Any]:
    results: list[dict[str, Any]] = []
    for raw, name in files:
        try:
            results.append(import_file(db, raw, name))
        except Exception as exc:  # noqa: BLE001
            db.rollback()
            results.append({"ok": False, "source_file": name, "error": str(exc)})
    ok = sum(1 for r in results if r.get("ok"))
    return {"ok_count": ok, "total": len(files), "results": results}


def refresh_import(db: Session, source_file: str) -> dict[str, Any]:
    name = source_file.strip().split("/")[-1]
    imp = db.execute(select(AppEmpowerImport).where(AppEmpowerImport.source_file == name)).scalar_one_or_none()
    if not imp or not imp.raw_bytes:
        raise ValueError("Kayıt bulunamadı veya yeniden okuma için ham dosya yok (dosyayı tekrar yükleyin)")
    return import_file(db, bytes(imp.raw_bytes), name)


def delete_source_file(db: Session, source_file: str) -> dict[str, Any]:
    name = source_file.strip().split("/")[-1]
    imp = db.execute(select(AppEmpowerImport).where(AppEmpowerImport.source_file == name)).scalar_one_or_none()
    if not imp:
        raise ValueError("Dosya bulunamadı")
    platform, doc_index = imp.platform, imp.doc_index
    db.execute(
        delete(AppEmpowerDailyRow).where(
            AppEmpowerDailyRow.platform == platform,
            AppEmpowerDailyRow.doc_index == doc_index,
        )
    )
    db.delete(imp)
    db.commit()
    return {"ok": True, "deleted": name, "platform": platform, "doc_index": doc_index}


def delete_source_files_bulk(db: Session, source_files: list[str]) -> dict[str, Any]:
    deleted: list[str] = []
    errors: list[dict[str, str]] = []
    for sf in source_files:
        try:
            delete_source_file(db, sf)
            deleted.append(sf.strip().split("/")[-1])
        except Exception as exc:  # noqa: BLE001
            errors.append({"source_file": sf, "error": str(exc)})
    return {"ok": not errors, "deleted": deleted, "errors": errors}


def _merged_rows_for_platform(
    db: Session,
    platform: str,
    *,
    start: date | None,
    end: date | None,
) -> list[AppEmpowerDailyRow]:
    q = db.query(AppEmpowerDailyRow).filter(AppEmpowerDailyRow.platform == platform)
    if start:
        q = q.filter(AppEmpowerDailyRow.report_date >= start)
    if end:
        q = q.filter(AppEmpowerDailyRow.report_date <= end)
    rows = q.order_by(AppEmpowerDailyRow.report_date, AppEmpowerDailyRow.doc_index.desc()).all()
    by_date: dict[date, AppEmpowerDailyRow] = {}
    for r in rows:
        if r.report_date not in by_date:
            by_date[r.report_date] = r
    return [by_date[d] for d in sorted(by_date.keys())]


def query_overlay(
    db: Session,
    *,
    platform: str,
    start: str | None = None,
    end: str | None = None,
    series_keys: list[str] | None = None,
) -> dict[str, Any]:
    plat = (platform or "").strip().lower()
    if plat not in ("android", "ios"):
        raise ValueError("platform android veya ios olmalı")
    keys = series_keys or [s.key for s in APP_EMPOWER_SERIES]
    keys = [k for k in keys if k in SERIES_BY_KEY]
    start_d = date.fromisoformat(start[:10]) if start else None
    end_d = date.fromisoformat(end[:10]) if end else None
    merged = _merged_rows_for_platform(db, plat, start=start_d, end=end_d)
    latest = db.query(func.max(AppEmpowerImport.uploaded_at)).scalar()
    series_out: dict[str, Any] = {}
    for k in keys:
        spec = SERIES_BY_KEY[k]
        pts = []
        for r in merged:
            val = getattr(r, k, None)
            pts.append({"date": r.report_date.isoformat(), "value": float(val or 0)})
        series_out[k] = {"key": k, "label": spec.label, "unit": spec.unit, "by_date": pts}
    return {
        "platform": plat,
        "synced_at": latest.isoformat() if latest else None,
        "range": {"start": start, "end": end},
        "series": series_out,
    }
