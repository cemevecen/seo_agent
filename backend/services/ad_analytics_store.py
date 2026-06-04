"""Reklam raporu (Excel/CSV) — DB saklama, filtreleme ve özet agregasyon."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
import re
import warnings
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from collections.abc import Callable
from typing import Any, Iterable, Iterator

ProgressCallback = Callable[[dict[str, Any]], None]

from openpyxl import load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.orm import Session

from backend.database import SessionLocal, engine
from backend.models import AdReportCatalog, AdReportRow

_IS_PG = "postgresql" in str(engine.url)

LOGGER = logging.getLogger(__name__)

_IMPORT_BATCH_SIZE = 400


def _load_workbook_bytes(data: bytes):
    """Ad platformu xlsx dosyalarında openpyxl 'no default style' uyarısı zararsızdır."""
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Workbook contains no default style*",
            category=UserWarning,
        )
        return load_workbook(io.BytesIO(data), read_only=True, data_only=True)


HEADERS_MAP = {
    "adunit": "ad_unit",
    "adunitname": "ad_unit",
    "reklambirimi": "ad_unit",
    "reklambirim": "ad_unit",
    "month": "month",
    "ay": "month",
    "date": "report_date",
    "tarih": "report_date",
    "incometype": "income_type",
    "gelirtipi": "income_type",
    "gelirturu": "income_type",
    "gelirtur": "income_type",
    "adrequest": "ad_request",
    "matchedrequest": "matched_request",
    "impression": "impression",
    "click": "click",
    "adrequestecpm": "ad_request_ecpm",
    "adimpressionecpm": "ad_impression_ecpm",
    "adimpressionecpmtl": "ad_impression_ecpm",
    "ctr": "ctr",
    "coverage": "coverage",
    "viewability": "viewability",
    "netrevenue": "net_revenue",
    "netrevenuetl": "net_revenue",
    "empowerpageview": "empower_pageview",
    "empoweruniquevisitor": "empower_unique_visitor",
    "pageviewecpm": "pageview_ecpm",
    "uniquevisitorecpm": "unique_visitor_ecpm",
    "abovethefoldratio": "above_the_fold_ratio",
}

# Bilinen alan dışı sütunlar extra_metrics JSON'da saklanır (ham başlık anahtarı).
_CORE_FIELDS = frozenset(
    {
        "ad_unit",
        "month",
        "report_date",
        "income_type",
        "ad_request",
        "matched_request",
        "impression",
        "click",
        "ad_request_ecpm",
        "ad_impression_ecpm",
        "ctr",
        "coverage",
        "viewability",
        "net_revenue",
        "empower_pageview",
        "empower_unique_visitor",
        "pageview_ecpm",
        "unique_visitor_ecpm",
        "above_the_fold_ratio",
    }
)

_HEADER_SCAN_MAX_ROWS = 30

_HEADER_SUBSTRING_RULES: list[tuple[str, str]] = [
    ("netrevenue", "net_revenue"),
    ("toplamgelir", "net_revenue"),
    ("totalrevenue", "net_revenue"),
    ("netgelir", "net_revenue"),
    ("adunit", "ad_unit"),
    ("reklambirim", "ad_unit"),
    ("incometype", "income_type"),
    ("gelirtip", "income_type"),
    ("gelirtur", "income_type"),
    ("adrequest", "ad_request"),
    ("matchedrequest", "matched_request"),
    ("impression", "impression"),
    ("click", "click"),
    ("adrequestecpm", "ad_request_ecpm"),
    ("adimpressionecpm", "ad_impression_ecpm"),
    ("pageviewecpm", "pageview_ecpm"),
    ("uniquevisitorecpm", "unique_visitor_ecpm"),
    ("empowerpageview", "empower_pageview"),
    ("empoweruniquevisitor", "empower_unique_visitor"),
    ("abovethefold", "above_the_fold_ratio"),
    ("viewability", "viewability"),
    ("coverage", "coverage"),
]


def _n(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value) if value == value else 0.0
    s = str(value).strip()
    if not s:
        return 0.0
    s = re.sub(r"[%\s]", "", s)
    has_dot = "." in s
    has_comma = "," in s
    if has_dot and has_comma:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif has_comma:
        parts = s.split(",")
        if len(parts) > 1 and all(re.fullmatch(r"\d{3}", p) for p in parts[1:]):
            s = "".join(parts)
        else:
            s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _normalize_header(h: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (h or "").lower())


def _slug_metric_key(raw: str) -> str:
    s = re.sub(r"[^a-zA-Z0-9]+", "_", (raw or "").strip()).strip("_").lower()
    return s[:80] or "metric"


def _resolve_field(norm: str, raw: str) -> str | None:
    if norm in HEADERS_MAP:
        return HEADERS_MAP[norm]
    for needle, field in _HEADER_SUBSTRING_RULES:
        if needle in norm:
            return field
    return None


def _map_header_row(raw_headers: list[Any]) -> tuple[dict[str, int], list[tuple[str, int]]]:
    col_map: dict[str, int] = {}
    extras: list[tuple[str, int]] = []
    for i, raw in enumerate(raw_headers):
        label = str(raw or "").strip()
        norm = _normalize_header(label)
        field = _resolve_field(norm, label)
        if field:
            if field not in col_map:
                col_map[field] = i
        elif label:
            extras.append((label, i))
    return col_map, extras


def _excel_serial_to_date(val: Any) -> date | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        serial = float(val)
    except (TypeError, ValueError):
        s = str(val).strip()
        m = re.match(r"^(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2,4})$", s)
        if not m:
            return None
        y = int(m.group(3))
        if y < 100:
            y += 2000
        return date(y, int(m.group(2)), int(m.group(1)))
    if serial < 1000:
        return None
    base = datetime(1899, 12, 30)
    return (base + timedelta(days=serial)).date()


def _month_from_serial(val: Any) -> str:
    d = _excel_serial_to_date(val)
    return d.strftime("%Y-%m") if d else ""


@dataclass(frozen=True)
class AdStream:
    key: str
    project: str
    branch: str
    label: str
    channel: str
    default_surface: str


AD_STREAMS: tuple[AdStream, ...] = (
    AdStream("doviz:desktop", "doviz", "desktop", "Döviz · Desktop", "dovizcom", "web"),
    AdStream("doviz:mweb", "doviz", "mweb", "Döviz · Mobil Web", "dovizcom", "mweb"),
    AdStream("doviz:ios", "doviz", "ios", "Döviz · iOS", "ios", "ios_app"),
    AdStream("doviz:android", "doviz", "android", "Döviz · Android", "android", "android_app"),
    AdStream("sinemalar:desktop", "sinemalar", "desktop", "Sinemalar · Desktop", "sinemalar", "web"),
    AdStream("sinemalar:mweb", "sinemalar", "mweb", "Sinemalar · Mobil Web", "sinemalar", "mweb"),
)

_STREAM_BY_KEY = {s.key: s for s in AD_STREAMS}


def detect_stream(filename: str) -> AdStream | None:
    """Dosya adından proje + dal (6 akış)."""
    low = (filename or "").lower().replace(" ", "")
    if "m.sinemalar" in low or "m_sinemalar" in low:
        return _STREAM_BY_KEY["sinemalar:mweb"]
    if "sinemalar" in low and ("desktop" in low or low.startswith("sinemalardesktop")):
        return _STREAM_BY_KEY["sinemalar:desktop"]
    if "m.doviz" in low or "m_doviz" in low or "mdovizcom" in low:
        return _STREAM_BY_KEY["doviz:mweb"]
    if "doviz_ios" in low or "doviz-ios" in low:
        return _STREAM_BY_KEY["doviz:ios"]
    if "doviz_android" in low or "doviz-android" in low:
        return _STREAM_BY_KEY["doviz:android"]
    if "dovizcom" in low or "doviz.com" in low:
        return _STREAM_BY_KEY["doviz:desktop"]
    return None


def _report_period_rank(filename: str) -> int:
    """2025 dosyası=1, 2026 güncel dosyası=2 (son import üstüne yazar)."""
    low = (filename or "").lower()
    if (
        "_2_report" in low
        or "com2_report" in low
        or "desktop_2_report" in low
        or "_2_" in low and "report" in low
    ):
        return 2
    if (
        "_1_report" in low
        or "com1_report" in low
        or "desktop_1_report" in low
        or "_1_" in low and "report" in low
        or "ios_1" in low
        or "android_1" in low
    ):
        return 1
    return 9


def _bulk_sort_key(item: tuple[bytes, str]) -> tuple[str, int, str]:
    _data, name = item
    stream = detect_stream(name)
    sk = stream.key if stream else f"zzz:{name}"
    return (sk, _report_period_rank(name), name.lower())


def _detect_channel(filename: str) -> str:
    stream = detect_stream(filename)
    if stream:
        return stream.channel
    low = (filename or "").lower()
    if "dovizcom" in low or "doviz.com" in low:
        return "dovizcom"
    if "android" in low:
        return "android"
    if "ios" in low:
        return "ios"
    if "web" in low:
        return "web"
    return "other"


def _detect_surface(ad_unit: str, channel: str) -> str:
    u = (ad_unit or "").lower()
    if u.startswith("m_"):
        return "mweb"
    if u.startswith("web_"):
        return "web"
    if "app-ios" in u or "doviz_ios" in u or "_ios_" in u or u.startswith("ios_"):
        return "ios_app"
    if "android" in u or "app-android" in u:
        return "android_app"
    if channel == "android":
        return "android_app"
    if channel == "ios":
        return "ios_app"
    if channel == "dovizcom":
        if u.startswith("m_"):
            return "mweb"
        if u.startswith("web_"):
            return "web"
        return "site"
    return "unknown"


def _platform_from_surface(surface: str, channel: str) -> str:
    if surface in ("web", "mweb", "site"):
        return "web"
    if surface in ("android_app", "ios_app") or channel in ("android", "ios"):
        return "app"
    if channel == "dovizcom":
        return "web"
    return channel or "other"


def _row_fingerprint(
    *,
    report_date: date,
    ad_unit: str,
    income_type: str,
    project: str,
    branch: str,
) -> str:
    """Dal içinde tekilleştirme — kaynak dosya dahil değil (güncelleme duplike etmez)."""
    raw = f"{project}|{branch}|{report_date.isoformat()}|{ad_unit}|{income_type}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _dict_from_mapped(
    mapped: dict[str, Any],
    source_file: str,
    *,
    project: str,
    branch: str,
    channel: str,
    platform: str,
    surface: str,
    extra_metrics: dict[str, Any] | None = None,
) -> dict[str, Any] | None:
    ad_unit = (mapped.get("ad_unit") or "").strip()
    income_type = (mapped.get("income_type") or "").strip()
    rd = mapped.get("report_date")
    if not isinstance(rd, date):
        rd = _excel_serial_to_date(rd)
    if not isinstance(rd, date):
        rd = _excel_serial_to_date(mapped.get("month"))
    if not ad_unit or not income_type or not rd:
        return None
    month_key = mapped.get("month_key") or _month_from_serial(mapped.get("month")) or rd.strftime("%Y-%m")
    fp = _row_fingerprint(
        report_date=rd,
        ad_unit=ad_unit,
        income_type=income_type,
        project=project,
        branch=branch,
    )
    extras = dict(extra_metrics or {})
    for key in list(extras.keys()):
        if key in _CORE_FIELDS:
            extras.pop(key, None)
    row: dict[str, Any] = {
        "fingerprint": fp,
        "source_file": source_file,
        "project": project,
        "branch": branch,
        "platform": platform,
        "channel": channel,
        "surface": surface,
        "ad_unit": ad_unit[:500],
        "month_key": month_key[:7],
        "report_date": rd,
        "income_type": income_type[:120],
        "ad_request": _n(mapped.get("ad_request")),
        "matched_request": _n(mapped.get("matched_request")),
        "impression": _n(mapped.get("impression")),
        "click": _n(mapped.get("click")),
        "ad_request_ecpm": _n(mapped.get("ad_request_ecpm")),
        "ad_impression_ecpm": _n(mapped.get("ad_impression_ecpm")),
        "ctr": _n(mapped.get("ctr")),
        "coverage": _n(mapped.get("coverage")),
        "viewability": _n(mapped.get("viewability")),
        "net_revenue": _n(mapped.get("net_revenue")),
    }
    for opt in _OPTIONAL_EXTRA_FIELDS:
        if opt in mapped and mapped.get(opt) not in (None, ""):
            extras[opt] = _n(mapped.get(opt))
    row["extra_metrics"] = json.dumps(extras, ensure_ascii=False)
    return row


def _row_from_values(
    values: tuple[Any, ...],
    col_map: dict[str, int],
    extras_idx: list[tuple[str, int]],
    *,
    source_file: str,
    stream: AdStream | None,
    channel: str,
) -> dict[str, Any] | None:
    mapped: dict[str, Any] = {}
    extra_metrics: dict[str, Any] = {}
    for field, idx in col_map.items():
        if idx < len(values):
            mapped[field] = values[idx]
    if "month" in col_map and col_map["month"] < len(values):
        mapped["month_key"] = _month_from_serial(values[col_map["month"]])
    for label, idx in extras_idx:
        if idx < len(values):
            norm = _normalize_header(label)
            field = _resolve_field(norm, label)
            if field in _OPTIONAL_EXTRA_FIELDS:
                mapped[field] = values[idx]
                continue
            key = _slug_metric_key(label)
            val = values[idx]
            extra_metrics[key] = _n(val) if isinstance(val, (int, float, str)) else val
    ad_unit = (mapped.get("ad_unit") or "").strip()
    if stream:
        project, branch = stream.project, stream.branch
        channel = stream.channel
        surface = _detect_surface(ad_unit, channel)
        if surface == "unknown" and stream.default_surface:
            surface = stream.default_surface
    else:
        project, branch = "", ""
        surface = _detect_surface(ad_unit, channel)
    platform = _platform_from_surface(surface, channel)
    return _dict_from_mapped(
        mapped,
        source_file,
        project=project,
        branch=branch,
        channel=channel,
        platform=platform,
        surface=surface,
        extra_metrics=extra_metrics,
    )


def _locate_header_in_rows(
    rows_iter: Iterator[tuple[Any, ...]],
) -> tuple[dict[str, int], list[tuple[str, int]], Iterator[tuple[Any, ...]], list[str]] | None:
    """Ad Manager exportlarında başlık genelde 1. satırda değil; ilk N satırı tarar."""
    for i, row in enumerate(rows_iter):
        if i >= _HEADER_SCAN_MAX_ROWS:
            break
        if not row or not any(c is not None and str(c).strip() for c in row):
            continue
        col_map, extras_idx = _map_header_row(list(row))
        if "ad_unit" in col_map and "income_type" in col_map:
            labels = [str(h or "").strip() for h in row if str(h or "").strip()]
            return col_map, extras_idx, rows_iter, labels
    return None


def iter_xlsx_rows(data: bytes, *, filename: str = "upload.xlsx") -> Iterator[dict[str, Any]]:
    wb = _load_workbook_bytes(data)
    try:
        ws = wb.active
        located = _locate_header_in_rows(ws.iter_rows(values_only=True))
        if not located:
            LOGGER.warning(
                "Ad xlsx: %s — ilk %s satırda Ad Unit + Income Type başlığı yok",
                filename,
                _HEADER_SCAN_MAX_ROWS,
            )
            return
        col_map, extras_idx, rows_iter, _header_labels = located
        stream = detect_stream(filename)
        channel = stream.channel if stream else _detect_channel(filename)
        for row in rows_iter:
            if not row:
                continue
            item = _row_from_values(
                row, col_map, extras_idx, source_file=filename, stream=stream, channel=channel
            )
            if item:
                yield item
    finally:
        wb.close()


def parse_xlsx_bytes(data: bytes, *, filename: str = "upload.xlsx") -> list[dict[str, Any]]:
    return list(iter_xlsx_rows(data, filename=filename))


def parse_csv_text(text: str, *, filename: str = "upload.csv") -> list[dict[str, Any]]:
    raw = (text or "").strip()
    if not raw:
        return []
    lines = [ln for ln in raw.splitlines() if ln.strip()]
    if len(lines) < 2:
        return []
    delim = ","
    if lines[0].count(";") > lines[0].count(","):
        delim = ";"
    reader = csv.reader(lines, delimiter=delim)
    header_row = next(reader, None)
    if not header_row:
        return []
    col_map, extras_idx = _map_header_row(header_row)
    if "ad_unit" not in col_map or "income_type" not in col_map:
        return []
    stream = detect_stream(filename)
    channel = stream.channel if stream else _detect_channel(filename)
    out: list[dict[str, Any]] = []
    for cols in reader:
        item = _row_from_values(
            tuple(cols), col_map, extras_idx, source_file=filename, stream=stream, channel=channel
        )
        if item:
            out.append(item)
    return out


def _upsert_catalog(
    db: Session,
    filename: str,
    channel: str,
    columns: list[str],
    row_count: int,
    *,
    stream: AdStream | None = None,
) -> None:
    payload = json.dumps(columns, ensure_ascii=False)
    row = db.execute(
        select(AdReportCatalog).where(AdReportCatalog.source_file == filename)
    ).scalars().first()
    ch = stream.key if stream else channel
    if row:
        row.channel = ch
        row.columns_json = payload
        row.row_count = row_count
        row.imported_at = datetime.utcnow()
    else:
        db.add(
            AdReportCatalog(
                source_file=filename,
                channel=ch,
                columns_json=payload,
                row_count=row_count,
            )
        )


_UPSERT_UPDATE_COLS = (
    "source_file",
    "project",
    "branch",
    "platform",
    "channel",
    "surface",
    "month_key",
    "ad_request",
    "matched_request",
    "impression",
    "click",
    "ad_request_ecpm",
    "ad_impression_ecpm",
    "ctr",
    "coverage",
    "viewability",
    "net_revenue",
    "extra_metrics",
)


def _flush_batch(db: Session, batch: list[dict[str, Any]]) -> tuple[int, int]:
    """Upsert: aynı gün+ad_unit+gelir tipi güncellenir (duplike satır oluşmaz)."""
    if not batch:
        return 0, 0
    if _IS_PG:
        stmt = pg_insert(AdReportRow).values(batch)
        stmt = stmt.on_conflict_do_update(
            index_elements=["fingerprint"],
            set_={col: getattr(stmt.excluded, col) for col in _UPSERT_UPDATE_COLS},
        )
        res = db.execute(stmt)
        rc = res.rowcount
        affected = len(batch) if rc is None or rc < 0 else int(rc)
        db.flush()
        return affected, 0
    from sqlalchemy.dialects.sqlite import insert as sqlite_insert

    stmt = sqlite_insert(AdReportRow).values(batch)
    stmt = stmt.on_conflict_do_update(
        index_elements=["fingerprint"],
        set_={col: stmt.excluded[col] for col in _UPSERT_UPDATE_COLS},
    )
    res = db.execute(stmt)
    rc = res.rowcount
    affected = len(batch) if rc is None or rc < 0 else int(rc)
    db.flush()
    return affected, 0


def import_rows(
    db: Session,
    rows: Iterable[dict[str, Any]],
    *,
    commit: bool = True,
    progress_cb: ProgressCallback | None = None,
    progress_every: int = 800,
    row_estimate: int | None = None,
) -> dict[str, int]:
    inserted = 0
    skipped = 0
    parsed = 0
    batch: list[dict[str, Any]] = []

    def _emit(phase: str, **extra: Any) -> None:
        if not progress_cb:
            return
        payload: dict[str, Any] = {
            "phase": phase,
            "parsed": parsed,
            "inserted": inserted,
            "skipped": skipped,
        }
        if row_estimate and row_estimate > 0:
            payload["row_estimate"] = row_estimate
            payload["pct"] = min(99, int(100 * parsed / row_estimate))
        payload.update(extra)
        progress_cb(payload)

    try:
        _emit("parse_start")
        for r in rows:
            parsed += 1
            batch.append(r)
            if progress_every > 0 and parsed % progress_every == 0:
                _emit("parsing")
            if len(batch) >= _IMPORT_BATCH_SIZE:
                ins, sk = _flush_batch(db, batch)
                inserted += ins
                skipped += sk
                batch.clear()
                _emit("db_flush")
        if batch:
            ins, sk = _flush_batch(db, batch)
            inserted += ins
            skipped += sk
            _emit("db_flush")
        if commit:
            db.commit()
        _emit("parse_done")
    except Exception:
        db.rollback()
        raise
    return {
        "inserted": inserted,
        "skipped": skipped,
        "parsed": parsed,
        "total": count_rows(db),
    }


def _estimate_xlsx_data_rows(data: bytes) -> int:
    try:
        wb = _load_workbook_bytes(data)
        try:
            n = int(wb.active.max_row or 0)
            return max(0, n - 1)
        finally:
            wb.close()
    except Exception:
        return 0


def import_upload_file(
    db: Session,
    data: bytes,
    *,
    filename: str,
    commit: bool = True,
    progress_cb: ProgressCallback | None = None,
) -> dict[str, Any]:
    low = filename.lower()
    stream = detect_stream(filename)
    channel = stream.channel if stream else _detect_channel(filename)
    try:
        if low.endswith(".xlsx") or low.endswith(".xlsm"):
            columns: list[str] = []
            wb = _load_workbook_bytes(data)
            try:
                rows_it = wb.active.iter_rows(values_only=True)
                located = _locate_header_in_rows(rows_it)
                if located:
                    _col_map, _extras, _, columns = located
                else:
                    preview = next(wb.active.iter_rows(values_only=True), None)
                    columns = [
                        str(h or "").strip() for h in (preview or []) if str(h or "").strip()
                    ][:20]
            finally:
                wb.close()
            row_est = _estimate_xlsx_data_rows(data)
            result = import_rows(
                db,
                iter_xlsx_rows(data, filename=filename),
                commit=False,
                progress_cb=progress_cb,
                row_estimate=row_est or None,
            )
            if not result.get("parsed"):
                result["parse_error"] = (
                    "Excel başlığı okunamadı (ilk satırda veya sonraki 30 satırda "
                    "'Ad Unit' / 'Reklam birimi' ve 'Income Type' / 'Gelir tipi' aranır). "
                    f"İlk satır önizleme: {columns[:8]!r}"
                )
            LOGGER.info(
                "Ad xlsx import: %s stream=%s parsed=%s",
                filename,
                stream.key if stream else "?",
                result.get("parsed"),
            )
        elif low.endswith(".csv") or low.endswith(".txt"):
            text = data.decode("utf-8", errors="replace")
            lines = [ln for ln in text.splitlines() if ln.strip()]
            columns = []
            if lines:
                delim = ";" if lines[0].count(";") > lines[0].count(",") else ","
                columns = [c.strip() for c in lines[0].split(delim)]
            result = import_rows(
                db,
                parse_csv_text(text, filename=filename),
                commit=False,
                progress_cb=progress_cb,
                row_estimate=max(0, len(lines) - 1) if lines else None,
            )
        else:
            raise ValueError("Yalnızca .xlsx veya .csv desteklenir")
        _upsert_catalog(db, filename, channel, columns, result.get("parsed", 0), stream=stream)
        if commit:
            db.commit()
    except Exception:
        db.rollback()
        raise
    result["filename"] = filename
    result["channel"] = channel
    result["columns"] = columns
    if stream:
        result["stream_key"] = stream.key
        result["project"] = stream.project
        result["branch"] = stream.branch
        result["period"] = _report_period_rank(filename)
    else:
        result["stream_key"] = None
        result["warning"] = "Dosya adından dal tanınamadı"
    return result


def build_upload_batch_summary(per_file: list[dict[str, Any]]) -> dict[str, Any]:
    """Yükleme sonrası özet: başarılı, hatalı, boş, çakışan dönem, tanınmayan dal."""
    ok_files: list[str] = []
    failed: list[dict[str, str]] = []
    empty: list[dict[str, str]] = []
    unknown_files: list[str] = []
    warnings: list[str] = []
    total_skipped = 0
    stream_period_files: dict[tuple[str, str], list[str]] = {}
    name_counts: dict[str, int] = {}

    for item in per_file:
        fn = (item.get("filename") or "?").strip()
        name_counts[fn] = name_counts.get(fn, 0) + 1
        if item.get("error"):
            failed.append({"filename": fn, "reason": str(item["error"])})
            continue
        parsed = int(item.get("parsed") or 0)
        skipped = int(item.get("skipped") or 0)
        total_skipped += skipped
        if item.get("warning"):
            unknown_files.append(fn)
            warnings.append(f"{fn}: dal/proje dosya adından çıkarılamadı")
        if item.get("parse_error"):
            empty.append({"filename": fn, "reason": str(item["parse_error"])})
        elif parsed <= 0:
            empty.append({"filename": fn, "reason": "Dosyadan satır okunamadı (eksik veya hatalı içerik)"})
        else:
            ok_files.append(fn)
        sk = item.get("stream_key")
        period = str(item.get("period") or "")
        if sk:
            stream_period_files.setdefault((str(sk), period), []).append(fn)

    duplicate_names = [n for n, c in name_counts.items() if c > 1]
    overlapping_periods = [
        {"stream_key": k[0], "period": k[1], "files": v}
        for k, v in stream_period_files.items()
        if len(v) > 1
    ]
    duplicate_period_notes: list[str] = []
    for ov in overlapping_periods:
        duplicate_period_notes.append(
            f"Aynı dal/dönem için birden fazla dosya (veri üst üste birleşir): {', '.join(ov['files'])}"
        )

    file_count = len(per_file)
    integrated_rows = sum(int(x.get("parsed") or 0) for x in per_file if not x.get("error"))
    inserted_rows = sum(int(x.get("inserted") or 0) for x in per_file if not x.get("error"))
    has_errors = bool(failed) or bool(empty)
    has_warnings = bool(unknown_files) or bool(duplicate_names) or bool(duplicate_period_notes)
    nothing_imported = file_count > 0 and integrated_rows <= 0

    return {
        "file_count": file_count,
        "ok_count": len(ok_files),
        "failed_count": len(failed),
        "empty_count": len(empty),
        "unknown_count": len(unknown_files),
        "ok_files": ok_files,
        "failed": failed,
        "empty": empty,
        "unknown_files": unknown_files,
        "warnings": warnings,
        "duplicate_filenames": duplicate_names,
        "overlapping_periods": overlapping_periods,
        "duplicate_period_notes": duplicate_period_notes,
        "total_skipped": total_skipped,
        "integrated_rows": integrated_rows,
        "inserted_rows": inserted_rows,
        "has_errors": has_errors,
        "has_warnings": has_warnings,
        "nothing_imported": nothing_imported,
        "all_ok": file_count > 0 and not has_errors and len(ok_files) == file_count,
    }


def iter_bulk_import_events(
    files: list[tuple[bytes, str]],
) -> Iterator[dict[str, Any]]:
    """Gerçek içe aktarma aşamaları — NDJSON stream için."""
    ordered = sorted(files, key=_bulk_sort_key)
    total_files = len(ordered)
    yield {
        "phase": "batch_start",
        "total_files": total_files,
        "pct": 0,
    }
    per_file: list[dict[str, Any]] = []
    unknown: list[str] = []
    total_inserted = 0
    total_parsed = 0
    for idx, (data, name) in enumerate(ordered, start=1):
        if not data:
            per_file.append({"filename": name, "error": "boş dosya"})
            yield {
                "phase": "file_error",
                "file_index": idx,
                "total_files": total_files,
                "filename": name,
                "error": "boş dosya",
                "pct": int(100 * (idx - 1) / max(total_files, 1)),
            }
            continue
        yield {
            "phase": "file_start",
            "file_index": idx,
            "total_files": total_files,
            "filename": name,
            "bytes": len(data),
            "pct": int(100 * (idx - 1) / max(total_files, 1)),
        }
        pending: list[dict[str, Any]] = []

        def _file_progress(ev: dict[str, Any]) -> None:
            ev = dict(ev)
            ev["file_index"] = idx
            ev["total_files"] = total_files
            ev["filename"] = name
            if total_files:
                base = int(100 * (idx - 1) / total_files)
                span = int(100 / total_files)
                inner = int(ev.get("pct") or 0)
                ev["pct"] = min(99, base + int(span * inner / 100))
            pending.append(ev)

        with SessionLocal() as db:
            try:
                out = import_upload_file(
                    db,
                    data,
                    filename=name,
                    commit=True,
                    progress_cb=_file_progress,
                )
                for ev in pending:
                    yield ev
                per_file.append(out)
                total_inserted += int(out.get("inserted") or 0)
                total_parsed += int(out.get("parsed") or 0)
                if not out.get("stream_key"):
                    unknown.append(name)
                yield {
                    "phase": "file_done",
                    "file_index": idx,
                    "total_files": total_files,
                    "filename": name,
                    "parsed": out.get("parsed", 0),
                    "inserted": out.get("inserted", 0),
                    "stream_key": out.get("stream_key"),
                    "pct": int(100 * idx / max(total_files, 1)),
                }
            except Exception as exc:  # noqa: BLE001
                db.rollback()
                per_file.append({"filename": name, "error": str(exc)})
                yield {
                    "phase": "file_error",
                    "file_index": idx,
                    "total_files": total_files,
                    "filename": name,
                    "error": str(exc),
                    "pct": int(100 * idx / max(total_files, 1)),
                }
    with SessionLocal() as db:
        total_rows = count_rows(db)
    summary = build_upload_batch_summary(per_file)
    yield {
        "phase": "batch_done",
        "pct": 100,
        "file_count": total_files,
        "parsed": total_parsed,
        "inserted": total_inserted,
        "total": total_rows,
        "unknown_files": unknown,
        "files": per_file,
        "summary": summary,
    }


def import_upload_files_bulk(files: list[tuple[bytes, str]]) -> dict[str, Any]:
    """Çoklu xlsx: dosya başına ayrı transaction (hata sonrası invalid session kalmaz)."""
    ordered = sorted(files, key=_bulk_sort_key)
    per_file: list[dict[str, Any]] = []
    unknown: list[str] = []
    total_inserted = 0
    total_parsed = 0
    for data, name in ordered:
        if not data:
            per_file.append({"filename": name, "error": "boş dosya"})
            continue
        with SessionLocal() as db:
            try:
                out = import_upload_file(db, data, filename=name, commit=True)
                per_file.append(out)
                total_inserted += int(out.get("inserted") or 0)
                total_parsed += int(out.get("parsed") or 0)
                if not out.get("stream_key"):
                    unknown.append(name)
                LOGGER.info(
                    "Ad bulk file ok: %s stream=%s parsed=%s",
                    name,
                    out.get("stream_key"),
                    out.get("parsed"),
                )
            except Exception as exc:  # noqa: BLE001
                db.rollback()
                LOGGER.warning("Ad bulk file failed: %s — %s", name, exc)
                per_file.append({"filename": name, "error": str(exc)})
    with SessionLocal() as db:
        total_rows = count_rows(db)
    LOGGER.info(
        "Ad bulk import done: files=%s parsed=%s total_rows=%s",
        len(ordered),
        total_parsed,
        total_rows,
    )
    summary = build_upload_batch_summary(per_file)
    return {
        "files": per_file,
        "file_count": len(ordered),
        "inserted": total_inserted,
        "parsed": total_parsed,
        "total": total_rows,
        "unknown_files": unknown,
        "summary": summary,
    }


def count_rows(db: Session) -> int:
    return int(db.scalar(select(func.count()).select_from(AdReportRow)) or 0)


def reset_all(db: Session) -> dict[str, int]:
    db.execute(delete(AdReportRow))
    db.execute(delete(AdReportCatalog))
    db.commit()
    return {"total": 0}


def date_bounds(db: Session) -> dict[str, str | None]:
    row = db.execute(
        select(
            func.min(AdReportRow.report_date),
            func.max(AdReportRow.report_date),
        )
    ).one()
    dmin, dmax = row[0], row[1]
    return {
        "min_date": dmin.isoformat() if dmin else None,
        "max_date": dmax.isoformat() if dmax else None,
    }


def facets(db: Session) -> dict[str, Any]:
    income = [
        r[0]
        for r in db.execute(
            select(AdReportRow.income_type).distinct().order_by(AdReportRow.income_type)
        ).all()
        if r[0]
    ]
    platforms = [
        r[0]
        for r in db.execute(
            select(AdReportRow.platform).distinct().order_by(AdReportRow.platform)
        ).all()
        if r[0]
    ]
    sources = [
        r[0]
        for r in db.execute(
            select(AdReportRow.source_file).distinct().order_by(AdReportRow.source_file.desc())
        ).all()
        if r[0]
    ]
    channels = [
        r[0]
        for r in db.execute(
            select(AdReportRow.channel).distinct().order_by(AdReportRow.channel)
        ).all()
        if r[0]
    ]
    surfaces = [
        r[0]
        for r in db.execute(
            select(AdReportRow.surface).distinct().order_by(AdReportRow.surface)
        ).all()
        if r[0]
    ]
    catalogs = db.execute(select(AdReportCatalog)).scalars().all()
    extra_columns: list[str] = []
    seen_cols: set[str] = set()
    for cat in catalogs:
        try:
            cols = json.loads(cat.columns_json or "[]")
        except json.JSONDecodeError:
            cols = []
        for c in cols:
            norm = _normalize_header(str(c))
            if norm in HEADERS_MAP or _resolve_field(norm, str(c)):
                continue
            slug = _slug_metric_key(str(c))
            if slug not in seen_cols:
                seen_cols.add(slug)
                extra_columns.append(str(c))
    stream_stats = db.execute(
        select(
            AdReportRow.project,
            AdReportRow.branch,
            func.min(AdReportRow.report_date),
            func.max(AdReportRow.report_date),
            func.count(),
        )
        .where(AdReportRow.project != "", AdReportRow.branch != "")
        .group_by(AdReportRow.project, AdReportRow.branch)
    ).all()
    stats_map = {(r[0], r[1]): r for r in stream_stats}
    streams_out: list[dict[str, Any]] = []
    kpi_union: set[str] = set()
    for meta in AD_STREAMS:
        hit = stats_map.get((meta.project, meta.branch))
        stream_kpis: list[str] = []
        if hit and hit[4]:
            # KPI uygunluğu facets'te ağır JSON taraması yapılmaz — query_summary döner.
            stream_kpis = []
        streams_out.append(
            {
                "key": meta.key,
                "project": meta.project,
                "branch": meta.branch,
                "label": meta.label,
                "channel": meta.channel,
                "default_surface": meta.default_surface,
                "min_date": hit[2].isoformat() if hit and hit[2] else None,
                "max_date": hit[3].isoformat() if hit and hit[3] else None,
                "row_count": int(hit[4] or 0) if hit else 0,
                "has_data": bool(hit and hit[4]),
                "available_kpis": stream_kpis,
            }
        )

    file_rows = db.execute(
        select(
            AdReportRow.source_file,
            AdReportRow.project,
            AdReportRow.branch,
            func.count(),
        )
        .where(AdReportRow.source_file != "")
        .group_by(AdReportRow.source_file, AdReportRow.project, AdReportRow.branch)
    ).all()
    import_audit: list[dict[str, Any]] = []
    for fn, proj, br, cnt in file_rows:
        expected = detect_stream(fn)
        if expected and (expected.project != proj or expected.branch != br):
            import_audit.append(
                {
                    "source_file": fn,
                    "expected_stream": expected.key,
                    "actual_project": proj,
                    "actual_branch": br,
                    "row_count": int(cnt or 0),
                }
            )

    return {
        "streams": streams_out,
        "kpi_union": sorted(kpi_union, key=lambda k: {x: i for i, x in enumerate(KPI_METRIC_KEYS)}.get(k, 999)),
        "import_audit": import_audit,
        "income_types": income,
        "platforms": platforms,
        "channels": channels,
        "surfaces": surfaces,
        "source_files": sources,
        "extra_columns": extra_columns,
        "imports": [
            {
                "source_file": c.source_file,
                "channel": c.channel,
                "row_count": c.row_count,
                "columns": json.loads(c.columns_json or "[]"),
            }
            for c in catalogs
        ],
        **date_bounds(db),
        "total_rows": count_rows(db),
    }


def _parse_filter_list(raw: str | None) -> list[str]:
    if not raw:
        return []
    return [x.strip() for x in raw.split(",") if x.strip()]


def _apply_filters(
    q,
    *,
    start: str | None,
    end: str | None,
    income_types: list[str],
    ad_units: list[str],
    platforms: list[str],
    channels: list[str],
    surfaces: list[str],
    sources: list[str],
    search: str | None,
    project: str | None = None,
    branch: str | None = None,
):
    if project:
        q = q.where(AdReportRow.project == project)
    if branch:
        q = q.where(AdReportRow.branch == branch)
    if start:
        try:
            q = q.where(AdReportRow.report_date >= date.fromisoformat(start[:10]))
        except ValueError:
            pass
    if end:
        try:
            q = q.where(AdReportRow.report_date <= date.fromisoformat(end[:10]))
        except ValueError:
            pass
    if income_types:
        q = q.where(AdReportRow.income_type.in_(income_types))
    if ad_units:
        q = q.where(AdReportRow.ad_unit.in_(ad_units))
    if platforms:
        q = q.where(AdReportRow.platform.in_(platforms))
    if channels:
        q = q.where(AdReportRow.channel.in_(channels))
    if surfaces:
        q = q.where(AdReportRow.surface.in_(surfaces))
    if sources:
        q = q.where(AdReportRow.source_file.in_(sources))
    if search:
        q = q.where(AdReportRow.ad_unit.ilike(f"%{search}%"))
    return q


_OPTIONAL_EXTRA_FIELDS = frozenset(
    {
        "empower_pageview",
        "empower_unique_visitor",
        "pageview_ecpm",
        "unique_visitor_ecpm",
        "above_the_fold_ratio",
    }
)

# extra_metrics JSON anahtarları export başlıklarına göre farklı yazılabilir
_EXTRA_JSON_ALIASES: dict[str, tuple[str, ...]] = {
    "empower_pageview": ("empower_pageview", "empower_pageviews", "empowerpageview"),
    "empower_unique_visitor": (
        "empower_unique_visitor",
        "empower_unique_visitors",
        "empoweruniquevisitor",
    ),
    "pageview_ecpm": ("pageview_ecpm", "pageviewecpm"),
    "unique_visitor_ecpm": ("unique_visitor_ecpm", "uniquevisitorecpm"),
    "above_the_fold_ratio": (
        "above_the_fold_ratio",
        "above_the_fold",
        "abovethefoldratio",
        "abovethefold",
    ),
}

KPI_METRIC_KEYS = (
    "empower_pageview",
    "empower_unique_visitor",
    "ad_request",
    "matched_request",
    "impression",
    "click",
    "net_revenue",
    "ad_request_ecpm",
    "pageview_ecpm",
    "unique_visitor_ecpm",
    "ad_ecpm",
    "viewability_pct",
    "above_the_fold_ratio_pct",
    "ctr_pct",
    "coverage_pct",
)

COMPARE_MODES = frozenset({"previous_period", "previous_year", "custom"})


def _parse_iso_date(raw: str | None) -> date | None:
    if not raw:
        return None
    try:
        return date.fromisoformat(str(raw)[:10])
    except ValueError:
        return None


def resolve_compare_range(
    primary_start: str | None,
    primary_end: str | None,
    mode: str | None,
    custom_start: str | None = None,
    custom_end: str | None = None,
) -> tuple[str | None, str | None]:
    """Karşılaştırma dönemi ISO tarihleri (primary ile aynı uzunlukta veya özel aralık)."""
    if not mode or mode not in COMPARE_MODES:
        return None, None
    if mode == "custom":
        cs = _parse_iso_date(custom_start)
        ce = _parse_iso_date(custom_end)
        if cs and ce:
            return cs.isoformat(), ce.isoformat()
        return None, None
    ps = _parse_iso_date(primary_start)
    pe = _parse_iso_date(primary_end)
    if not ps or not pe:
        return None, None
    span_days = (pe - ps).days + 1
    if mode == "previous_period":
        cmp_end = ps - timedelta(days=1)
        cmp_start = cmp_end - timedelta(days=span_days - 1)
        return cmp_start.isoformat(), cmp_end.isoformat()
    if mode == "previous_year":
        try:
            cmp_start = ps.replace(year=ps.year - 1)
            cmp_end = pe.replace(year=pe.year - 1)
        except ValueError:
            cmp_start = ps - timedelta(days=365)
            cmp_end = pe - timedelta(days=365)
        return cmp_start.isoformat(), cmp_end.isoformat()
    return None, None


def _kpi_delta(current: float, previous: float) -> dict[str, Any]:
    abs_d = current - previous
    if previous == 0:
        pct = 100.0 if current > 0 else (0.0 if current == 0 else -100.0)
    else:
        pct = (abs_d / abs(previous)) * 100.0
    return {
        "current": current,
        "compare": previous,
        "abs": round(abs_d, 6),
        "pct": round(pct, 2),
    }


def compute_kpi_deltas(primary: dict[str, Any], compare: dict[str, Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for key in KPI_METRIC_KEYS:
        if key in primary or key in compare:
            out[key] = _kpi_delta(float(primary.get(key) or 0), float(compare.get(key) or 0))
    return out


def _merge_breakdown(
    primary_rows: list[dict[str, Any]],
    compare_rows: list[dict[str, Any]],
    key_field: str,
    metrics: tuple[str, ...] = ("net_revenue", "impression"),
) -> list[dict[str, Any]]:
    cmap = {r.get(key_field): r for r in compare_rows if r.get(key_field) is not None}
    keys: list[str] = []
    seen: set[str] = set()
    for row in primary_rows + compare_rows:
        k = row.get(key_field)
        if k is None or k in seen:
            continue
        seen.add(str(k))
        keys.append(str(k))

    def _fmt_metric(name: str, val: float) -> float | int:
        if name == "net_revenue":
            return round(val, 2)
        if name == "impression":
            return int(val)
        return round(val, 3)

    merged: list[dict[str, Any]] = []
    for key in keys:
        prow = next((x for x in primary_rows if str(x.get(key_field)) == key), {})
        crow = cmap.get(key, {})
        item: dict[str, Any] = {key_field: key}
        for metric in metrics:
            pv = float(prow.get(metric) or 0)
            cv = float(crow.get(metric) or 0)
            item[metric] = _fmt_metric(metric, pv)
            item[f"{metric}_compare"] = _fmt_metric(metric, cv)
            item[f"{metric}_delta_pct"] = _kpi_delta(pv, cv)["pct"]
        merged.append(item)
    merged.sort(key=lambda x: float(x.get("net_revenue") or 0), reverse=True)
    return merged


def align_by_date_series(
    primary: list[dict[str, Any]],
    compare: list[dict[str, Any]],
    kpi_key: str,
) -> list[dict[str, Any]]:
    """İki dönemin günlük serisini gün indeksi ile hizalar (tarih bazlı trend karşılaştırması)."""
    n = max(len(primary), len(compare))
    aligned: list[dict[str, Any]] = []
    for i in range(n):
        p = primary[i] if i < len(primary) else {}
        c = compare[i] if i < len(compare) else {}
        pv = float(p.get(kpi_key) or 0)
        cv = float(c.get(kpi_key) or 0)
        aligned.append(
            {
                "day_index": i + 1,
                "primary_date": p.get("date"),
                "compare_date": c.get("date"),
                "primary": round(pv, 6),
                "compare": round(cv, 6),
                "delta_pct": _kpi_delta(pv, cv)["pct"],
            }
        )
    return aligned


def _attach_compare_block(
    primary: dict[str, Any],
    compare: dict[str, Any],
    *,
    range_start: str,
    range_end: str,
    mode: str,
) -> dict[str, Any]:
    pk = primary.get("kpis") or {}
    ck = compare.get("kpis") or {}
    return {
        "mode": mode,
        "range": {"start": range_start, "end": range_end},
        "kpis": ck,
        "deltas": compute_kpi_deltas(pk, ck),
        "by_date": compare.get("by_date") or [],
        "by_date_aligned": {
            k: align_by_date_series(primary.get("by_date") or [], compare.get("by_date") or [], k)
            for k in KPI_METRIC_KEYS
        },
        "by_income_type": _merge_breakdown(
            primary.get("by_income_type") or [],
            compare.get("by_income_type") or [],
            "income_type",
        ),
        "by_ad_unit": _merge_breakdown(
            primary.get("by_ad_unit") or [],
            compare.get("by_ad_unit") or [],
            "ad_unit",
        ),
        "by_month": _merge_breakdown(
            primary.get("by_month") or [],
            compare.get("by_month") or [],
            "month",
        ),
        "by_surface": _merge_breakdown(
            primary.get("by_surface") or [],
            compare.get("by_surface") or [],
            "surface",
        ),
        "by_channel": _merge_breakdown(
            primary.get("by_channel") or [],
            compare.get("by_channel") or [],
            "channel",
        ),
    }


def _norm_ratio_sql(col):
    from sqlalchemy import case

    return case((col > 1, col / 100.0), else_=col)


def _extra_json_expr(sub, key: str):
    """Tek JSON anahtarı → Float ifadesi (subquery satırı)."""
    from sqlalchemy import Float, cast

    if _IS_PG:
        from sqlalchemy.dialects.postgresql import JSONB

        safe_json = func.coalesce(func.nullif(sub.c.extra_metrics, ""), "{}")
        raw = cast(safe_json, JSONB)[key].astext
        return cast(func.nullif(raw, ""), Float)
    return cast(func.nullif(func.json_extract(sub.c.extra_metrics, f"$.{key}"), ""), Float)


def _extra_json_expr_multi(sub, canonical: str):
    """Bilinen alias'larla extra JSON alanı (satır başına coalesce)."""
    keys = _EXTRA_JSON_ALIASES.get(canonical, (canonical,))
    expr = _extra_json_expr(sub, keys[0])
    for alias in keys[1:]:
        expr = func.coalesce(expr, _extra_json_expr(sub, alias))
    return expr


def _sum_extra_multi(sub, canonical: str):
    return func.coalesce(func.sum(_extra_json_expr_multi(sub, canonical)), 0)


def _weighted_extra_ratio(sub, canonical: str, weight_col):
    """Oran metrikleri impression/ad_request ile ağırlıklı topla (0–100 normalize)."""
    val = _extra_json_expr_multi(sub, canonical)
    norm = _norm_ratio_sql(val)
    return func.coalesce(func.sum(norm * weight_col), 0)


def _has_extra_data(db: Session, q, canonical: str) -> bool:
    sub = q.subquery()
    total = db.scalar(select(_sum_extra_multi(sub, canonical)))
    return float(total or 0) > 0


def _derive_kpi_availability(
    *,
    net_rev: float,
    impr: float,
    clicks: float,
    ad_req: float,
    matched_req: float,
    empower_pv: float,
    empower_uv: float,
    view_w_sum: float,
    cov_w_sum: float,
    atf_w_sum: float,
    has_pageview_ecpm_extra: bool,
    has_uv_ecpm_extra: bool,
) -> list[str]:
    avail: list[str] = []
    if net_rev != 0:
        avail.append("net_revenue")
    if ad_req > 0:
        avail.append("ad_request")
        if net_rev != 0:
            avail.append("ad_request_ecpm")
    if matched_req > 0:
        avail.append("matched_request")
    if impr > 0:
        avail.extend(["impression", "ad_ecpm", "ctr_pct"])
    if clicks > 0 or impr > 0:
        avail.append("click")
    if empower_pv > 0:
        avail.append("empower_pageview")
        if net_rev != 0:
            avail.append("pageview_ecpm")
    elif has_pageview_ecpm_extra:
        avail.append("pageview_ecpm")
    if empower_uv > 0:
        avail.append("empower_unique_visitor")
        if net_rev != 0:
            avail.append("unique_visitor_ecpm")
    elif has_uv_ecpm_extra:
        avail.append("unique_visitor_ecpm")
    if impr > 0 and view_w_sum > 0:
        avail.append("viewability_pct")
    if ad_req > 0 and cov_w_sum > 0:
        avail.append("coverage_pct")
    if impr > 0 and atf_w_sum > 0:
        avail.append("above_the_fold_ratio_pct")
    # sıra KPI kartları ile uyumlu
    order = {k: i for i, k in enumerate(KPI_METRIC_KEYS)}
    return sorted(set(avail), key=lambda k: order.get(k, 999))


def scan_stream_kpis(db: Session, project: str, branch: str) -> list[str]:
    """Dal için DB'de gerçekten var olan KPI kaynaklarını tespit et (hızlı kolon + tek JSON turu)."""
    base = select(AdReportRow).where(
        AdReportRow.project == project,
        AdReportRow.branch == branch,
    )
    sub = base.subquery()
    norm_view = _norm_ratio_sql(sub.c.viewability)
    norm_cov = _norm_ratio_sql(sub.c.coverage)
    row = db.execute(
        select(
            func.coalesce(func.sum(sub.c.net_revenue), 0),
            func.coalesce(func.sum(sub.c.impression), 0),
            func.coalesce(func.sum(sub.c.click), 0),
            func.coalesce(func.sum(sub.c.ad_request), 0),
            func.coalesce(func.sum(sub.c.matched_request), 0),
            func.coalesce(_sum_extra_multi(sub, "empower_pageview"), 0),
            func.coalesce(_sum_extra_multi(sub, "empower_unique_visitor"), 0),
            func.coalesce(func.sum(norm_view * sub.c.impression), 0),
            func.coalesce(func.sum(norm_cov * sub.c.ad_request), 0),
            func.coalesce(_weighted_extra_ratio(sub, "above_the_fold_ratio", sub.c.impression), 0),
            func.coalesce(_sum_extra_multi(sub, "pageview_ecpm"), 0),
            func.coalesce(_sum_extra_multi(sub, "unique_visitor_ecpm"), 0),
        )
    ).one()
    vals = [float(x or 0) for x in row]
    return _derive_kpi_availability(
        net_rev=vals[0],
        impr=vals[1],
        clicks=vals[2],
        ad_req=vals[3],
        matched_req=vals[4],
        empower_pv=vals[5],
        empower_uv=vals[6],
        view_w_sum=vals[7],
        cov_w_sum=vals[8],
        atf_w_sum=vals[9],
        has_pageview_ecpm_extra=vals[10] > 0,
        has_uv_ecpm_extra=vals[11] > 0,
    )


def query_summary(
    db: Session,
    *,
    start: str | None = None,
    end: str | None = None,
    income_types: str | None = None,
    ad_units: str | None = None,
    platforms: str | None = None,
    channels: str | None = None,
    surfaces: str | None = None,
    sources: str | None = None,
    search: str | None = None,
    project: str | None = None,
    branch: str | None = None,
    compare_mode: str | None = None,
    compare_start: str | None = None,
    compare_end: str | None = None,
) -> dict[str, Any]:
    it = _parse_filter_list(income_types)
    au = _parse_filter_list(ad_units)
    pl = _parse_filter_list(platforms)
    ch = _parse_filter_list(channels)
    su = _parse_filter_list(surfaces)
    sf = _parse_filter_list(sources)

    base = select(AdReportRow)
    base = _apply_filters(
        base,
        start=start,
        end=end,
        income_types=it,
        ad_units=au,
        platforms=pl,
        channels=ch,
        surfaces=su,
        sources=sf,
        search=search,
        project=project,
        branch=branch,
    )
    sub = base.subquery()

    norm_view = _norm_ratio_sql(sub.c.viewability)
    norm_cov = _norm_ratio_sql(sub.c.coverage)

    totals = db.execute(
        select(
            func.coalesce(func.sum(sub.c.net_revenue), 0),
            func.coalesce(func.sum(sub.c.impression), 0),
            func.coalesce(func.sum(sub.c.click), 0),
            func.coalesce(func.sum(sub.c.ad_request), 0),
            func.coalesce(func.sum(sub.c.matched_request), 0),
            func.coalesce(_sum_extra_multi(sub, "empower_pageview"), 0),
            func.coalesce(_sum_extra_multi(sub, "empower_unique_visitor"), 0),
            func.coalesce(func.sum(norm_view * sub.c.impression), 0),
            func.coalesce(func.sum(norm_cov * sub.c.ad_request), 0),
            func.coalesce(_weighted_extra_ratio(sub, "above_the_fold_ratio", sub.c.impression), 0),
            func.coalesce(_sum_extra_multi(sub, "pageview_ecpm"), 0),
            func.coalesce(_sum_extra_multi(sub, "unique_visitor_ecpm"), 0),
        )
    ).one()
    (
        net_rev,
        impr,
        clicks,
        ad_req,
        matched_req,
        empower_pv,
        empower_uv,
        view_w_sum,
        cov_w_sum,
        atf_w_sum,
        pv_ecpm_sum,
        uv_ecpm_sum,
    ) = [float(x or 0) for x in totals]
    avg_ecpm = (net_rev / impr * 1000.0) if impr > 0 else 0.0
    ctr = (clicks / impr * 100.0) if impr > 0 else 0.0
    req_ecpm = (net_rev / ad_req * 1000.0) if ad_req > 0 else 0.0
    if empower_pv > 0:
        pv_ecpm = net_rev / empower_pv * 1000.0
    elif pv_ecpm_sum > 0:
        cnt_pv = db.scalar(
            select(func.count())
            .select_from(sub)
            .where(_extra_json_expr_multi(sub, "pageview_ecpm") > 0)
        )
        pv_ecpm = pv_ecpm_sum / float(cnt_pv or 1)
    else:
        pv_ecpm = 0.0
    if empower_uv > 0:
        uv_ecpm = net_rev / empower_uv * 1000.0
    elif uv_ecpm_sum > 0:
        cnt_uv = db.scalar(
            select(func.count())
            .select_from(sub)
            .where(_extra_json_expr_multi(sub, "unique_visitor_ecpm") > 0)
        )
        uv_ecpm = uv_ecpm_sum / float(cnt_uv or 1)
    else:
        uv_ecpm = 0.0
    coverage_pct = (cov_w_sum / ad_req * 100.0) if ad_req > 0 else 0.0
    viewability_pct = (view_w_sum / impr * 100.0) if impr > 0 else 0.0
    atf_pct = (atf_w_sum / impr * 100.0) if impr > 0 else 0.0

    has_pv_ecpm_extra = pv_ecpm_sum > 0
    has_uv_ecpm_extra = uv_ecpm_sum > 0
    kpi_available = _derive_kpi_availability(
        net_rev=net_rev,
        impr=impr,
        clicks=clicks,
        ad_req=ad_req,
        matched_req=matched_req,
        empower_pv=empower_pv,
        empower_uv=empower_uv,
        view_w_sum=view_w_sum,
        cov_w_sum=cov_w_sum,
        atf_w_sum=atf_w_sum,
        has_pageview_ecpm_extra=has_pv_ecpm_extra,
        has_uv_ecpm_extra=has_uv_ecpm_extra,
    )

    by_date = db.execute(
        select(
            sub.c.report_date,
            func.sum(sub.c.net_revenue),
            func.sum(sub.c.impression),
            func.sum(sub.c.click),
            func.sum(sub.c.ad_request),
            func.sum(sub.c.matched_request),
            _sum_extra_multi(sub, "empower_pageview"),
            _sum_extra_multi(sub, "empower_unique_visitor"),
            func.sum(norm_view * sub.c.impression),
            func.sum(norm_cov * sub.c.ad_request),
            _weighted_extra_ratio(sub, "above_the_fold_ratio", sub.c.impression),
        )
        .group_by(sub.c.report_date)
        .order_by(sub.c.report_date)
    ).all()

    by_income = db.execute(
        select(
            sub.c.income_type,
            func.sum(sub.c.net_revenue),
            func.sum(sub.c.impression),
        )
        .group_by(sub.c.income_type)
        .order_by(func.sum(sub.c.net_revenue).desc())
    ).all()

    by_unit = db.execute(
        select(
            sub.c.ad_unit,
            func.sum(sub.c.net_revenue),
            func.sum(sub.c.impression),
        )
        .group_by(sub.c.ad_unit)
        .order_by(func.sum(sub.c.net_revenue).desc())
        .limit(25)
    ).all()

    by_month = db.execute(
        select(
            sub.c.month_key,
            func.sum(sub.c.net_revenue),
            func.sum(sub.c.impression),
        )
        .group_by(sub.c.month_key)
        .order_by(sub.c.month_key)
    ).all()

    by_surface = db.execute(
        select(
            sub.c.surface,
            func.sum(sub.c.net_revenue),
            func.sum(sub.c.impression),
        )
        .group_by(sub.c.surface)
        .order_by(func.sum(sub.c.net_revenue).desc())
    ).all()

    by_channel = db.execute(
        select(
            sub.c.channel,
            func.sum(sub.c.net_revenue),
            func.sum(sub.c.impression),
        )
        .group_by(sub.c.channel)
        .order_by(func.sum(sub.c.net_revenue).desc())
    ).all()

    def _day_series(row: tuple) -> dict[str, Any]:
        d_rev = float(row[1] or 0)
        d_impr = float(row[2] or 0)
        d_clk = float(row[3] or 0)
        d_req = float(row[4] or 0)
        d_match = float(row[5] or 0)
        d_epv = float(row[6] or 0)
        d_euv = float(row[7] or 0)
        d_view_w = float(row[8] or 0)
        d_cov_w = float(row[9] or 0)
        d_atf_w = float(row[10] or 0)
        return {
            "date": row[0].isoformat(),
            "net_revenue": round(d_rev, 2),
            "impression": int(d_impr),
            "click": int(d_clk),
            "ad_request": int(d_req),
            "matched_request": int(d_match),
            "empower_pageview": int(d_epv),
            "empower_unique_visitor": int(d_euv),
            "ad_request_ecpm": round((d_rev / d_req * 1000.0) if d_req > 0 else 0.0, 3),
            "pageview_ecpm": round((d_rev / d_epv * 1000.0) if d_epv > 0 else 0.0, 3),
            "unique_visitor_ecpm": round((d_rev / d_euv * 1000.0) if d_euv > 0 else 0.0, 3),
            "ad_ecpm": round((d_rev / d_impr * 1000.0) if d_impr > 0 else 0.0, 3),
            "ctr_pct": round((d_clk / d_impr * 100.0) if d_impr > 0 else 0.0, 3),
            "coverage_pct": round((d_cov_w / d_req * 100.0) if d_req > 0 else 0.0, 3),
            "viewability_pct": round((d_view_w / d_impr * 100.0) if d_impr > 0 else 0.0, 3),
            "above_the_fold_ratio_pct": round((d_atf_w / d_impr * 100.0) if d_impr > 0 else 0.0, 3),
        }

    payload: dict[str, Any] = {
        "range": {"start": start, "end": end},
        "kpi_available": kpi_available,
        "kpis": {
            "empower_pageview": int(empower_pv),
            "empower_unique_visitor": int(empower_uv),
            "ad_request": int(ad_req),
            "matched_request": int(matched_req),
            "impression": int(impr),
            "click": int(clicks),
            "net_revenue": round(net_rev, 2),
            "ad_request_ecpm": round(req_ecpm, 3),
            "pageview_ecpm": round(pv_ecpm, 3),
            "unique_visitor_ecpm": round(uv_ecpm, 3),
            "ad_ecpm": round(avg_ecpm, 3),
            "viewability_pct": round(viewability_pct, 3),
            "above_the_fold_ratio_pct": round(atf_pct, 3),
            "ctr_pct": round(ctr, 3),
            "coverage_pct": round(coverage_pct, 3),
            # geriye uyumluluk
            "impressions": int(impr),
            "clicks": int(clicks),
            "ad_requests": int(ad_req),
            "avg_ad_ecpm": round(avg_ecpm, 3),
        },
        "by_date": [_day_series(r) for r in by_date],
        "by_income_type": [
            {
                "income_type": r[0],
                "net_revenue": round(float(r[1] or 0), 2),
                "impression": int(r[2] or 0),
            }
            for r in by_income
        ],
        "by_ad_unit": [
            {
                "ad_unit": r[0],
                "net_revenue": round(float(r[1] or 0), 2),
                "impression": int(r[2] or 0),
            }
            for r in by_unit
        ],
        "by_month": [
            {
                "month": r[0],
                "net_revenue": round(float(r[1] or 0), 2),
                "impression": int(r[2] or 0),
            }
            for r in by_month
        ],
        "by_surface": [
            {
                "surface": r[0],
                "net_revenue": round(float(r[1] or 0), 2),
                "impression": int(r[2] or 0),
            }
            for r in by_surface
        ],
        "by_channel": [
            {
                "channel": r[0],
                "net_revenue": round(float(r[1] or 0), 2),
                "impression": int(r[2] or 0),
            }
            for r in by_channel
        ],
    }

    cmp_start, cmp_end = resolve_compare_range(
        start, end, compare_mode, compare_start, compare_end
    )
    if cmp_start and cmp_end and compare_mode in COMPARE_MODES:
        compare_payload = query_summary(
            db,
            start=cmp_start,
            end=cmp_end,
            income_types=income_types,
            ad_units=ad_units,
            platforms=platforms,
            channels=channels,
            surfaces=surfaces,
            sources=sources,
            search=search,
            project=project,
            branch=branch,
            compare_mode=None,
        )
        payload["compare"] = _attach_compare_block(
            payload,
            compare_payload,
            range_start=cmp_start,
            range_end=cmp_end,
            mode=compare_mode,
        )

    return payload


def query_table(
    db: Session,
    *,
    start: str | None = None,
    end: str | None = None,
    income_types: str | None = None,
    ad_units: str | None = None,
    platforms: str | None = None,
    channels: str | None = None,
    surfaces: str | None = None,
    sources: str | None = None,
    search: str | None = None,
    project: str | None = None,
    branch: str | None = None,
    breakdown: str | "date,month,ad_unit,income_type",
    limit: int = 500,
    offset: int = 0,
    compare_mode: str | None = None,
    compare_start: str | None = None,
    compare_end: str | None = None,
) -> dict[str, Any]:
    it = _parse_filter_list(income_types)
    au = _parse_filter_list(ad_units)
    pl = _parse_filter_list(platforms)
    ch = _parse_filter_list(channels)
    su = _parse_filter_list(surfaces)
    sf = _parse_filter_list(sources)
    parts = [p.strip() for p in (breakdown or "").split(",") if p.strip()]
    allowed = {"date", "month", "ad_unit", "income_type", "platform", "channel", "surface"}
    group_cols = [p for p in parts if p in allowed] or ["date", "ad_unit", "income_type"]
    dim_fields = [c for c in group_cols]

    sub = select(AdReportRow)
    sub = _apply_filters(
        sub,
        start=start,
        end=end,
        income_types=it,
        ad_units=au,
        platforms=pl,
        channels=ch,
        surfaces=su,
        sources=sf,
        search=search,
        project=project,
        branch=branch,
    )
    sub = sub.subquery()

    select_cols = []
    group_by_cols = []
    if "date" in group_cols:
        select_cols.append(sub.c.report_date.label("report_date"))
        group_by_cols.append(sub.c.report_date)
    if "month" in group_cols:
        select_cols.append(sub.c.month_key.label("month_key"))
        group_by_cols.append(sub.c.month_key)
    if "ad_unit" in group_cols:
        select_cols.append(sub.c.ad_unit.label("ad_unit"))
        group_by_cols.append(sub.c.ad_unit)
    if "income_type" in group_cols:
        select_cols.append(sub.c.income_type.label("income_type"))
        group_by_cols.append(sub.c.income_type)
    if "platform" in group_cols:
        select_cols.append(sub.c.platform.label("platform"))
        group_by_cols.append(sub.c.platform)
    if "channel" in group_cols:
        select_cols.append(sub.c.channel.label("channel"))
        group_by_cols.append(sub.c.channel)
    if "surface" in group_cols:
        select_cols.append(sub.c.surface.label("surface"))
        group_by_cols.append(sub.c.surface)

    metrics = [
        func.sum(sub.c.ad_request).label("ad_request"),
        func.sum(sub.c.impression).label("impression"),
        func.sum(sub.c.click).label("click"),
        func.sum(sub.c.net_revenue).label("net_revenue"),
    ]
    q = select(*select_cols, *metrics).group_by(*group_by_cols)
    if "report_date" in [c.name for c in group_by_cols]:
        q = q.order_by(sub.c.report_date.desc())
    else:
        q = q.order_by(func.sum(sub.c.net_revenue).desc())

    grouped = q.subquery()
    total = int(db.scalar(select(func.count()).select_from(grouped)) or 0)
    rows = db.execute(
        select(grouped).limit(min(limit, 2000)).offset(max(offset, 0))
    ).all()

    out_rows = []
    for r in rows:
        m = r._mapping if hasattr(r, "_mapping") else dict(zip(grouped.c.keys(), r))
        item: dict[str, Any] = {}
        if "report_date" in m and m["report_date"] is not None:
            item["date"] = m["report_date"].isoformat()
        if "month_key" in m:
            item["month"] = m["month_key"]
        if "ad_unit" in m:
            item["ad_unit"] = m["ad_unit"]
        if "income_type" in m:
            item["income_type"] = m["income_type"]
        if "platform" in m:
            item["platform"] = m["platform"]
        if "channel" in m:
            item["channel"] = m["channel"]
        if "surface" in m:
            item["surface"] = m["surface"]
        impr = float(m["impression"] or 0)
        rev = float(m["net_revenue"] or 0)
        item.update({
            "ad_request": int(m["ad_request"] or 0),
            "impression": int(impr),
            "click": int(m["click"] or 0),
            "net_revenue": round(rev, 2),
            "computed_ecpm": round(rev / impr * 1000, 3) if impr > 0 else 0.0,
        })
        out_rows.append(item)

    result: dict[str, Any] = {
        "rows": out_rows,
        "total": int(total),
        "limit": limit,
        "offset": offset,
        "breakdown": group_cols,
    }

    cmp_start, cmp_end = resolve_compare_range(
        start, end, compare_mode, compare_start, compare_end
    )
    if cmp_start and cmp_end and compare_mode in COMPARE_MODES:
        cmp_table = query_table(
            db,
            start=cmp_start,
            end=cmp_end,
            income_types=income_types,
            ad_units=ad_units,
            platforms=platforms,
            channels=channels,
            surfaces=surfaces,
            sources=sources,
            search=search,
            project=project,
            branch=branch,
            breakdown=breakdown,
            limit=limit,
            offset=offset,
            compare_mode=None,
        )
        result["compare_range"] = {"start": cmp_start, "end": cmp_end}
        result["rows"] = _merge_table_rows(out_rows, cmp_table.get("rows") or [], dim_fields)

    return result


def _table_dimension_key(row: dict[str, Any], dim_fields: list[str]) -> str:
    parts: list[str] = []
    for field in dim_fields:
        if field == "date":
            parts.append(str(row.get("date") or ""))
        else:
            parts.append(str(row.get(field) or ""))
    return "|".join(parts)


def _merge_table_rows(
    primary_rows: list[dict[str, Any]],
    compare_rows: list[dict[str, Any]],
    dim_fields: list[str],
) -> list[dict[str, Any]]:
    numeric_metrics = (
        "ad_request",
        "impression",
        "click",
        "net_revenue",
        "computed_ecpm",
    )
    cmap = {_table_dimension_key(r, dim_fields): r for r in compare_rows}
    keys: list[str] = []
    seen: set[str] = set()
    for row in primary_rows + compare_rows:
        k = _table_dimension_key(row, dim_fields)
        if k in seen:
            continue
        seen.add(k)
        keys.append(k)

    merged: list[dict[str, Any]] = []
    for key in keys:
        prow = next((x for x in primary_rows if _table_dimension_key(x, dim_fields) == key), {})
        crow = cmap.get(key, {})
        item: dict[str, Any] = {}
        for field in dim_fields:
            if field == "date":
                item["date"] = prow.get("date") or crow.get("date")
            else:
                item[field] = prow.get(field) or crow.get(field)
        for metric in numeric_metrics:
            pv = float(prow.get(metric) or 0)
            cv = float(crow.get(metric) or 0)
            item[metric] = prow.get(metric, 0) if prow else 0
            item[f"{metric}_compare"] = crow.get(metric, 0) if crow else 0
            item[f"{metric}_delta_pct"] = _kpi_delta(pv, cv)["pct"]
        merged.append(item)
    merged.sort(key=lambda x: float(x.get("net_revenue") or 0), reverse=True)
    return merged
