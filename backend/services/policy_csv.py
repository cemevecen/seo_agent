"""Ad Manager Policy Center — scrape ingest + (legacy) CSV helpers.

Birincil kaynak: Mac bridge Playwright scrape
(`scripts/admanager_policy_scrape.py` → `/api/policy/ingest`).
CSV yükleme UI'dan kaldırıldı; parse_csv yalnızca scrape CSV indirmesi için kalır.
"""
from __future__ import annotations

import csv
import io
import json
import logging
import re
import threading
from datetime import date, datetime
from html import unescape
from typing import Any, Iterable
from urllib.parse import urlparse

import requests

logger = logging.getLogger(__name__)

ALLOWED_HOST_SUFFIXES = ("sinemalar.com",)

# ── CSV / scrape parse ────────────────────────────────────────────────────────

# Olası header isimleri → standart alan adı.
# Google Ad Manager Policy Center CSV'sinde sütun isimleri dilime/sürüme göre
# farklı olabilir — bu yüzden esnek eşleme yapıyoruz.
_HEADER_ALIASES: dict[str, list[str]] = {
    "url": [
        "url", "page url", "page", "sayfa", "site url", "site", "destination url",
        "landing page", "domain", "page_url",
        # Ad Manager TR export
        "sorunun konumu",
    ],
    "issue_type": [
        "violation type", "violation", "ihlal türü", "ihlal", "issue", "issue type",
        "policy", "policy violation", "policy issue", "reason", "neden",
        # Ad Manager TR export: "Sorunlar" = detaylı ihlal açıklaması
        "sorunlar", "sorun",
    ],
    "policy_topic": [
        # Ad Manager TR: "Sorun türü" = üst-kategori (Politika sorunu / Reklamveren tercihi / Yayıncı politikası)
        "sorun türü", "policy topic", "policy_topic",
    ],
    "enforcement": [
        "enforcement", "enforcement status", "status", "yaptırım", "uygulama",
        "action taken", "enforcement_status",
        # Ad Manager TR: "Durum" = "Kısıtlanmış reklam sunumu" vb.
        "durum",
    ],
    "ad_requests_7d": [
        "ad requests", "ad requests (7 days)", "ad_requests_7d", "ad requests 7d",
        "reklam istekleri", "reklam isteği", "weekly ad requests",
        "weekly_ad_request_count", "weeklyadrequestcount", "ad request count",
        "ad requests (last 7 days)", "ad requests (7d)", "requests",
        # Ad Manager TR export
        "reklam istekleri: son 7 gün", "reklam istekleri son 7 gün",
    ],
    "first_reported": [
        "first detected", "first reported", "first seen", "ilk tespit",
        "ilk bildirim", "first_detected_date", "detected on", "ilk_görülme",
        "ilk görülme",
        # Ad Manager TR: "Bildirim tarihi" = ilk tespit
        "bildirim tarihi",
    ],
    "last_reported": [
        "last detected", "last reported", "last seen", "son tespit",
        "son bildirim", "last_detected_date", "last updated", "son güncelleme",
        "son_görülme", "son görülme",
        # Ad Manager TR
        "son bulunma tarihi",
    ],
    "asset_type": [
        # Ad Manager TR: "Varlık" = Sayfa / Uygulama
        "varlık", "asset", "asset type",
    ],
    "property_codes": [
        # Ad Manager TR: "Mülk kodları" = ca-pub-XXX;ca-video-pub-XXX
        "mülk kodları", "property codes", "ad unit codes",
    ],
    "site_or_app": [
        "site veya uygulama", "site or app", "site/app",
    ],
}


def _norm(s: str) -> str:
    return (s or "").strip().lower().replace("﻿", "")


def _build_header_map(headers: list[str]) -> dict[str, int]:
    """{standart_alan: csv_kolon_index} eşlemesi üret."""
    norm_headers = [_norm(h) for h in headers]
    out: dict[str, int] = {}
    for std_key, aliases in _HEADER_ALIASES.items():
        for alias in aliases:
            alias_norm = _norm(alias)
            for i, h in enumerate(norm_headers):
                if h == alias_norm:
                    out[std_key] = i
                    break
            if std_key in out:
                break
    return out


def host_from_url(url: str) -> str:
    try:
        return (urlparse(url).hostname or "").lower()
    except Exception:
        return ""


def host_bucket(url: str) -> str:
    """m.sinemalar.com | sinemalar.com | other."""
    host = host_from_url(url)
    if host.startswith("m.") and host.endswith("sinemalar.com"):
        return "m.sinemalar.com"
    if host.endswith("sinemalar.com"):
        return "sinemalar.com"
    return "other"


def is_sinemalar_url(url: str) -> bool:
    host = host_from_url(url)
    return bool(host) and (
        host == "sinemalar.com" or host.endswith(".sinemalar.com")
    )


def _parse_int(v: Any) -> int:
    """'3300' / '3,3 B' / '3.3K' / '1,8 B' → int."""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return max(0, int(v))
    s = str(v).strip().replace("\u00a0", " ")
    if not s or s in ("—", "-", "–", "N/A"):
        return 0
    s_up = s.upper().replace(" ", "")
    mult = 1.0
    # TR: B = Bin, BİN; EN: K
    if s_up.endswith("BİN") or (s_up.endswith("B") and not s_up.endswith("MB")):
        mult = 1_000.0
        s_up = re.sub(r"(BİN|B)$", "", s_up)
    elif s_up.endswith("K"):
        mult = 1_000.0
        s_up = s_up[:-1]
    elif s_up.endswith("M") or s_up.endswith("MİLYON"):
        mult = 1_000_000.0
        s_up = re.sub(r"(MİLYON|M)$", "", s_up)
    # TR ondalık: 3,3 → 3.3 ; binlik: 1.234 → 1234
    if re.search(r"\d+\.\d{3}(\D|$)", s_up) and "," not in s_up:
        s_up = s_up.replace(".", "")
    else:
        s_up = s_up.replace(".", "").replace(",", ".") if "," in s_up else s_up.replace(",", "")
    m = re.search(r"[-+]?\d+(?:\.\d+)?", s_up)
    if not m:
        return 0
    try:
        return max(0, int(float(m.group(0)) * mult))
    except ValueError:
        return 0


_DATE_PATTERNS = [
    "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d.%m.%Y",
    "%m/%d/%Y", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S",
    "%b %d, %Y", "%d %b %Y",
    "%d %b %Y", "%d %B %Y",
]

_TR_MONTHS = {
    "oca": "Jan", "şub": "Feb", "sub": "Feb", "mar": "Mar", "nis": "Apr",
    "may": "May", "haz": "Jun", "tem": "Jul", "ağu": "Aug", "agu": "Aug",
    "eyl": "Sep", "eki": "Oct", "kas": "Nov", "ara": "Dec",
}


def _parse_date(v: Any) -> date | None:
    if not v:
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip()
    if not s or s.lower() in ("-", "n/a", "na", "null"):
        return None
    # TR: "3 Ağu 2026"
    m_tr = re.match(r"^(\d{1,2})\s+([A-Za-zÇĞİÖŞÜçğıöşü\.]+)\s+(\d{4})$", s)
    if m_tr:
        day, mon_raw, year = m_tr.group(1), m_tr.group(2).rstrip("."), m_tr.group(3)
        key = mon_raw.lower()[:3]
        # normalize turkish chars for key
        key = (
            key.replace("ğ", "g").replace("ü", "u").replace("ş", "s")
            .replace("ı", "i").replace("ö", "o").replace("ç", "c")
        )
        en = _TR_MONTHS.get(mon_raw.lower()[:3]) or _TR_MONTHS.get(key)
        if en:
            s = f"{day} {en} {year}"
    for fmt in _DATE_PATTERNS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).date()
    except ValueError:
        pass
    return None


def _categorize(issue_type: str) -> str:
    t = (issue_type or "").lower()
    if any(k in t for k in ("sexual", "adult", "cinsel", "yetişkin", "porn", "çıplak")):
        return "Yetişkinlere özel"
    if any(k in t for k in ("shocking", "şok", "violence", "şiddet", "graphic", "kanlı")):
        return "Şok edici içerik"
    if any(k in t for k in ("malware", "phishing", "güvenlik", "security", "harmful", "kötü amaçlı")):
        return "Güvenlik"
    if any(k in t for k in ("copyright", "telif", "trademark", "marka")):
        return "Telif/Marka"
    if any(k in t for k in ("yayıncı içeriği olmayan", "no content", "low value", "düşük değer")):
        return "Yayıncı içeriği yok"
    if any(k in t for k in ("dangerous", "tehlikeli", "weapons", "silah", "uyuşturucu", "drug")):
        return "Tehlikeli içerik"
    if any(k in t for k in ("misleading", "yanıltıcı", "deceptive", "aldatıcı")):
        return "Yanıltıcı içerik"
    if any(k in t for k in ("nefret", "hate", "ırk", "ayrımcılık", "discrimination")):
        return "Nefret/Ayrımcılık"
    if any(k in t for k in ("siyasi", "political", "seçim", "election")):
        return "Siyasi içerik"
    if any(k in t for k in ("hassas", "sensitive", "trajedi", "tragedy")):
        return "Hassas içerik"
    if any(k in t for k in ("müstehcen", "obscene")):
        return "Müstehcen içerik"
    return "Politika sorunu"


def parse_csv(content: bytes) -> tuple[list[dict], list[str], str | None]:
    """CSV'yi parse et.

    Döner: (rows, headers, error_message)
    """
    # Encoding dene
    text = None
    for enc in ("utf-8-sig", "utf-8", "cp1254", "iso-8859-9", "latin-1"):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        return [], [], "CSV dosyası okunamadı (encoding sorunu)."

    # Delimiter sniff
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        class _D:
            delimiter = ","
            quotechar = '"'
        dialect = _D  # type: ignore

    reader = csv.reader(io.StringIO(text), dialect)
    rows_raw = list(reader)
    if not rows_raw:
        return [], [], "CSV boş."

    headers = rows_raw[0]
    header_map = _build_header_map(headers)

    if "url" not in header_map:
        return [], headers, (
            f"CSV'de URL kolonu bulunamadı. Mevcut başlıklar: {headers}. "
            "URL içeren bir kolon olmalı (örn. 'URL', 'Page URL', 'Sayfa')."
        )
    if "issue_type" not in header_map:
        return [], headers, (
            f"CSV'de ihlal türü kolonu bulunamadı. Mevcut başlıklar: {headers}. "
            "İhlal/violation içeren bir kolon olmalı."
        )

    rows: list[dict] = []
    skipped_columns = set(range(len(headers))) - set(header_map.values())

    for raw in rows_raw[1:]:
        if not raw or not any(raw):
            continue

        def get(key: str) -> str:
            i = header_map.get(key)
            if i is None or i >= len(raw):
                return ""
            return (raw[i] or "").strip()

        url = get("url")
        issue_type = get("issue_type")
        site_or_app = get("site_or_app")
        if not url or not issue_type:
            continue

        # URL eğer http(s) yoksa ekle — Ad Manager "m.sinemalar.com/..." veya path verir
        if not url.startswith(("http://", "https://")):
            if site_or_app and "sinemalar" in site_or_app.lower():
                host = site_or_app.strip().split()[0].rstrip("/")
                if not host.startswith("http"):
                    url = f"https://{host}/{url.lstrip('./')}"
                else:
                    url = f"{host.rstrip('/')}/{url.lstrip('./')}"
            else:
                url = "https://" + url.lstrip("/")

        if not is_sinemalar_url(url):
            continue

        # Bilinmeyen + bilinen ek bilgiler extras'a
        extras: dict[str, str] = {}
        for i in skipped_columns:
            if i < len(raw) and raw[i]:
                col_name = headers[i] if i < len(headers) else f"col_{i}"
                extras[col_name] = raw[i].strip()
        # Bilinen ama ana alanlarda yer almayan bilgiler
        for key in ("policy_topic", "asset_type", "property_codes", "site_or_app"):
            v = get(key)
            if v:
                extras[key] = v
        extras["host_bucket"] = host_bucket(url)
        extras["source"] = extras.get("source") or "admanager_csv"

        # Category her zaman issue_type'tan otomatik üretilir (chip renkleri buna bağlı)
        category = _categorize(issue_type)

        rows.append({
            "url": url,
            "issue_type": issue_type,
            "category": category,
            "enforcement": get("enforcement"),
            "ad_requests_7d": _parse_int(get("ad_requests_7d")),
            "first_reported": _parse_date(get("first_reported")),
            "last_reported": _parse_date(get("last_reported")),
            "extra_json": json.dumps(extras, ensure_ascii=False) if extras else "",
        })

    return rows, headers, None


def rows_from_scrape_payload(payload: dict[str, Any]) -> list[dict]:
    """Bridge ingest satırlarını import_rows formatına çevir (yalnız sinemalar)."""
    raw_rows = payload.get("rows") if isinstance(payload, dict) else None
    if not isinstance(raw_rows, list):
        raw_rows = []
    out: list[dict] = []
    for item in raw_rows:
        if not isinstance(item, dict):
            continue
        url = str(item.get("url") or "").strip()
        issue_type = str(item.get("issue_type") or "").strip()
        if not url or not issue_type:
            continue
        if not url.startswith(("http://", "https://")):
            url = "https://" + url.lstrip("/")
        if not is_sinemalar_url(url):
            continue
        extras: dict[str, Any] = {}
        if item.get("site_host"):
            extras["site_host"] = item["site_host"]
        extras["host_bucket"] = host_bucket(url)
        extras["source"] = str(item.get("source") or payload.get("source") or "admanager_policy_scrape")
        out.append(
            {
                "url": url,
                "issue_type": issue_type,
                "category": _categorize(issue_type),
                "enforcement": str(item.get("enforcement") or "").strip(),
                "ad_requests_7d": _parse_int(item.get("ad_requests_7d")),
                "first_reported": _parse_date(item.get("first_reported")),
                "last_reported": _parse_date(item.get("last_reported") or item.get("first_reported")),
                "extra_json": json.dumps(extras, ensure_ascii=False),
            }
        )
    return out


def ingest_scrape_payload(db, payload: dict[str, Any]) -> dict[str, Any]:
    """Scrape sonucu UPSERT + blob kaydı + non-sinemalar temizliği."""
    import base64

    rows = rows_from_scrape_payload(payload)
    csv_b64 = payload.get("csv_base64") if isinstance(payload, dict) else None
    if not rows and csv_b64:
        try:
            content = base64.b64decode(csv_b64)
            parsed, _h, err = parse_csv(content)
            if err:
                return {"ok": False, "message": err}
            rows = parsed
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "message": f"csv_base64 decode: {exc}"}

    if not rows:
        return {"ok": False, "message": "Sinemalar satırı yok — filtre veya tarama kaynağını kontrol et"}

    new_count, upd_count = import_rows(db, rows)
    pruned = prune_non_sinemalar(db)

    filename = f"admanager-policy-{datetime.utcnow().strftime('%Y%m%d-%H%M')}.json"
    blob = json.dumps(
        {
            "source": payload.get("source"),
            "scraped_at": payload.get("scraped_at"),
            "method": payload.get("method"),
            "site_filter": payload.get("site_filter"),
            "row_count": len(rows),
        },
        ensure_ascii=False,
    ).encode("utf-8")
    if csv_b64:
        try:
            blob = base64.b64decode(csv_b64)
            filename = filename.replace(".json", ".csv")
        except Exception:
            pass
    save_csv_blob(
        db,
        filename=filename,
        content=blob,
        row_count=len(rows),
        new_count=new_count,
        updated_count=upd_count,
    )
    return {
        "ok": True,
        "imported": len(rows),
        "new_count": new_count,
        "updated_count": upd_count,
        "pruned_non_sinemalar": pruned,
        "message": f"{len(rows)} satır · {new_count} yeni · {upd_count} güncellendi",
    }


def prune_non_sinemalar(db) -> int:
    """Yalnız sinemalar.com / m.sinemalar.com kalsın."""
    from backend.models import AdPolicyViolation

    deleted = 0
    for row in db.query(AdPolicyViolation).all():
        if not is_sinemalar_url(row.url or ""):
            db.delete(row)
            deleted += 1
    if deleted:
        db.commit()
    return deleted


# ── DB import (UPSERT, duplicate koruması) ────────────────────────────────────

def import_rows(db, rows: list[dict]) -> tuple[int, int]:
    """Satırları DB'ye UPSERT et.

    Aynı (url, issue_type) çiftine sahip satır varsa günceller —
    our_status, our_notes, page_title, page_title_fetched_at korunur.

    Döner: (yeni_eklenen, güncellenen)
    """
    from backend.models import AdPolicyViolation

    if not rows:
        return 0, 0

    now = datetime.utcnow()
    new_count = 0
    upd_count = 0

    for r in rows:
        if not is_sinemalar_url(r.get("url") or ""):
            continue
        existing = (
            db.query(AdPolicyViolation)
            .filter(
                AdPolicyViolation.url == r["url"],
                AdPolicyViolation.issue_type == r["issue_type"],
            )
            .first()
        )
        if existing:
            existing.category = r["category"] or existing.category
            existing.enforcement = r["enforcement"] or existing.enforcement
            existing.ad_requests_7d = r["ad_requests_7d"]
            if r["first_reported"]:
                if not existing.first_reported or r["first_reported"] < existing.first_reported:
                    existing.first_reported = r["first_reported"]
            existing.last_reported = r["last_reported"] or date.today()
            existing.extra_json = r["extra_json"] or existing.extra_json
            existing.fetched_at = now
            # our_status, our_notes, page_title, page_title_fetched_at KORUNUR
            upd_count += 1
        else:
            db.add(AdPolicyViolation(
                url=r["url"],
                issue_type=r["issue_type"],
                category=r["category"],
                enforcement=r["enforcement"],
                ad_requests_7d=r["ad_requests_7d"],
                first_reported=r["first_reported"] or date.today(),
                last_reported=r["last_reported"] or date.today(),
                page_title="",
                page_title_fetched_at=None,
                extra_json=r["extra_json"],
                our_status="new",
                our_notes="",
                fetched_at=now,
                updated_at=now,
                first_seen_at=now,
            ))
            new_count += 1

    db.commit()
    return new_count, upd_count


def save_csv_blob(db, filename: str, content: bytes, row_count: int,
                  new_count: int, updated_count: int) -> int:
    """Yüklenen CSV'yi DB'de sakla (sadece son 5 tanesini tut)."""
    from backend.models import PolicyCSVUpload

    upload = PolicyCSVUpload(
        filename=filename,
        row_count=row_count,
        new_count=new_count,
        updated_count=updated_count,
        content=content,
        uploaded_at=datetime.utcnow(),
    )
    db.add(upload)
    db.flush()
    upload_id = upload.id

    # Eski yüklemeleri temizle (son 5'i tut)
    old = (
        db.query(PolicyCSVUpload)
        .order_by(PolicyCSVUpload.uploaded_at.desc())
        .offset(5)
        .all()
    )
    for o in old:
        db.delete(o)
    db.commit()
    return upload_id


def get_latest_upload(db):
    from backend.models import PolicyCSVUpload
    return (
        db.query(PolicyCSVUpload)
        .order_by(PolicyCSVUpload.uploaded_at.desc())
        .first()
    )


# ── Sorgular ──────────────────────────────────────────────────────────────────

def get_violations(db, *, status: str | None = None, category: str | None = None,
                   order_by: str = "ad_requests", limit: int = 5000,
                   new_threshold: datetime | None = None,
                   host: str | None = None) -> list[dict]:
    """`host`: all | sinemalar.com | m.sinemalar.com"""
    from backend.models import AdPolicyViolation
    from sqlalchemy import desc

    q = db.query(AdPolicyViolation)
    if status and status != "all":
        q = q.filter(AdPolicyViolation.our_status == status)
    if category and category != "all":
        q = q.filter(AdPolicyViolation.category == category)

    if order_by == "date":
        q = q.order_by(desc(AdPolicyViolation.last_reported), desc(AdPolicyViolation.ad_requests_7d))
    elif order_by == "url":
        q = q.order_by(AdPolicyViolation.url)
    else:
        q = q.order_by(desc(AdPolicyViolation.ad_requests_7d))

    host_filter = (host or "all").strip().lower()
    results = []
    for r in q.limit(limit).all():
        if not is_sinemalar_url(r.url or ""):
            continue
        bucket = host_bucket(r.url or "")
        if host_filter in ("sinemalar.com", "m.sinemalar.com") and bucket != host_filter:
            continue
        d = _violation_to_dict(r)
        d["is_new_import"] = bool(
            new_threshold and r.first_seen_at and r.first_seen_at >= new_threshold
        )
        results.append(d)
    return results


def get_stats(db) -> dict:
    from backend.models import AdPolicyViolation
    from sqlalchemy import func

    total = db.query(func.count(AdPolicyViolation.id)).scalar() or 0
    new_count = db.query(func.count(AdPolicyViolation.id)).filter(
        AdPolicyViolation.our_status == "new"
    ).scalar() or 0
    total_requests = db.query(func.sum(AdPolicyViolation.ad_requests_7d)).scalar() or 0
    last_fetch = db.query(func.max(AdPolicyViolation.fetched_at)).scalar()
    try:
        with_title = db.query(func.count(AdPolicyViolation.id)).filter(
            AdPolicyViolation.page_title != ""
        ).scalar() or 0
    except Exception:
        db.rollback()
        with_title = 0

    by_category: dict[str, int] = {}
    for row in db.query(AdPolicyViolation.category, func.count(AdPolicyViolation.id)).group_by(
        AdPolicyViolation.category
    ).order_by(func.count(AdPolicyViolation.id).desc()).all():
        by_category[row[0] or "Diğer"] = row[1]

    by_status: dict[str, int] = {}
    for row in db.query(AdPolicyViolation.our_status, func.count(AdPolicyViolation.id)).group_by(
        AdPolicyViolation.our_status
    ).all():
        by_status[row[0]] = row[1]

    by_host = {"sinemalar.com": 0, "m.sinemalar.com": 0, "other": 0}
    for row in db.query(AdPolicyViolation.url).all():
        by_host[host_bucket(row[0] or "")] = by_host.get(host_bucket(row[0] or ""), 0) + 1

    return {
        "total": total,
        "new": new_count,
        "with_title": with_title,
        "without_title": total - with_title,
        "total_ad_requests_7d": int(total_requests),
        "last_fetch": last_fetch.isoformat() if last_fetch else None,
        "by_category": by_category,
        "by_status": by_status,
        "by_host": by_host,
    }


def _admin_link(url: str) -> str | None:
    # Person pages: mobileweb/person/, mobileweb/personMovies/, or /sanatci/.../id
    m_person = re.search(r"mobileweb/person(?:Movies)?/(\d+)", url)
    if m_person:
        return f"https://www.sinemalar.com/management/person/{m_person.group(1)}"
    m_person2 = re.search(r"/sanatci/[^/]+/(\d+)", url)
    if m_person2:
        return f"https://www.sinemalar.com/management/person/{m_person2.group(1)}"

    # Movie pages: mobileweb/movieInfo/ or /film/.../id
    m_movie = re.search(r"mobileweb/movieInfo/(\d+)", url)
    if m_movie:
        return f"https://www.sinemalar.com/management/movie/{m_movie.group(1)}"
    m_movie2 = re.search(r"/film/[^/]+/(\d+)", url)
    if m_movie2:
        return f"https://www.sinemalar.com/management/movie/{m_movie2.group(1)}"

    return None


def _violation_to_dict(r) -> dict:
    extras = {}
    if r.extra_json:
        try:
            extras = json.loads(r.extra_json)
        except (json.JSONDecodeError, TypeError):
            extras = {}
    return {
        "id": r.id,
        "url": r.url,
        "host": host_from_url(r.url or ""),
        "host_bucket": host_bucket(r.url or ""),
        "page_title": r.page_title or "",
        "page_title_fetched_at": r.page_title_fetched_at.isoformat() if r.page_title_fetched_at else None,
        "issue_type": r.issue_type,
        "category": r.category,
        "ad_requests_7d": r.ad_requests_7d,
        "enforcement": r.enforcement,
        "first_reported": r.first_reported.isoformat() if r.first_reported else None,
        "last_reported": r.last_reported.isoformat() if r.last_reported else None,
        "our_status": r.our_status,
        "our_notes": r.our_notes,
        "updated_at": r.updated_at.isoformat() if r.updated_at else None,
        "fetched_at": r.fetched_at.isoformat() if r.fetched_at else None,
        "first_seen_at": r.first_seen_at.isoformat() if r.first_seen_at else None,
        "admin_link": _admin_link(r.url),
        "in_noads": r.in_noads,
        "noads_checked_at": r.noads_checked_at.isoformat() if getattr(r, "noads_checked_at", None) else None,
        "extras": extras,
    }


# ── Sayfa başlığı çekme ───────────────────────────────────────────────────────

_TITLE_RE = re.compile(r"<title[^>]*>(.*?)</title>", re.IGNORECASE | re.DOTALL)
_OG_TITLE_RE = re.compile(
    r"""<meta\s+[^>]*property\s*=\s*['"]og:title['"]\s+[^>]*content\s*=\s*['"]([^'"]+)['"]""",
    re.IGNORECASE,
)
_TWITTER_TITLE_RE = re.compile(
    r"""<meta\s+[^>]*name\s*=\s*['"]twitter:title['"]\s+[^>]*content\s*=\s*['"]([^'"]+)['"]""",
    re.IGNORECASE,
)

_USER_AGENT = (
    "Mozilla/5.0 (compatible; SeoAgentPolicyBot/1.0; "
    "+https://www.sinemalar.com/admin)"
)


def _clean_title(t: str) -> str:
    t = unescape(t)
    t = re.sub(r"\s+", " ", t).strip()
    return t[:480]


def fetch_title(url: str, timeout: float = 10.0) -> str | None:
    """URL'den HTML <title> ya da og:title çek."""
    if not url or not url.startswith(("http://", "https://")):
        return None
    try:
        resp = requests.get(
            url,
            headers={"User-Agent": _USER_AGENT, "Accept-Language": "tr-TR,tr;q=0.9,en;q=0.5"},
            timeout=timeout,
            allow_redirects=True,
        )
        if resp.status_code >= 400:
            return f"[HTTP {resp.status_code}]"
        # İlk 100 KB yeter
        html = resp.text[:100_000]

        m = _TITLE_RE.search(html)
        if m and m.group(1).strip():
            return _clean_title(m.group(1))

        m = _OG_TITLE_RE.search(html)
        if m:
            return _clean_title(m.group(1))

        m = _TWITTER_TITLE_RE.search(html)
        if m:
            return _clean_title(m.group(1))

        return "[başlık yok]"
    except requests.Timeout:
        return "[timeout]"
    except requests.RequestException as exc:
        return f"[hata: {str(exc)[:80]}]"


# ── Batch title fetch job (in-process state) ──────────────────────────────────

_TITLE_JOB_STATE: dict[str, Any] = {
    "running": False,
    "total": 0,
    "done": 0,
    "started_at": None,
    "finished_at": None,
    "error": None,
}
_TITLE_JOB_LOCK = threading.Lock()


def get_title_job_state() -> dict:
    with _TITLE_JOB_LOCK:
        s = dict(_TITLE_JOB_STATE)
    if s["started_at"]:
        s["started_at"] = s["started_at"].isoformat() if isinstance(s["started_at"], datetime) else s["started_at"]
    if s["finished_at"]:
        s["finished_at"] = s["finished_at"].isoformat() if isinstance(s["finished_at"], datetime) else s["finished_at"]
    return s


def start_title_job(session_factory, *, only_missing: bool = True) -> bool:
    """Sayfa başlıklarını arka planda çek. True döner: başladı; False: zaten çalışıyor."""
    with _TITLE_JOB_LOCK:
        if _TITLE_JOB_STATE["running"]:
            return False
        _TITLE_JOB_STATE["running"] = True
        _TITLE_JOB_STATE["done"] = 0
        _TITLE_JOB_STATE["total"] = 0
        _TITLE_JOB_STATE["started_at"] = datetime.utcnow()
        _TITLE_JOB_STATE["finished_at"] = None
        _TITLE_JOB_STATE["error"] = None

    def _worker():
        from concurrent.futures import ThreadPoolExecutor, as_completed
        from backend.models import AdPolicyViolation
        db = session_factory()
        try:
            q = db.query(AdPolicyViolation)
            if only_missing:
                q = q.filter(AdPolicyViolation.page_title == "")
            urls = [u for (u,) in q.with_entities(AdPolicyViolation.url).distinct().all()]
            with _TITLE_JOB_LOCK:
                _TITLE_JOB_STATE["total"] = len(urls)

            if not urls:
                return

            # 20 paralel worker — 558 URL ~30-60s sürer
            results: dict[str, str] = {}
            with ThreadPoolExecutor(max_workers=20, thread_name_prefix="title-fetch") as ex:
                futures = {ex.submit(fetch_title, u): u for u in urls}
                for fut in as_completed(futures):
                    url = futures[fut]
                    try:
                        title = fut.result() or "[başlık yok]"
                    except Exception as fe:
                        title = f"[hata: {str(fe)[:60]}]"
                    results[url] = title
                    with _TITLE_JOB_LOCK:
                        _TITLE_JOB_STATE["done"] += 1

                    # Her 25 satırda bir commit (DB trafiğini azalt)
                    if len(results) % 25 == 0:
                        _flush_titles(db, results)
                        results = {}

            # Kalanları yaz
            if results:
                _flush_titles(db, results)
        except Exception as exc:
            logger.exception("Title fetch job hatası")
            with _TITLE_JOB_LOCK:
                _TITLE_JOB_STATE["error"] = str(exc)[:300]
        finally:
            db.close()
            with _TITLE_JOB_LOCK:
                _TITLE_JOB_STATE["running"] = False
                _TITLE_JOB_STATE["finished_at"] = datetime.utcnow()

    threading.Thread(target=_worker, daemon=True, name="policy-title-fetch").start()
    return True


def _flush_titles(db, url_to_title: dict[str, str]) -> None:
    """Toplu update — her URL için page_title alanını yaz."""
    from backend.models import AdPolicyViolation
    if not url_to_title:
        return
    now = datetime.utcnow()
    try:
        for url, title in url_to_title.items():
            db.query(AdPolicyViolation).filter(
                AdPolicyViolation.url == url
            ).update({
                AdPolicyViolation.page_title: title,
                AdPolicyViolation.page_title_fetched_at: now,
            }, synchronize_session=False)
        db.commit()
    except Exception:
        db.rollback()
        raise


def refresh_single_title(db, vid: int) -> str | None:
    """Tek bir satırın sayfa başlığını yeniden çek."""
    from backend.models import AdPolicyViolation
    row = db.query(AdPolicyViolation).filter(AdPolicyViolation.id == vid).first()
    if not row:
        return None
    title = fetch_title(row.url) or "[başlık yok]"
    now = datetime.utcnow()
    # Aynı URL'ye sahip tüm satırları güncelle
    db.query(AdPolicyViolation).filter(
        AdPolicyViolation.url == row.url
    ).update({
        AdPolicyViolation.page_title: title,
        AdPolicyViolation.page_title_fetched_at: now,
    }, synchronize_session=False)
    db.commit()
    return title


# ── Excel export ──────────────────────────────────────────────────────────────

def build_xlsx(violations: list[dict]) -> bytes:
    """Violations listesini .xlsx olarak serialize et."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Policy İhlalleri"

    headers = [
        "URL", "Sayfa Başlığı", "İhlal Türü", "Kategori",
        "Yaptırım", "Reklam İsteği (7g)", "İlk Tespit", "Son Tespit",
        "Admin Link", "Güncellendi",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_i, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_i)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for v in violations:
        ws.append([
            v.get("url", ""),
            v.get("page_title", ""),
            v.get("issue_type", ""),
            v.get("category", ""),
            v.get("enforcement", ""),
            v.get("ad_requests_7d", 0),
            v.get("first_reported") or "",
            v.get("last_reported") or "",
            v.get("admin_link") or "",
            (v.get("updated_at") or "")[:19],
        ])

    widths = [55, 50, 40, 22, 18, 14, 12, 12, 60, 19]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
